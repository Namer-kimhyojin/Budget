from django.test import TestCase
from django.test.utils import override_settings
from django.core.management import call_command
from django.contrib.auth.models import User
from rest_framework.test import APIClient
from openpyxl import Workbook, load_workbook
from io import BytesIO, StringIO
from pathlib import Path
import json
import tempfile
from unittest import mock

from .calculation import parse_calc_expression
from .models import UserProfile, Organization, BudgetSubject, BudgetEntry, BudgetDetail, BudgetVersion, EntrustedProject, ApprovalLog
from .services.budget_book_export import build_budget_book_file


class CalcExpressionParserTest(TestCase):
    def test_parse_simple_expression(self):
        parsed = parse_calc_expression('50000x3x12')
        self.assertEqual(parsed.price, 50000)
        self.assertEqual(parsed.qty, 3.0)
        self.assertEqual(parsed.freq, 12)
        self.assertEqual(parsed.amount, 1800000)

    def test_parse_expression_with_commas_and_symbols(self):
        parsed = parse_calc_expression('50,000 * 2 * 12')
        self.assertEqual(parsed.price, 50000)
        self.assertEqual(parsed.qty, 2.0)
        self.assertEqual(parsed.freq, 12)
        self.assertEqual(parsed.amount, 1200000)

    def test_parse_expression_with_korean_unit(self):
        parsed = parse_calc_expression('5만원 x 3 x 12')
        self.assertEqual(parsed.price, 50000)
        self.assertEqual(parsed.qty, 3.0)
        self.assertEqual(parsed.freq, 12)

    def test_invalid_expression_returns_error(self):
        with self.assertRaises(ValueError):
            parse_calc_expression('50000x3')


class CalcExpressionApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_parse_expression_action_success(self):
        response = self.client.post('/api/details/parse_expression/', {'expression': '40000x2x10'}, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['price'], 40000)
        self.assertEqual(response.data['qty'], 2.0)
        self.assertEqual(response.data['freq'], 10)
        self.assertEqual(response.data['amount'], 800000)

    def test_parse_expression_action_failure(self):
        response = self.client.post('/api/details/parse_expression/', {'expression': 'invalid text'}, format='json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.data)

    def test_detail_accepts_region_weather_and_evidence_fields(self):
        org = Organization.objects.create(name='테스트부서', code='T001', org_type='dept')
        subject = BudgetSubject.objects.create(code='9999', name='테스트목', level=4, subject_type='expense')
        entry = BudgetEntry.objects.create(subject=subject, organization=org, year=2026, supplemental_round=0)
        response = self.client.post('/api/details/', {
            'entry': entry.id,
            'name': '현장 출장',
            'price': 1000,
            'qty': 1,
            'freq': 1,
            'unit': '회',
            'source': '자체',
            'region_context': '해안 지역',
            'weather_context': '강풍 빈발',
            'evidence_source_name': '기상청',
            'evidence_source_url': 'https://www.weather.go.kr/',
        }, format='json')
        self.assertEqual(response.status_code, 201)


class AuthApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.dept = Organization.objects.create(name='Auth Dept', code='AUTHD', org_type='dept')
        self.team = Organization.objects.create(name='Auth Team', code='AUTHT', org_type='team', parent=self.dept)
        self.other_dept = Organization.objects.create(name='Other Dept', code='AUTHD2', org_type='dept')
        self.other_team = Organization.objects.create(name='Other Team', code='AUTHT2', org_type='team', parent=self.other_dept)

    def test_signup_and_me_flow(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'tester01',
            'password': 'StrongPass!234',
            'name': '테스터',
            'email': 'tester01@example.com',
        }, format='json')
        self.assertEqual(signup.status_code, 201)
        token = signup.data['token']
        self.assertTrue(token)

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')
        me = self.client.get('/api/auth/me/')
        self.assertEqual(me.status_code, 200)
        self.assertEqual(me.data['user']['username'], 'tester01')
        self.assertIn(me.data['profile']['role'], ['ADMIN', 'REQUESTOR'])

    def test_signup_with_team_sets_department_and_team(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'tester_team',
            'password': 'StrongPass!234',
            'name': 'Team User',
            'email': 'tester_team@example.com',
            'team': self.team.id,
        }, format='json')
        self.assertEqual(signup.status_code, 201)
        self.assertEqual(signup.data['profile']['organization'], self.dept.id)
        self.assertEqual(signup.data['profile']['team'], self.team.id)

    def test_login_and_logout(self):
        self.client.post('/api/auth/signup/', {
            'username': 'tester02',
            'password': 'StrongPass!234',
            'name': '테스터2',
            'email': 'tester02@example.com',
        }, format='json')

        login = self.client.post('/api/auth/login/', {
            'username': 'tester02',
            'password': 'StrongPass!234',
        }, format='json')
        self.assertEqual(login.status_code, 200)
        token = login.data['token']

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')
        logout = self.client.post('/api/auth/logout/', {}, format='json')
        self.assertEqual(logout.status_code, 200)

    def test_find_id(self):
        self.client.post('/api/auth/signup/', {
            'username': 'findme',
            'password': 'StrongPass!234',
            'name': '찾기유저',
            'email': 'find@example.com',
        }, format='json')
        response = self.client.post('/api/auth/find-id/', {
            'name': '찾기유저',
            'email': 'find@example.com',
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.data['count'], 1)

    def test_assign_role_by_admin(self):
        admin_signup = self.client.post('/api/auth/signup/', {
            'username': 'adminuser',
            'password': 'StrongPass!234',
            'name': '관리자',
            'email': 'admin@example.com',
        }, format='json')
        admin_token = admin_signup.data['token']

        user_signup = self.client.post('/api/auth/signup/', {
            'username': 'normaluser',
            'password': 'StrongPass!234',
            'name': '일반',
            'email': 'user@example.com',
        }, format='json')
        user_id = user_signup.data['user']['id']

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {admin_token}')
        assign = self.client.post('/api/auth/assign-role/', {
            'user_id': user_id,
            'role': 'MANAGER',
            'organization': self.dept.id,
            'team': self.team.id,
        }, format='json')
        self.assertEqual(assign.status_code, 200)
        profile = UserProfile.objects.get(user_id=user_id)
        self.assertEqual(profile.role, 'MANAGER')
        self.assertEqual(profile.organization_id, self.dept.id)
        self.assertEqual(profile.team_id, self.team.id)

    def test_login_with_email(self):
        self.client.post('/api/auth/signup/', {
            'username': 'mailuser',
            'password': 'StrongPass!234',
            'name': 'Mail User',
            'email': 'mailuser@example.com',
        }, format='json')
        response = self.client.post('/api/auth/login/', {
            'username': 'mailuser@example.com',
            'password': 'StrongPass!234',
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertIn('token', response.data)

    def test_change_password(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'pwuser',
            'password': 'StrongPass!234',
            'name': 'PW User',
            'email': 'pwuser@example.com',
        }, format='json')
        token = signup.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')
        changed = self.client.post('/api/auth/change-password/', {
            'current_password': 'StrongPass!234',
            'new_password': 'NewStrong!789',
        }, format='json')
        self.assertEqual(changed.status_code, 200)

        self.client.credentials()
        login_new = self.client.post('/api/auth/login/', {
            'username': 'pwuser',
            'password': 'NewStrong!789',
        }, format='json')
        self.assertEqual(login_new.status_code, 200)

    def test_password_policy_endpoint(self):
        response = self.client.get('/api/auth/password-policy/')
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data.get('lock_policy'))

    def test_admin_user_management_endpoints(self):
        admin = self.client.post('/api/auth/signup/', {
            'username': 'admin_mng',
            'password': 'StrongPass!234',
            'name': 'AdminMng',
            'email': 'admin_mng@example.com',
        }, format='json')
        admin_token = admin.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {admin_token}')

        create_user = self.client.post('/api/auth/users/', {
            'username': 'managed01',
            'password': 'StrongPass!234',
            'name': 'Managed User',
            'email': 'managed01@example.com',
            'role': 'STAFF',
            'organization': self.dept.id,
            'team': self.team.id,
        }, format='json')
        self.assertEqual(create_user.status_code, 201)

        users = self.client.get('/api/auth/users/')
        self.assertEqual(users.status_code, 200)
        managed = next((u for u in users.data if u['username'] == 'managed01'), None)
        self.assertIsNotNone(managed)

        patch = self.client.patch(f"/api/auth/users/{managed['id']}/", {
            'name': 'Managed Updated',
            'role': 'MANAGER',
            'organization': self.other_dept.id,
            'team': self.other_team.id,
            'is_active': True,
        }, format='json')
        self.assertEqual(patch.status_code, 200)

        delete = self.client.delete(f"/api/auth/users/{managed['id']}/")
        self.assertEqual(delete.status_code, 200)

    def test_admin_user_management_rejects_team_department_mismatch(self):
        admin = self.client.post('/api/auth/signup/', {
            'username': 'admin_mismatch',
            'password': 'StrongPass!234',
            'name': 'AdminMismatch',
            'email': 'admin_mismatch@example.com',
        }, format='json')
        admin_token = admin.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {admin_token}')

        create_user = self.client.post('/api/auth/users/', {
            'username': 'managed02',
            'password': 'StrongPass!234',
            'name': 'Managed User2',
            'email': 'managed02@example.com',
            'role': 'STAFF',
            'organization': self.dept.id,
            'team': self.team.id,
        }, format='json')
        self.assertEqual(create_user.status_code, 201)
        user_id = create_user.data['user']['id']

        mismatch = self.client.patch(f"/api/auth/users/{user_id}/", {
            'organization': self.dept.id,
            'team': self.other_team.id,
        }, format='json')
        self.assertEqual(mismatch.status_code, 400)
        self.assertEqual(mismatch.data['error'], 'team does not belong to organization')


class AuditLogApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        admin_signup = self.client.post('/api/auth/signup/', {
            'username': 'audit_admin',
            'password': 'StrongPass!234',
            'name': 'Audit Admin',
            'email': 'audit_admin@example.com',
        }, format='json')
        self.assertEqual(admin_signup.status_code, 201)
        self.admin_token = admin_signup.data['token']

    def test_login_and_logout_are_recorded(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'audit_user',
            'password': 'StrongPass!234',
            'name': 'Audit User',
            'email': 'audit_user@example.com',
        }, format='json')
        self.assertEqual(signup.status_code, 201)
        user_id = signup.data['user']['id']

        login = self.client.post('/api/auth/login/', {
            'username': 'audit_user',
            'password': 'StrongPass!234',
        }, format='json')
        self.assertEqual(login.status_code, 200)
        user_token = login.data['token']

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {user_token}')
        logout = self.client.post('/api/auth/logout/', {}, format='json')
        self.assertEqual(logout.status_code, 200)

        self.assertTrue(
            ApprovalLog.objects.filter(
                log_type='AUTH',
                action='LOGIN',
                resource_type='auth',
                resource_id=str(user_id),
            ).exists()
        )
        self.assertTrue(
            ApprovalLog.objects.filter(
                log_type='AUTH',
                action='LOGOUT',
                resource_type='auth',
                resource_id=str(user_id),
            ).exists()
        )

    def test_admin_user_crud_is_recorded(self):
        dept = Organization.objects.create(name='Audit Dept', code='AUDIT_D', org_type='dept')

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {self.admin_token}')
        create_user = self.client.post('/api/auth/users/', {
            'username': 'managed_audit',
            'password': 'StrongPass!234',
            'name': 'Managed Audit',
            'email': 'managed_audit@example.com',
            'role': 'STAFF',
            'organization': dept.id,
        }, format='json')
        self.assertEqual(create_user.status_code, 201)
        target_user_id = create_user.data['user']['id']

        patch = self.client.patch(f'/api/auth/users/{target_user_id}/', {
            'name': 'Managed Audit Updated',
            'is_active': True,
        }, format='json')
        self.assertEqual(patch.status_code, 200)

        delete = self.client.delete(f'/api/auth/users/{target_user_id}/')
        self.assertEqual(delete.status_code, 200)

        self.assertTrue(ApprovalLog.objects.filter(log_type='CRUD', action='CREATE', resource_type='user', resource_id=str(target_user_id)).exists())
        self.assertTrue(ApprovalLog.objects.filter(log_type='CRUD', action='UPDATE', resource_type='user', resource_id=str(target_user_id)).exists())
        self.assertTrue(ApprovalLog.objects.filter(log_type='CRUD', action='DELETE', resource_type='user', resource_id=str(target_user_id)).exists())

    def test_middleware_records_detail_create(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'audit_worker',
            'password': 'StrongPass!234',
            'name': 'Audit Worker',
            'email': 'audit_worker@example.com',
        }, format='json')
        self.assertEqual(signup.status_code, 201)
        token = signup.data['token']

        org = Organization.objects.create(name='Audit Org', code='AUDIT_ORG', org_type='dept')
        subject = BudgetSubject.objects.create(code='AUDIT_SUB_01', name='Audit Subject', level=4, subject_type='expense')
        entry = BudgetEntry.objects.create(subject=subject, organization=org, year=2026, supplemental_round=0)
        profile = UserProfile.objects.get(user__username='audit_worker')
        profile.organization = org
        profile.save(update_fields=['organization'])

        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')
        detail_res = self.client.post('/api/details/', {
            'entry': entry.id,
            'name': 'Audit Detail',
            'price': 1000,
            'qty': 1,
            'freq': 1,
            'unit': '식',
            'source': '자체',
        }, format='json')
        self.assertEqual(detail_res.status_code, 201)

        self.assertTrue(
            ApprovalLog.objects.filter(
                log_type='CRUD',
                action='CREATE',
                resource_type='details',
                method='POST',
                status_code=201,
            ).exists()
        )


class BudgetSubjectBulkUpdateApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.root_a = BudgetSubject.objects.create(code='1000', name='Root A', level=1, subject_type='expense')
        self.root_b = BudgetSubject.objects.create(code='2000', name='Root B', level=1, subject_type='expense')
        self.child_a = BudgetSubject.objects.create(code='1100', name='Child A', level=2, parent=self.root_a, subject_type='expense')

    def test_bulk_update_allows_code_swap_atomically(self):
        response = self.client.post('/api/subjects/bulk-update-tree/', {
            'updates': [
                {'id': self.root_a.id, 'code': '2000'},
                {'id': self.root_b.id, 'code': '1000'},
                {'id': self.child_a.id, 'code': '2100'},
            ]
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.root_a.refresh_from_db()
        self.root_b.refresh_from_db()
        self.assertEqual(self.root_a.code, '2000')
        self.assertEqual(self.root_b.code, '1000')

    def test_bulk_update_rejects_invalid_hierarchy_code_and_rolls_back(self):
        original_a = self.root_a.code
        original_child = self.child_a.code

        response = self.client.post('/api/subjects/bulk-update-tree/', {
            'updates': [
                {'id': self.root_a.id, 'code': '3000'},
                {'id': self.child_a.id, 'code': '4100'},  # parent prefix mismatch (should start with 3)
            ]
        }, format='json')
        self.assertEqual(response.status_code, 400)

        self.root_a.refresh_from_db()
        self.child_a.refresh_from_db()
        self.assertEqual(self.root_a.code, original_a)
        self.assertEqual(self.child_a.code, original_child)


class EntrustedProjectDeleteGuardApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        signup = self.client.post('/api/auth/signup/', {
            'username': 'project_guard_admin',
            'password': 'StrongPass!234',
            'name': 'Project Guard',
            'email': 'project_guard_admin@example.com',
        }, format='json')
        self.token = signup.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {self.token}')
        self.org = Organization.objects.create(name='Guard Dept', code='GD01', org_type='dept')
        self.root = BudgetSubject.objects.create(code='7100', name='Guard Root', level=1, subject_type='expense')
        self.child = BudgetSubject.objects.create(code='7110', name='Guard Child', level=2, parent=self.root, subject_type='expense')

    def test_delete_project_keeps_budget_subjects_unchanged(self):
        project = EntrustedProject.objects.create(
            organization=self.org,
            year=2026,
            code='EP_GUARD_001',
            name='Guard Project',
            status='PLANNED',
        )
        subject_ids_before = list(BudgetSubject.objects.order_by('id').values_list('id', flat=True))

        response = self.client.delete(f'/api/entrusted-projects/{project.id}/')
        self.assertEqual(response.status_code, 204)
        subject_ids_after = list(BudgetSubject.objects.order_by('id').values_list('id', flat=True))
        self.assertEqual(subject_ids_after, subject_ids_before)

    def test_delete_project_with_linked_entries_is_blocked(self):
        project = EntrustedProject.objects.create(
            organization=self.org,
            year=2026,
            code='EP_GUARD_002',
            name='Guard Project Linked',
            status='PLANNED',
        )
        BudgetEntry.objects.create(
            subject=self.child,
            organization=self.org,
            entrusted_project=project,
            year=2026,
            supplemental_round=0,
        )

        response = self.client.delete(f'/api/entrusted-projects/{project.id}/')
        self.assertEqual(response.status_code, 409)
        self.assertTrue(EntrustedProject.objects.filter(id=project.id).exists())


class DashboardAndBulkUpsertApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.org = Organization.objects.create(name='Dept A', code='D001', org_type='dept')
        self.sub = BudgetSubject.objects.create(code='1111', name='Sub 1', level=4, subject_type='expense')

    def test_dashboard_summary_api(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'dashuser', 'password': 'StrongPass!234', 'name': 'Dash', 'email': 'dash@example.com'
        }, format='json')
        token = signup.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        entry = BudgetEntry.objects.create(subject=self.sub, organization=self.org, year=2026, status='FINALIZED')
        BudgetDetail.objects.create(entry=entry, name='Item 1', price=1000, qty=2, freq=1, source='자체')

        response = self.client.get('/api/dashboard/summary/', {'year': 2026})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['total_expense'], 2000)
        self.assertEqual(response.data['status_counts']['FINALIZED'], 1)

    def test_bulk_upsert_api(self):
        signup = self.client.post('/api/auth/signup/', {
            'username': 'bulkuser', 'password': 'StrongPass!234', 'name': 'Bulk', 'email': 'bulk@example.com'
        }, format='json')
        token = signup.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        payload = {
            'year': 2026,
            'entries': [
                {
                    'subject_code': '1111',
                    'org_code': 'D001',
                    'details': [
                        {'name': 'New Bulk Item', 'price': 5000, 'qty': 1, 'freq': 1}
                    ]
                }
            ]
        }
        response = self.client.post('/api/entries/bulk-upsert/', payload, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], 1)
        
        entry = BudgetEntry.objects.get(year=2026, subject__code='1111', organization__code='D001')
        self.assertEqual(entry.details.count(), 1)
        self.assertEqual(entry.details.first().price, 5000)


class BudgetVersionTransferApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        signup = self.client.post('/api/auth/signup/', {
            'username': 'transfer_admin',
            'password': 'StrongPass!234',
            'name': 'Transfer Admin',
            'email': 'transfer_admin@example.com',
        }, format='json')
        self.token = signup.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {self.token}')

        self.org = Organization.objects.create(name='Dept Transfer', code='DT01', org_type='dept')
        self.sub = BudgetSubject.objects.create(code='9111', name='Transfer Subject', level=4, subject_type='expense')

        self.source_version = BudgetVersion.objects.create(year=2025, round=0, name='2025년 본예산', status='CONFIRMED')
        self.source_entry = BudgetEntry.objects.create(
            subject=self.sub,
            organization=self.org,
            year=2025,
            supplemental_round=0,
            status='FINALIZED',
            budget_category='ORIGINAL',
        )
        self.source_detail = BudgetDetail.objects.create(
            entry=self.source_entry,
            name='기준 산출',
            price=1500,
            qty=2,
            freq=3,
            source='자체',
            unit='식',
            currency_unit='원',
            freq_unit='회',
        )

    def test_create_next_round_transfer_includes_before_and_after(self):
        response = self.client.post('/api/versions/create_next_round/', {
            'year': 2026,
            'name': '2026년 본예산(이관)',
            'creation_mode': 'TRANSFER',
            'source_version_id': self.source_version.id,
        }, format='multipart')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data.get('creation_mode'), 'TRANSFER')
        self.assertEqual(response.data.get('source_version'), self.source_version.id)
        self.assertEqual(response.data.get('cloned_count'), 1)

        target_round = response.data.get('round', 0)
        entry_list = self.client.get('/api/entries/', {'year': 2026, 'round': target_round})
        self.assertEqual(entry_list.status_code, 200)
        payload = entry_list.data if isinstance(entry_list.data, list) else entry_list.data.get('results', [])
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0].get('original_amount'), 9000)
        self.assertEqual(payload[0].get('variance_amount'), 0)

        details = payload[0].get('details') or []
        self.assertEqual(len(details), 1)
        detail = details[0]
        before_data = detail.get('before_data')
        after_data = detail.get('after_data')

        self.assertIsNotNone(before_data)
        self.assertIsNotNone(after_data)
        self.assertEqual(before_data.get('name'), '기준 산출')
        self.assertEqual(after_data.get('name'), '기준 산출')
        self.assertEqual(before_data.get('total_price'), 9000)
        self.assertEqual(after_data.get('total_price'), 9000)

        created_detail = BudgetDetail.objects.get(id=detail['id'])
        self.assertEqual(created_detail.transfer_source_detail_id, self.source_detail.id)
        self.assertEqual(created_detail.before_snapshot.get('name'), self.source_detail.name)

    def test_create_next_round_transfer_uses_source_total_for_original_amount(self):
        self.source_entry.total_amount = 12_345
        self.source_entry.save(update_fields=['total_amount'])

        response = self.client.post('/api/versions/create_next_round/', {
            'year': 2026,
            'name': '2026년 본예산(이관-기준금액)',
            'creation_mode': 'TRANSFER',
            'source_version_id': self.source_version.id,
        }, format='multipart')
        self.assertEqual(response.status_code, 201)

        created_entry = BudgetEntry.objects.get(
            year=2026,
            supplemental_round=response.data.get('round', 0),
            subject=self.sub,
            organization=self.org,
        )
        self.assertEqual(created_entry.last_year_amount, 12_345)
        self.assertEqual(created_entry.total_amount, 12_345)
        self.assertEqual(created_entry.remaining_amount, 12_345)

    def test_create_next_round_transfer_clones_all_source_entries(self):
        second_subject = BudgetSubject.objects.create(
            code='9112',
            name='Transfer Subject 2',
            level=4,
            subject_type='expense',
        )
        second_entry = BudgetEntry.objects.create(
            subject=second_subject,
            organization=self.org,
            year=2025,
            supplemental_round=0,
            status='DRAFT',
            budget_category='ORIGINAL',
        )
        BudgetDetail.objects.create(
            entry=second_entry,
            name='Detail 2',
            price=1000,
            qty=2,
            freq=4,
            source='SELF',
            unit='EA',
            currency_unit='KRW',
            freq_unit='TIMES',
        )

        response = self.client.post('/api/versions/create_next_round/', {
            'year': 2026,
            'name': '2026 transfer all',
            'creation_mode': 'TRANSFER',
            'source_version_id': self.source_version.id,
        }, format='multipart')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data.get('cloned_count'), 2)

        created_entries = BudgetEntry.objects.filter(
            year=2026,
            supplemental_round=response.data.get('round', 0),
        ).order_by('subject_id')
        self.assertEqual(created_entries.count(), 2)
        self.assertSetEqual(set(created_entries.values_list('subject_id', flat=True)), {self.sub.id, second_subject.id})

        second_cloned = created_entries.get(subject=second_subject)
        self.assertEqual(second_cloned.last_year_amount, 8_000)
        self.assertEqual(second_cloned.total_amount, 8_000)
        self.assertEqual(second_cloned.remaining_amount, 8_000)

    def test_transfer_entry_baseline_amount_is_immutable(self):
        response = self.client.post('/api/versions/create_next_round/', {
            'year': 2026,
            'name': '2026 transfer baseline lock',
            'creation_mode': 'TRANSFER',
            'source_version_id': self.source_version.id,
        }, format='multipart')
        self.assertEqual(response.status_code, 201)

        cloned_entry = BudgetEntry.objects.get(
            year=2026,
            supplemental_round=response.data.get('round', 0),
            subject=self.sub,
            organization=self.org,
        )
        original_amount = cloned_entry.last_year_amount

        patch_response = self.client.patch(
            f'/api/entries/{cloned_entry.id}/',
            {'last_year_amount': original_amount + 1000},
            format='json',
        )
        self.assertEqual(patch_response.status_code, 400)
        self.assertIn('last_year_amount', patch_response.data)

        cloned_entry.refresh_from_db()
        self.assertEqual(cloned_entry.last_year_amount, original_amount)

    def test_transfer_entry_variance_changes_immediately_after_detail_edit(self):
        response = self.client.post('/api/versions/create_next_round/', {
            'year': 2026,
            'name': '2026 transfer variance update',
            'creation_mode': 'TRANSFER',
            'source_version_id': self.source_version.id,
        }, format='multipart')
        self.assertEqual(response.status_code, 201)

        cloned_entry = BudgetEntry.objects.get(
            year=2026,
            supplemental_round=response.data.get('round', 0),
            subject=self.sub,
            organization=self.org,
        )
        cloned_detail = cloned_entry.details.first()
        self.assertIsNotNone(cloned_detail)

        update_response = self.client.patch(
            f'/api/details/{cloned_detail.id}/',
            {'price': 2000},
            format='json',
        )
        self.assertEqual(update_response.status_code, 200)

        cloned_entry.refresh_from_db()
        self.assertEqual(cloned_entry.last_year_amount, 9000)
        self.assertEqual(cloned_entry.total_amount, 12_000)
        self.assertEqual(cloned_entry.remaining_amount, 12_000)

        entry_response = self.client.get(f'/api/entries/{cloned_entry.id}/')
        self.assertEqual(entry_response.status_code, 200)
        self.assertEqual(entry_response.data.get('original_amount'), 9000)
        self.assertEqual(entry_response.data.get('variance_amount'), 3000)


class BudgetScopeAndConcurrencyApiTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.org_a = Organization.objects.create(name='Scope Dept A', code='SCOPE_A', org_type='dept')
        self.org_b = Organization.objects.create(name='Scope Dept B', code='SCOPE_B', org_type='dept')
        self.subject = BudgetSubject.objects.create(code='SCP1', name='Scope Subject', level=4, subject_type='expense')

        self.user_a = User.objects.create_user(username='scope_user_a', password='StrongPass!234')
        UserProfile.objects.create(user=self.user_a, role='STAFF', organization=self.org_a)

        self.entry_a = BudgetEntry.objects.create(
            subject=self.subject,
            organization=self.org_a,
            year=2026,
            supplemental_round=0,
            status='DRAFT',
        )
        self.entry_b = BudgetEntry.objects.create(
            subject=self.subject,
            organization=self.org_b,
            year=2026,
            supplemental_round=0,
            status='DRAFT',
        )

    def test_staff_cannot_read_or_write_other_org_data(self):
        self.client.force_authenticate(user=self.user_a)

        list_response = self.client.get('/api/entries/', {'year': 2026, 'round': 0})
        self.assertEqual(list_response.status_code, 200)
        payload = list_response.data if isinstance(list_response.data, list) else list_response.data.get('results', [])
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]['id'], self.entry_a.id)

        other_detail_response = self.client.get(f'/api/entries/{self.entry_b.id}/')
        self.assertEqual(other_detail_response.status_code, 404)

        create_detail_response = self.client.post('/api/details/', {
            'entry': self.entry_b.id,
            'name': 'Blocked Detail',
            'price': 1000,
            'qty': 1,
            'freq': 1,
            'unit': 'EA',
            'source': 'SELF',
        }, format='json')
        self.assertEqual(create_detail_response.status_code, 403)

    def test_detail_update_returns_409_on_stale_updated_at(self):
        detail = BudgetDetail.objects.create(
            entry=self.entry_a,
            name='Concurrency Item',
            price=1000,
            qty=1,
            freq=1,
            source='SELF',
            unit='EA',
        )
        self.client.force_authenticate(user=self.user_a)

        first_get = self.client.get(f'/api/details/{detail.id}/')
        self.assertEqual(first_get.status_code, 200)
        stale_updated_at = first_get.data.get('updated_at')
        self.assertTrue(stale_updated_at)

        update_ok = self.client.patch(
            f'/api/details/{detail.id}/',
            {'price': 2000, '_updated_at': stale_updated_at},
            format='json',
        )
        self.assertEqual(update_ok.status_code, 200)

        update_conflict = self.client.patch(
            f'/api/details/{detail.id}/',
            {'price': 3000, '_updated_at': stale_updated_at},
            format='json',
        )
        self.assertEqual(update_conflict.status_code, 409)
        self.assertEqual(update_conflict.data.get('code'), 'DETAIL_CONFLICT')


class BudgetBookExportServiceTest(TestCase):
    def _make_template(self, include_official_sheets: bool = True) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = '예산목별조서'
        ws['A4'] = '수입 계'
        ws['A17'] = '지출 계'
        ws['F3'] = '기존부서1'
        ws['G3'] = '기존부서2'
        ws['F16'] = '=F3'
        ws['G16'] = '=G3'

        ws_income = wb.create_sheet('수입(총괄)')
        ws_income['A4'] = '수입계'

        ws_expense = wb.create_sheet('지출(총괄)')
        ws_expense['A4'] = '지출계'

        if include_official_sheets:
            ws_basic = wb.create_sheet('기본재산명세서')
            ws_basic['G3'] = '<기준일 : 2000. 12. 31 >'
            ws_basic.merge_cells('G7:G14')
            ws_basic['B9'] = '삭제대상'
            ws_basic['E15'] = '=SUM(D6:D10)'
            ws_basic['F15'] = '=E15'
            ws_basic['F16'] = '=F15'
            ws_basic['F19'] = '=SUM(F15:F18)'

            ws_ordinary = wb.create_sheet('보통재산명세서')
            ws_ordinary['F3'] = '<기준일 : 2000. 12. 31 >'
            ws_ordinary['E7'] = '=SUM(E6:E6)'
            ws_ordinary['F7'] = '=SUM(F6:F6)'
            ws_ordinary['E9'] = '=SUM(E8:E8)'
            ws_ordinary['F9'] = '=SUM(F8:F8)'
            ws_ordinary['E12'] = '=SUM(E10:E11)'
            ws_ordinary['F12'] = '=SUM(F10:F11)'
            ws_ordinary['E19'] = '=SUM(E6:E18)'
            ws_ordinary['F19'] = '=SUM(F6:F18)'
            ws_ordinary['G19'] = '=F19-E19'

            ws_labor = wb.create_sheet('총인건비명세서 ')
            ws_labor['B3'] = '구분'
            ws_labor['F5'] = '=SUM(F6:F10)'
            ws_labor['G5'] = '=SUM(G6:G10)'
            ws_labor['F11'] = '=SUM(F6:F10)'
            ws_labor['G11'] = '=SUM(G6:G10)'
            ws_labor['F12'] = '=G12*0.1'
            ws_labor['F13'] = '=G13*0.1'
            ws_labor['F14'] = '=G14*0.1'
            ws_labor['F15'] = '=G15*0.1'
            ws_labor['F16'] = '=G16*0.1'
            ws_labor['F17'] = '=G17*0.1'
            ws_labor['F18'] = '=SUM(F12:F17)'
            ws_labor['F19'] = '=SUM(F18:G18)'
            ws_labor['H6'] = '=F6+G6'
            ws_labor['K6'] = '=IF(H6=0,0,G6/H6)'

        tmp = tempfile.NamedTemporaryFile(prefix='budget_template_', suffix='.xlsx', delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()
        wb.save(tmp_path)
        wb.close()
        return tmp_path

    def test_build_budget_book_file_writes_template_totals_and_seed_sheets(self):
        dept = Organization.objects.create(name='테스트부서', code='TDEP', org_type='dept', sort_order=1)
        Organization.objects.create(name='예비부서', code='SDEP', org_type='dept', sort_order=2)
        team = Organization.objects.create(name='테스트팀', code='TTEAM', org_type='team', parent=dept)

        inc_j = BudgetSubject.objects.create(code='INC_T_01', name='수입장', level=1, subject_type='income')
        inc_g = BudgetSubject.objects.create(code='INC_T_01_01', name='수입관', level=2, subject_type='income', parent=inc_j)
        inc_h = BudgetSubject.objects.create(code='INC_T_01_01_01', name='수입항', level=3, subject_type='income', parent=inc_g)
        inc_m = BudgetSubject.objects.create(code='INC_T_01_01_01_01', name='수입목', level=4, subject_type='income', parent=inc_h)

        exp_j = BudgetSubject.objects.create(code='EXP_T_01', name='지출장', level=1, subject_type='expense')
        exp_g = BudgetSubject.objects.create(code='EXP_T_01_01', name='지출관', level=2, subject_type='expense', parent=exp_j)
        exp_h = BudgetSubject.objects.create(code='EXP_T_01_01_01', name='지출항', level=3, subject_type='expense', parent=exp_g)
        exp_m = BudgetSubject.objects.create(code='EXP_T_01_01_01_01', name='지출목', level=4, subject_type='expense', parent=exp_h)

        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='DRAFT')

        BudgetEntry.objects.create(
            subject=inc_m,
            organization=team,
            year=2026,
            supplemental_round=0,
            total_amount=5_000_000,
            last_year_amount=4_000_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )
        BudgetEntry.objects.create(
            subject=exp_m,
            organization=team,
            year=2026,
            supplemental_round=0,
            total_amount=8_000_000,
            last_year_amount=6_000_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        template_path = self._make_template()
        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                file_bytes, file_name, disposition, meta = build_budget_book_file(version)
        finally:
            template_path.unlink(missing_ok=True)

        self.assertTrue(file_name.endswith('.xlsx'))
        self.assertIn('attachment;', disposition)
        self.assertEqual(meta.get('row_count'), 2)
        self.assertIn('template_overrides', meta)
        self.assertGreater(len(file_bytes), 0)

        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        self.assertIn('IBMS_기초데이터', wb.sheetnames)
        self.assertIn('IBMS_수입총괄_자동', wb.sheetnames)
        self.assertIn('IBMS_지출총괄_자동', wb.sheetnames)
        self.assertIn('IBMS_보류시트_샘플', wb.sheetnames)
        self.assertIn('IBMS_자산명세서_샘플', wb.sheetnames)
        self.assertIn('IBMS_총인건비_샘플', wb.sheetnames)

        matrix = wb['예산목별조서']
        self.assertEqual(matrix['F3'].value, '테스트부서')
        self.assertEqual(matrix['G3'].value, '예비부서')
        self.assertEqual(matrix['F16'].value, '=F3')
        self.assertEqual(matrix['G16'].value, '=G3')
        self.assertEqual(matrix['E4'].value, 5000)
        self.assertEqual(matrix['E17'].value, 8000)
        self.assertEqual(matrix['F4'].value, 5000)
        self.assertEqual(matrix['F17'].value, 8000)
        self.assertEqual(matrix['G4'].value, 0)
        self.assertEqual(matrix['G17'].value, 0)

        income_total = wb['수입(총괄)']
        self.assertEqual(income_total['E4'].value, 5000)
        self.assertEqual(income_total['F4'].value, 4000)
        self.assertEqual(income_total['G4'].value, 1000)

        expense_total = wb['지출(총괄)']
        self.assertEqual(expense_total['E4'].value, 8000)
        self.assertEqual(expense_total['F4'].value, 6000)
        self.assertEqual(expense_total['G4'].value, 2000)

        seed = wb['IBMS_기초데이터']
        self.assertEqual(seed['A1'].value, '연도')
        self.assertGreaterEqual(seed.max_row, 3)

        asset_sample = wb['IBMS_자산명세서_샘플']
        self.assertEqual(asset_sample['A1'].value, '구분')
        self.assertEqual(asset_sample['A2'].value, '기본재산')

        labor_sample = wb['IBMS_총인건비_샘플']
        self.assertEqual(labor_sample['A1'].value, '구분')
        self.assertEqual(labor_sample['A4'].value, '합계')

        basic_asset = wb['기본재산명세서']
        self.assertEqual(basic_asset['G3'].value, '<기준일 : 2025. 12. 31 >')
        self.assertEqual(basic_asset['B5'].value, '현 금')
        self.assertEqual(basic_asset['F5'].value, 200_000_000)
        self.assertEqual(basic_asset['B7'].value, '토지1')
        self.assertEqual(basic_asset['F7'].value, 1_374_911_200)
        self.assertEqual(basic_asset['G7'].value, '출연부지 및 \n매입토지')
        self.assertEqual(basic_asset['B9'].value, '토지3')
        self.assertEqual(basic_asset['F14'].value, 829_148_580)
        self.assertEqual(basic_asset['E15'].value, '=SUM(D6:D10)')
        self.assertEqual(basic_asset['F15'].value, '=E15')
        self.assertEqual(basic_asset['F16'].value, '=F15')
        self.assertEqual(basic_asset['F19'].value, '=SUM(F15:F18)')

        ordinary_asset = wb['보통재산명세서']
        self.assertEqual(ordinary_asset['F3'].value, '<기준일 : 2025. 12. 31 >')
        self.assertEqual(ordinary_asset['E5'].value, 5_441_863_488)
        self.assertEqual(ordinary_asset['F5'].value, 5_441_863_488)
        self.assertEqual(ordinary_asset['E6'].value, 25_416_589_448)
        self.assertEqual(ordinary_asset['F6'].value, 25_416_589_448)
        self.assertEqual(ordinary_asset['E7'].value, '=SUM(E6:E6)')
        self.assertEqual(ordinary_asset['F7'].value, '=SUM(F6:F6)')
        self.assertEqual(ordinary_asset['E9'].value, '=SUM(E8:E8)')
        self.assertEqual(ordinary_asset['F9'].value, '=SUM(F8:F8)')
        self.assertEqual(ordinary_asset['E12'].value, '=SUM(E10:E11)')
        self.assertEqual(ordinary_asset['F12'].value, '=SUM(F10:F11)')
        self.assertEqual(ordinary_asset['E19'].value, '=SUM(E6:E18)')
        self.assertEqual(ordinary_asset['F19'].value, '=SUM(F6:F18)')
        self.assertEqual(ordinary_asset['G19'].value, '=F19-E19')

        labor_statement = wb['총인건비명세서 ']
        self.assertEqual(labor_statement['B3'].value, '구분')
        self.assertEqual(labor_statement['F6'].value, 120_315)
        self.assertEqual(labor_statement['G6'].value, 114_586)
        self.assertEqual(labor_statement['G18'].value, 7_170_495)
        self.assertEqual(labor_statement['G19'].value, 780_583)
        self.assertEqual(labor_statement['F5'].value, '=SUM(F6:F10)')
        self.assertEqual(labor_statement['G5'].value, '=SUM(G6:G10)')
        self.assertEqual(labor_statement['F11'].value, '=SUM(F6:F10)')
        self.assertEqual(labor_statement['G11'].value, '=SUM(G6:G10)')
        self.assertEqual(labor_statement['F12'].value, '=G12*0.1')
        self.assertEqual(labor_statement['F13'].value, '=G13*0.1')
        self.assertEqual(labor_statement['F14'].value, '=G14*0.1')
        self.assertEqual(labor_statement['F15'].value, '=G15*0.1')
        self.assertEqual(labor_statement['F16'].value, '=G16*0.1')
        self.assertEqual(labor_statement['F17'].value, '=G17*0.1')
        self.assertEqual(labor_statement['F18'].value, '=SUM(F12:F17)')
        self.assertEqual(labor_statement['F19'].value, '=SUM(F18:G18)')
        self.assertEqual(labor_statement['H6'].value, '=F6+G6')
        self.assertEqual(labor_statement['K6'].value, '=IF(H6=0,0,G6/H6)')

        overrides = meta.get('template_overrides') or []
        self.assertGreaterEqual(len(overrides), 6)
        by_sheet = {item.get('sheet'): item for item in overrides}
        self.assertTrue(by_sheet['예산목별조서']['applied'])
        self.assertTrue(by_sheet['수입(총괄)']['applied'])
        self.assertTrue(by_sheet['지출(총괄)']['applied'])
        self.assertTrue(by_sheet['기본재산명세서']['applied'])
        self.assertTrue(by_sheet['보통재산명세서']['applied'])
        labor_override = next((item for item in overrides if str(item.get('sheet', '')).strip() == '총인건비명세서'), None)
        self.assertIsNotNone(labor_override)
        self.assertTrue(labor_override['applied'])
        self.assertGreater(labor_override['updated_cells_count'], 0)

    def test_build_budget_book_file_uses_original_budget_header_for_transfer_version(self):
        dept = Organization.objects.create(name='이관부서', code='TRDEP', org_type='dept')
        inc_m = BudgetSubject.objects.create(code='INC_TR_1', name='이관수입목', level=4, subject_type='income')
        exp_m = BudgetSubject.objects.create(code='EXP_TR_1', name='이관지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(
            year=2026,
            round=1,
            name='2026년 1차 추경',
            status='DRAFT',
            creation_mode='TRANSFER',
        )

        BudgetEntry.objects.create(
            subject=inc_m,
            organization=dept,
            year=2026,
            supplemental_round=1,
            total_amount=3_000_000,
            last_year_amount=2_500_000,
            status='DRAFT',
            budget_category='SUPPLEMENTAL',
        )
        BudgetEntry.objects.create(
            subject=exp_m,
            organization=dept,
            year=2026,
            supplemental_round=1,
            total_amount=4_000_000,
            last_year_amount=3_500_000,
            status='DRAFT',
            budget_category='SUPPLEMENTAL',
        )

        template_path = self._make_template()
        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                file_bytes, _file_name, _disposition, _meta = build_budget_book_file(version)
        finally:
            template_path.unlink(missing_ok=True)

        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        seed = wb['IBMS_기초데이터']
        income_summary = wb['IBMS_수입총괄_자동']
        expense_summary = wb['IBMS_지출총괄_자동']
        self.assertEqual(seed['O1'].value, '당초예산액(원)')
        self.assertEqual(income_summary['H1'].value, '당초예산액(원)')
        self.assertEqual(expense_summary['H1'].value, '당초예산액(원)')
        wb.close()

    def test_build_budget_book_file_skips_missing_official_sheets_without_error(self):
        dept = Organization.objects.create(name='테스트부서', code='NODEP', org_type='dept')
        leaf = BudgetSubject.objects.create(code='MIN_EXP_1', name='최소지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='DRAFT')
        BudgetEntry.objects.create(
            subject=leaf,
            organization=dept,
            year=2026,
            supplemental_round=0,
            total_amount=1_000_000,
            last_year_amount=500_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        template_path = self._make_template(include_official_sheets=False)
        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                file_bytes, _file_name, _disposition, meta = build_budget_book_file(version)
        finally:
            template_path.unlink(missing_ok=True)

        self.assertGreater(len(file_bytes), 0)
        overrides = meta.get('template_overrides') or []
        by_sheet = {item.get('sheet'): item for item in overrides}
        self.assertFalse(by_sheet['기본재산명세서']['applied'])
        self.assertEqual(by_sheet['기본재산명세서']['reason'], 'sheet_not_found')
        self.assertFalse(by_sheet['보통재산명세서']['applied'])
        self.assertEqual(by_sheet['보통재산명세서']['reason'], 'sheet_not_found')
        self.assertFalse(by_sheet['총인건비명세서']['applied'])
        self.assertEqual(by_sheet['총인건비명세서']['reason'], 'sheet_not_found')

    def test_build_budget_book_file_preserves_formula_cells_in_template(self):
        dept = Organization.objects.create(name='수식부서', code='FDEP', org_type='dept')
        leaf = BudgetSubject.objects.create(code='FORM_EXP_1', name='수식지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='DRAFT')
        BudgetEntry.objects.create(
            subject=leaf,
            organization=dept,
            year=2026,
            supplemental_round=0,
            total_amount=2_000_000,
            last_year_amount=1_000_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        wb = Workbook()
        ws = wb.active
        ws.title = '예산목별조서'
        ws['A4'] = '수입 계'
        ws['A17'] = '지출 계'
        ws['F3'] = '수식부서'
        ws['F16'] = '=F3'
        ws['E4'] = '=SUM(E5:E15)'
        ws['F4'] = '=SUM(F5:F15)'
        ws['E17'] = '=SUM(E18:E27)'
        ws['F17'] = '=SUM(F18:F27)'
        ws_income = wb.create_sheet('수입(총괄)')
        ws_income['A4'] = '수입계'
        ws_income['E4'] = '=E5+E502'
        ws_income['F4'] = '=F5+F502'
        ws_income['G4'] = '=G5+G502'
        ws_expense = wb.create_sheet('지출(총괄)')
        ws_expense['A4'] = '지출계'
        ws_expense['E4'] = '=E5+E254+E2830+E2838+E2936'
        ws_expense['F4'] = '=F5+F254+F2830+F2838+F2936'
        ws_expense['G4'] = '=G5+G254+G2830+G2838+G2936'

        tmp = tempfile.NamedTemporaryFile(prefix='budget_formula_template_', suffix='.xlsx', delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()
        wb.save(tmp_path)
        wb.close()

        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(tmp_path)):
                file_bytes, _file_name, _disposition, _meta = build_budget_book_file(version)
        finally:
            tmp_path.unlink(missing_ok=True)

        out = load_workbook(BytesIO(file_bytes), data_only=False)
        matrix = out['예산목별조서']
        self.assertEqual(matrix['E4'].value, '=SUM(E5:E15)')
        self.assertEqual(matrix['F4'].value, '=SUM(F5:F15)')
        self.assertEqual(matrix['E17'].value, '=SUM(E18:E27)')
        self.assertEqual(matrix['F17'].value, '=SUM(F18:F27)')
        self.assertEqual(matrix['F16'].value, '=F3')
        income_total = out['수입(총괄)']
        self.assertEqual(income_total['E4'].value, '=E5+E502')
        self.assertEqual(income_total['F4'].value, '=F5+F502')
        self.assertEqual(income_total['G4'].value, '=G5+G502')
        expense_total = out['지출(총괄)']
        self.assertEqual(expense_total['E4'].value, '=E5+E254+E2830+E2838+E2936')
        self.assertEqual(expense_total['F4'].value, '=F5+F254+F2830+F2838+F2936')
        self.assertEqual(expense_total['G4'].value, '=G5+G254+G2830+G2838+G2936')
        out.close()

    def test_build_budget_book_file_continues_when_single_override_fails(self):
        dept = Organization.objects.create(name='오류부서', code='EDEP', org_type='dept')
        leaf = BudgetSubject.objects.create(code='ERR_EXP_1', name='오류지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='DRAFT')
        BudgetEntry.objects.create(
            subject=leaf,
            organization=dept,
            year=2026,
            supplemental_round=0,
            total_amount=1_000_000,
            last_year_amount=800_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        template_path = self._make_template(include_official_sheets=True)
        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                with mock.patch(
                    'budget_mgmt.services.budget_book_export._write_basic_asset_sheet_overrides',
                    side_effect=RuntimeError('forced template error'),
                ):
                    file_bytes, _file_name, _disposition, meta = build_budget_book_file(version)
        finally:
            template_path.unlink(missing_ok=True)

        self.assertGreater(len(file_bytes), 0)
        self.assertGreaterEqual(meta.get('template_warning_count', 0), 1)
        warnings = meta.get('template_warnings') or []
        by_sheet = {item.get('sheet'): item for item in (meta.get('template_overrides') or [])}
        self.assertIn('기본재산명세서', by_sheet)
        self.assertFalse(by_sheet['기본재산명세서']['applied'])
        self.assertEqual(by_sheet['기본재산명세서']['reason'], 'error:RuntimeError')
        self.assertTrue(any(item.get('sheet') == '기본재산명세서' for item in warnings))


class BudgetBookExportApiTest(TestCase):
    def _make_template(self) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = '예산목별조서'
        ws['A4'] = '수입 계'
        ws['A17'] = '지출 계'
        ws['F3'] = '테스트부서'
        ws['F16'] = '=F3'
        wb.create_sheet('수입(총괄)')['A4'] = '수입계'
        wb.create_sheet('지출(총괄)')['A4'] = '지출계'
        tmp = tempfile.NamedTemporaryFile(prefix='budget_template_api_', suffix='.xlsx', delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()
        wb.save(tmp_path)
        wb.close()
        return tmp_path

    def test_export_budget_book_endpoint_returns_xlsx(self):
        client = APIClient()
        dept = Organization.objects.create(name='테스트부서', code='APIDEP', org_type='dept')
        leaf = BudgetSubject.objects.create(code='API_EXP_1', name='API지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='PENDING')
        BudgetEntry.objects.create(
            subject=leaf,
            organization=dept,
            year=2026,
            supplemental_round=0,
            total_amount=1_000_000,
            last_year_amount=900_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        template_path = self._make_template()
        try:
            with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                response = client.get(f'/api/versions/{version.id}/export-budget-book/')
        finally:
            template_path.unlink(missing_ok=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment;', response['Content-Disposition'])
        self.assertIn('filename*=', response['Content-Disposition'])
        self.assertIn('X-Export-Template-Warnings', response)
        self.assertIn('X-Export-Template-Overrides', response)
        self.assertGreaterEqual(int(response['X-Export-Template-Warnings']), 0)
        self.assertGreaterEqual(int(response['X-Export-Template-Overrides']), 3)
        self.assertGreater(len(response.content), 0)

        wb = load_workbook(BytesIO(response.content), data_only=False)
        self.assertIn('IBMS_기초데이터', wb.sheetnames)
        self.assertIn('IBMS_자산명세서_샘플', wb.sheetnames)
        self.assertIn('IBMS_총인건비_샘플', wb.sheetnames)
        wb.close()


class BudgetBookExportAuditCommandTest(TestCase):
    def _make_template(self) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = '예산목별조서'
        ws['A4'] = '수입 계'
        ws['A17'] = '지출 계'
        ws['F3'] = '테스트부서'
        ws['F16'] = '=F3'
        wb.create_sheet('수입(총괄)')['A4'] = '수입계'
        wb.create_sheet('지출(총괄)')['A4'] = '지출계'
        wb.create_sheet('기본재산명세서')['G3'] = '<기준일 : 2025. 12. 31 >'
        wb.create_sheet('보통재산명세서')['F3'] = '<기준일 : 2025. 12. 31 >'
        wb.create_sheet('총인건비명세서 ')['B1'] = '총인건비 명세서'
        tmp = tempfile.NamedTemporaryFile(prefix='budget_template_cmd_', suffix='.xlsx', delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()
        wb.save(tmp_path)
        wb.close()
        return tmp_path

    def test_audit_budget_book_export_command_outputs_json_report(self):
        dept = Organization.objects.create(name='감사부서', code='AUDITD', org_type='dept')
        leaf = BudgetSubject.objects.create(code='AUDIT_EXP_1', name='감사지출목', level=4, subject_type='expense')
        version = BudgetVersion.objects.create(year=2026, round=0, name='2026년 본예산', status='DRAFT')
        BudgetEntry.objects.create(
            subject=leaf,
            organization=dept,
            year=2026,
            supplemental_round=0,
            total_amount=1_000_000,
            last_year_amount=900_000,
            status='DRAFT',
            budget_category='ORIGINAL',
        )

        template_path = self._make_template()
        stdout = StringIO()
        with tempfile.TemporaryDirectory(prefix='audit_budget_') as tmpdir:
            json_path = Path(tmpdir) / 'audit.json'
            try:
                with override_settings(BUDGET_BOOK_TEMPLATE_PATH=str(template_path)):
                    call_command(
                        'audit_budget_book_export',
                        version_id=version.id,
                        output_json=str(json_path),
                        max_diffs_per_sheet=5,
                        stdout=stdout,
                    )
            finally:
                template_path.unlink(missing_ok=True)

            self.assertTrue(json_path.exists())
            payload = json.loads(json_path.read_text(encoding='utf-8'))

        self.assertEqual(payload.get('version_id'), version.id)
        self.assertIn('sheets', payload)
        self.assertIsInstance(payload['sheets'], list)
        self.assertIn('Budget book export audit completed.', stdout.getvalue())
