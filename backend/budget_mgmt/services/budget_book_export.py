from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import logging
from typing import Any, Iterable
from urllib.parse import quote

from django.conf import settings

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from ..models import BudgetEntry, BudgetVersion, Organization

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_DEFAULT_TEMPLATE_NAMES = (
    "2026\ub144 \ubcf8\uc608\uc0b0\uc11c.xlsx",
    "tmp_budget_2026.xlsx",
)

_SHEET_SEED = "IBMS_\uae30\ucd08\ub370\uc774\ud130"
_SHEET_INCOME_SUMMARY = "IBMS_\uc218\uc785\ucd1d\uad04_\uc790\ub3d9"
_SHEET_EXPENSE_SUMMARY = "IBMS_\uc9c0\ucd9c\ucd1d\uad04_\uc790\ub3d9"
_SHEET_DEFERRED = "IBMS_\ubcf4\ub958\uc2dc\ud2b8_\uc0d8\ud50c"
_SHEET_ASSET_SAMPLE = "IBMS_\uc790\uc0b0\uba85\uc138\uc11c_\uc0d8\ud50c"
_SHEET_LABOR_SAMPLE = "IBMS_\ucd1d\uc778\uac74\ube44_\uc0d8\ud50c"
_SHEET_MATRIX = "\uc608\uc0b0\ubaa9\ubcc4\uc870\uc11c"
_SHEET_INCOME_TOTAL = "\uc218\uc785(\ucd1d\uad04)"
_SHEET_EXPENSE_TOTAL = "\uc9c0\ucd9c(\ucd1d\uad04)"
_SHEET_BASIC_ASSET = "\uae30\ubcf8\uc7ac\uc0b0\uba85\uc138\uc11c"
_SHEET_ORDINARY_ASSET = "\ubcf4\ud1b5\uc7ac\uc0b0\uba85\uc138\uc11c"
_SHEET_LABOR_STATEMENT = "\ucd1d\uc778\uac74\ube44\uba85\uc138\uc11c"

logger = logging.getLogger(__name__)


@dataclass
class _EntryRow:
    subject_type: str
    jang: str
    gwan: str
    hang: str
    mok: str
    subject_code: str
    subject_name: str
    description: str
    department_name: str
    organization_name: str
    project_name: str
    current_amount: int
    previous_amount: int
    detail_count: int
    status: str
    budget_category: str
    carryover_type: str


def _repo_root() -> Path:
    return Path(getattr(settings, "BASE_DIR", Path.cwd())).resolve().parent


def _resolve_template_path() -> Path | None:
    configured = str(getattr(settings, "BUDGET_BOOK_TEMPLATE_PATH", "") or "").strip()
    if configured:
        candidate = Path(configured).expanduser()
        if not candidate.is_absolute():
            candidate = _repo_root() / candidate
        if candidate.exists():
            return candidate

    root = _repo_root()
    for file_name in _DEFAULT_TEMPLATE_NAMES:
        candidate = root / file_name
        if candidate.exists():
            return candidate
    return None


def _open_workbook() -> tuple[Workbook, Path | None]:
    template_path = _resolve_template_path()
    if template_path is not None:
        wb = load_workbook(template_path)
        return wb, template_path
    return Workbook(), None


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key(value: str) -> str:
    text = _safe_str(value).lower()
    return re.sub(r"[\s\-\(\)\[\]{}<>\.,/:;|_·'\"`]+", "", text)


def _to_thousand_won(amount: int) -> int:
    return int(round((int(amount or 0)) / 1000.0))


def _subject_path(subject) -> tuple[str, str, str, str]:
    if subject is None:
        return "", "", "", ""

    by_level = {}
    cursor = subject
    guard = 0
    while cursor is not None and guard < 8:
        level = int(getattr(cursor, "level", 0) or 0)
        if level not in by_level and level > 0:
            by_level[level] = _safe_str(getattr(cursor, "name", ""))
        cursor = getattr(cursor, "parent", None)
        guard += 1

    jang = by_level.get(1, "")
    gwan = by_level.get(2, "")
    hang = by_level.get(3, "")
    mok = by_level.get(4, "")
    return jang, gwan, hang, mok


def _root_department_name(org) -> str:
    if org is None:
        return ""
    cursor = org
    guard = 0
    while getattr(cursor, "parent", None) is not None and guard < 8:
        cursor = cursor.parent
        guard += 1
    return _safe_str(getattr(cursor, "name", ""))


def _entry_amount(entry, detail_items) -> int:
    total_amount = int(getattr(entry, "total_amount", 0) or 0)
    if total_amount:
        return total_amount
    detail_total = 0
    for item in detail_items:
        try:
            detail_total += int(item.total_price or 0)
        except Exception:
            continue
    return detail_total


def _collect_entry_rows(version: BudgetVersion) -> list[_EntryRow]:
    queryset = (
        BudgetEntry.objects
        .filter(year=version.year, supplemental_round=version.round)
        .select_related(
            "subject",
            "subject__parent",
            "subject__parent__parent",
            "subject__parent__parent__parent",
            "organization",
            "organization__parent",
            "entrusted_project",
        )
        .prefetch_related("details")
        .order_by("subject__subject_type", "subject__code", "organization__id", "id")
    )

    rows: list[_EntryRow] = []
    for entry in queryset:
        details = list(entry.details.all())
        subject = entry.subject
        jang, gwan, hang, mok = _subject_path(subject)
        subject_name = _safe_str(getattr(subject, "name", ""))
        leaf_name = mok or hang or gwan or jang or subject_name
        rows.append(_EntryRow(
            subject_type=_safe_str(getattr(subject, "subject_type", "")),
            jang=jang,
            gwan=gwan,
            hang=hang,
            mok=mok or leaf_name,
            subject_code=_safe_str(getattr(subject, "code", "")),
            subject_name=subject_name,
            description=_safe_str(getattr(subject, "description", "")),
            department_name=_root_department_name(entry.organization),
            organization_name=_safe_str(getattr(entry.organization, "name", "")),
            project_name=_safe_str(getattr(entry.entrusted_project, "name", "")),
            current_amount=_entry_amount(entry, details),
            previous_amount=int(getattr(entry, "last_year_amount", 0) or 0),
            detail_count=len(details),
            status=_safe_str(getattr(entry, "status", "")),
            budget_category=_safe_str(getattr(entry, "budget_category", "")),
            carryover_type=_safe_str(getattr(entry, "carryover_type", "")),
        ))
    return rows


def _get_or_create_sheet(wb: Workbook, title: str):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def _clear_sheet(ws):
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def _base_amount_header(version: BudgetVersion) -> str:
    # TRANSFER 회차에서는 원본 회차 금액을 "당초 예산액"으로 표기한다.
    if str(getattr(version, "creation_mode", "") or "").upper() == "TRANSFER":
        return "\ub2f9\ucd08\uc608\uc0b0\uc561(\uc6d0)"
    return "\uc804\ub144\ub3c4\uc608\uc0b0\uc561(\uc6d0)"


def _write_seed_sheet(wb: Workbook, version: BudgetVersion, rows: list[_EntryRow]) -> None:
    ws = _get_or_create_sheet(wb, _SHEET_SEED)
    _clear_sheet(ws)

    base_header = _base_amount_header(version)
    headers = [
        "\uc5f0\ub3c4",
        "\ud68c\ucc28",
        "\uc218\uc785/\uc9c0\ucd9c",
        "\uc7a5",
        "\uad00",
        "\ud56d",
        "\ubaa9",
        "\uacfc\ubaa9\ucf54\ub4dc",
        "\uacfc\ubaa9\uba85",
        "\uc124\uba85",
        "\ubd80\uc11c",
        "\uc870\uc9c1",
        "\uacfc\uc81c\uba85",
        "\ud604\uc7ac\uc608\uc0b0\uc561(\uc6d0)",
        base_header,
        "\uc99d\uac10\uc561(\uc6d0)",
        "\uc0b0\ucd9c\ub0b4\uc5ed\uc218",
        "\uc0c1\ud0dc",
        "\uc608\uc0b0\uad6c\ubd84",
        "\uc774\uc6d4\uad6c\ubd84",
    ]
    ws.append(headers)

    for row in rows:
        diff = int(row.current_amount or 0) - int(row.previous_amount or 0)
        ws.append([
            version.year,
            version.round,
            row.subject_type,
            row.jang,
            row.gwan,
            row.hang,
            row.mok,
            row.subject_code,
            row.subject_name,
            row.description,
            row.department_name,
            row.organization_name,
            row.project_name,
            int(row.current_amount or 0),
            int(row.previous_amount or 0),
            diff,
            int(row.detail_count or 0),
            row.status,
            row.budget_category,
            row.carryover_type,
        ])

    ws.freeze_panes = "A2"


def _aggregate_rows(rows: Iterable[_EntryRow], subject_type: str):
    bucket = defaultdict(lambda: {"current": 0, "previous": 0, "detail_count": 0, "entry_count": 0})
    for row in rows:
        if row.subject_type != subject_type:
            continue
        key = (
            row.jang,
            row.gwan,
            row.hang,
            row.mok or row.hang or row.gwan or row.jang,
            row.department_name,
        )
        item = bucket[key]
        item["current"] += int(row.current_amount or 0)
        item["previous"] += int(row.previous_amount or 0)
        item["detail_count"] += int(row.detail_count or 0)
        item["entry_count"] += 1

    return sorted(bucket.items(), key=lambda kv: kv[0])


def _write_summary_sheet(wb: Workbook, version: BudgetVersion, rows: list[_EntryRow], subject_type: str) -> None:
    title = _SHEET_INCOME_SUMMARY if subject_type == "income" else _SHEET_EXPENSE_SUMMARY
    ws = _get_or_create_sheet(wb, title)
    _clear_sheet(ws)

    base_header = _base_amount_header(version)
    ws.append([
        "\uad6c\ubd84",
        "\uc7a5",
        "\uad00",
        "\ud56d",
        "\ubaa9",
        "\ubd80\uc11c",
        "\ud604\uc7ac\uc608\uc0b0\uc561(\uc6d0)",
        base_header,
        "\uc99d\uac10\uc561(\uc6d0)",
        "\uac74\uc218",
        "\uc0b0\ucd9c\ub0b4\uc5ed\uc218",
    ])

    agg = _aggregate_rows(rows, subject_type)
    current_sum = 0
    previous_sum = 0
    count_sum = 0
    detail_sum = 0
    type_label = "\uc218\uc785" if subject_type == "income" else "\uc9c0\ucd9c"

    for key, values in agg:
        jang, gwan, hang, mok, dept = key
        current = int(values["current"] or 0)
        previous = int(values["previous"] or 0)
        diff = current - previous
        entry_count = int(values["entry_count"] or 0)
        detail_count = int(values["detail_count"] or 0)
        ws.append([type_label, jang, gwan, hang, mok, dept, current, previous, diff, entry_count, detail_count])
        current_sum += current
        previous_sum += previous
        count_sum += entry_count
        detail_sum += detail_count

    if not agg:
        ws.append([
            type_label,
            "",
            "",
            "",
            "",
            "",
            0,
            0,
            0,
            0,
            0,
        ])

    ws.append([
        "\ud569\uacc4",
        "",
        "",
        "",
        "",
        "",
        current_sum,
        previous_sum,
        current_sum - previous_sum,
        count_sum,
        detail_sum,
    ])
    ws.freeze_panes = "A2"


def _write_deferred_sheet(wb: Workbook, version: BudgetVersion, row_count: int) -> None:
    ws = _get_or_create_sheet(wb, _SHEET_DEFERRED)
    _clear_sheet(ws)
    ws.append(["\uc2dc\ud2b8", "\ud604\ud669", "\ube44\uace0"])
    ws.append([
        "\uae30\ubcf8\uc7ac\uc0b0\uba85\uc138\uc11c",
        "\uc0d8\ud50c \ub370\uc774\ud130",
        "\ud6c4\uc18d \uad6c\ud604 \uc608\uc815",
    ])
    ws.append([
        "\ubcf4\ud1b5\uc7ac\uc0b0\uba85\uc138\uc11c",
        "\uc0d8\ud50c \ub370\uc774\ud130",
        "\ud6c4\uc18d \uad6c\ud604 \uc608\uc815",
    ])
    ws.append([
        "\ucd1d\uc778\uac74\ube44\uba85\uc138\uc11c",
        "\uc0d8\ud50c \ub370\uc774\ud130",
        "\ud6c4\uc18d \uad6c\ud604 \uc608\uc815",
    ])
    ws.append([])
    ws.append(["\uae30\uc900\uc815\ubcf4", "\uac12"])
    ws.append(["\uc5f0\ub3c4", version.year])
    ws.append(["\ud68c\ucc28", version.round])
    ws.append(["\ub300\uc0c1 \uc5d4\ud2b8\ub9ac \uc218", row_count])


def _write_asset_sample_sheet(wb: Workbook, version: BudgetVersion) -> None:
    ws = _get_or_create_sheet(wb, _SHEET_ASSET_SAMPLE)
    _clear_sheet(ws)

    ws.append([
        "\uad6c\ubd84",
        "\uc790\uc0b0\uba85",
        "\ucde8\ub4dd\uc77c",
        "\ucde8\ub4dd\uac00\uc561(\uc6d0)",
        "\ub0b4\uc6a9\uc5f0\uc218(\ub144)",
        "\ube44\uace0",
    ])
    ws.append([
        "\uae30\ubcf8\uc7ac\uc0b0",
        "\ud1a0\uc9c0(\uc0d8\ud50c)",
        f"{version.year - 1}-01-01",
        100_000_000,
        50,
        "\ud6c4\uc18d \uad6c\ud604 \uc804 \uc0d8\ud50c \ub370\uc774\ud130",
    ])
    ws.append([
        "\ubcf4\ud1b5\uc7ac\uc0b0",
        "\uc5c5\ubb34\uc6a9 \ube44\ud488(\uc0d8\ud50c)",
        f"{version.year}-03-01",
        3_000_000,
        5,
        "\ud6c4\uc18d \uad6c\ud604 \uc804 \uc0d8\ud50c \ub370\uc774\ud130",
    ])
    ws.append([])
    ws.append(["\uae30\uc900\uc5f0\ub3c4", version.year, "", "", "", ""])
    ws.freeze_panes = "A2"


def _write_labor_sample_sheet(wb: Workbook, version: BudgetVersion) -> None:
    ws = _get_or_create_sheet(wb, _SHEET_LABOR_SAMPLE)
    _clear_sheet(ws)

    ws.append([
        "\uad6c\ubd84",
        "\uc778\uc6d0\uc218",
        "\ud3c9\uade0\uc5f0\ubd09(\uc6d0)",
        "\uc778\uac74\ube44(\uc6d0)",
        "\ube44\uace0",
    ])
    ws.append([
        "\uc815\uaddc\uc9c1(\uc0d8\ud50c)",
        10,
        50_000_000,
        "=B2*C2",
        "\ud6c4\uc18d \uad6c\ud604 \uc804 \uc0d8\ud50c \ub370\uc774\ud130",
    ])
    ws.append([
        "\uacc4\uc57d\uc9c1(\uc0d8\ud50c)",
        4,
        32_000_000,
        "=B3*C3",
        "\ud6c4\uc18d \uad6c\ud604 \uc804 \uc0d8\ud50c \ub370\uc774\ud130",
    ])
    ws.append([
        "\ud569\uacc4",
        "=SUM(B2:B3)",
        "",
        "=SUM(D2:D3)",
        f"{version.year}\ub144 \uae30\uc900",
    ])
    ws.freeze_panes = "A2"


def _find_row(ws, column: int, candidates: Iterable[str]) -> int | None:
    key_set = {_normalize_key(item) for item in candidates if item}
    if not key_set:
        return None
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row_idx, column).value
        key = _normalize_key(value)
        if key in key_set:
            return row_idx
    return None


def _subject_totals(rows: list[_EntryRow], subject_type: str) -> dict[str, Any]:
    current_total = 0
    previous_total = 0
    by_department = defaultdict(int)
    for row in rows:
        if row.subject_type != subject_type:
            continue
        current_total += int(row.current_amount or 0)
        previous_total += int(row.previous_amount or 0)
        dept_key = _normalize_key(row.department_name)
        if dept_key:
            by_department[dept_key] += int(row.current_amount or 0)
    return {
        "current_total": current_total,
        "previous_total": previous_total,
        "by_department": by_department,
    }


def _find_sheet_name(wb: Workbook, candidates: Iterable[str]) -> str | None:
    key_set = {_normalize_key(name) for name in candidates if name}
    if not key_set:
        return None
    for name in wb.sheetnames:
        if _normalize_key(name) in key_set:
            return name
    return None


def _override_record(sheet: str, applied: bool, reason: str, updated_cells_count: int) -> dict[str, Any]:
    return {
        "sheet": sheet,
        "applied": bool(applied),
        "reason": reason,
        "updated_cells_count": int(updated_cells_count or 0),
    }


def _resolve_writable_cell(ws, row: int, col: int):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell
    coord = cell.coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            # Non top-left merged cells are read-only and should be skipped.
            if row == merged_range.min_row and col == merged_range.min_col:
                return ws.cell(merged_range.min_row, merged_range.min_col)
            return None
    return None


def _set_cell_value(ws, row: int, col: int, value, *, preserve_formula: bool = True) -> int:
    cell = _resolve_writable_cell(ws, row, col)
    if cell is None:
        return 0
    if preserve_formula and isinstance(cell.value, str) and cell.value.strip().startswith("="):
        return 0
    if cell.value == value:
        return 0
    cell.value = value
    return 1


def _set_cell_ref(ws, ref: str, value, *, preserve_formula: bool = True) -> int:
    cell = ws[ref]
    return _set_cell_value(ws, int(cell.row), int(cell.column), value, preserve_formula=preserve_formula)


def _top_level_organization_names() -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for raw in Organization.objects.filter(parent__isnull=True).order_by("sort_order", "id").values_list("name", flat=True):
        name = _safe_str(raw)
        key = _normalize_key(name)
        if not key or key in seen:
            continue
        names.append(name)
        seen.add(key)
    return names


def _matrix_header_columns(ws, income_header_row: int, expense_header_row: int) -> list[int]:
    max_col = max(int(ws.max_column or 0), 6)
    used_cols: list[int] = []
    for col in range(6, max_col + 1):
        income_val = ws.cell(income_header_row, col).value if income_header_row else None
        expense_val = ws.cell(expense_header_row, col).value if expense_header_row else None
        if _safe_str(income_val) or _safe_str(expense_val):
            used_cols.append(col)
    if not used_cols:
        return [6]
    return list(range(6, max(used_cols) + 1))


def _sync_reference_date_text(ws, year: int) -> int:
    target = f"<기준일 : {int(year) - 1}. 12. 31 >"
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, min(ws.max_column, 12) + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and "기준일" in value:
                return _set_cell_value(ws, row, col, target, preserve_formula=False)
    return 0


def _apply_values(ws, value_by_ref: dict[str, Any], *, preserve_formula: bool = True) -> int:
    updated = 0
    for ref, value in value_by_ref.items():
        updated += _set_cell_ref(ws, ref, value, preserve_formula=preserve_formula)
    return updated


def _clear_refs(ws, refs: Iterable[str], *, preserve_formula: bool = True) -> int:
    updated = 0
    for ref in refs:
        updated += _set_cell_ref(ws, ref, None, preserve_formula=preserve_formula)
    return updated


def _write_matrix_sheet_overrides(wb: Workbook, rows: list[_EntryRow]) -> dict[str, Any]:
    if _SHEET_MATRIX not in wb.sheetnames:
        return _override_record(_SHEET_MATRIX, False, "sheet_not_found", 0)
    ws = wb[_SHEET_MATRIX]
    updated = 0

    income = _subject_totals(rows, "income")
    expense = _subject_totals(rows, "expense")
    income_row = _find_row(ws, 1, ["\uc218\uc785\uacc4"])
    expense_row = _find_row(ws, 1, ["\uc9c0\ucd9c\uacc4"])
    if not income_row and not expense_row:
        return _override_record(_SHEET_MATRIX, False, "summary_rows_not_found", 0)
    income_header_row = max(1, income_row - 1) if income_row else 3
    expense_header_row = max(1, expense_row - 1) if expense_row else 16
    header_cols = _matrix_header_columns(ws, income_header_row, expense_header_row)
    top_org_names = _top_level_organization_names()

    if top_org_names:
        for idx, col in enumerate(header_cols):
            org_name = top_org_names[idx] if idx < len(top_org_names) else ""
            if income_row:
                updated += _set_cell_value(ws, income_header_row, col, org_name, preserve_formula=False)
            if expense_row:
                updated += _set_cell_value(ws, expense_header_row, col, org_name, preserve_formula=True)

    def _dept_key_from_header(row_idx: int, col_idx: int) -> str:
        value = ws.cell(row_idx, col_idx).value
        # Some official templates keep expense header row as formula (=F3 ...).
        if isinstance(value, str) and value.strip().startswith("="):
            return ""
        return _normalize_key(value)

    if income_row:
        updated += _set_cell_value(ws, income_row, 5, _to_thousand_won(income["current_total"]), preserve_formula=True)
        header_row = income_header_row
        for col in header_cols:
            dept_name = _dept_key_from_header(header_row, col)
            if not dept_name:
                continue
            amount = _to_thousand_won(income["by_department"].get(dept_name, 0))
            updated += _set_cell_value(ws, income_row, col, amount, preserve_formula=True)

    if expense_row:
        updated += _set_cell_value(ws, expense_row, 5, _to_thousand_won(expense["current_total"]), preserve_formula=True)
        header_row = expense_header_row
        for col in header_cols:
            dept_name = _dept_key_from_header(header_row, col)
            if not dept_name:
                dept_name = _dept_key_from_header(income_header_row, col)
            if not dept_name:
                continue
            amount = _to_thousand_won(expense["by_department"].get(dept_name, 0))
            updated += _set_cell_value(ws, expense_row, col, amount, preserve_formula=True)

    reason = "ok"
    if not top_org_names:
        reason = "no_root_organizations"
    return _override_record(_SHEET_MATRIX, True, reason, updated)


def _write_total_sheet_overrides(
    wb: Workbook,
    rows: list[_EntryRow],
    *,
    sheet_name: str,
    subject_type: str,
    label: str,
) -> dict[str, Any]:
    if sheet_name not in wb.sheetnames:
        return _override_record(sheet_name, False, "sheet_not_found", 0)
    ws = wb[sheet_name]
    total_row = _find_row(ws, 1, [label])
    if not total_row:
        return _override_record(sheet_name, False, "summary_row_not_found", 0)

    totals = _subject_totals(rows, subject_type)
    current = _to_thousand_won(totals["current_total"])
    previous = _to_thousand_won(totals["previous_total"])
    updated = 0
    updated += _set_cell_value(ws, total_row, 5, current, preserve_formula=True)
    updated += _set_cell_value(ws, total_row, 6, previous, preserve_formula=True)
    updated += _set_cell_value(ws, total_row, 7, current - previous, preserve_formula=True)
    return _override_record(sheet_name, True, "ok", updated)


def _write_basic_asset_sheet_overrides(wb: Workbook, version: BudgetVersion) -> dict[str, Any]:
    sheet_name = _find_sheet_name(wb, [_SHEET_BASIC_ASSET])
    if not sheet_name:
        return _override_record(_SHEET_BASIC_ASSET, False, "sheet_not_found", 0)
    ws = wb[sheet_name]

    updated = 0
    updated += _sync_reference_date_text(ws, version.year)
    updated += _apply_values(
        ws,
        {
            "B5": "현 금",
            "C5": "-",
            "D5": "-",
            "E5": "-",
            "F5": 200_000_000,
            "G5": "설립시 기본재산",
            "B6": "현 금",
            "C6": "-",
            "D6": "-",
            "E6": "-",
            "F6": 6_309_484_579,
            "G6": "시설임대 ·장비활용 수익금 기본재산 ",
            "B7": "토지1",
            "C7": "포항 남 지곡로",
            "D7": 394,
            "E7": 16_729,
            "F7": 1_374_911_200,
            "G7": "출연부지 및 \n매입토지",
            "B8": "토지2",
            "C8": "포항 남 지곡동",
            "D8": 909,
            "E8": 1_809,
            "F8": 220_698_000,
            "B9": "토지3",
            "C9": "포항 남 지곡동",
            "D9": 911,
            "E9": 2_030,
            "F9": 262_367_900,
            "B10": "토지4",
            "C10": "포항 남 지곡동",
            "D10": 918,
            "E10": 304,
            "F10": 9_728_000,
            "B11": "토지5",
            "C11": "포항 남 지곡동",
            "D11": "산116",
            "E11": 121_800,
            "F11": 8_728_331_640,
            "B12": "토지6",
            "C12": "포항 남 지곡동",
            "D12": 905,
            "E12": 160,
            "F12": 39_040_000,
            "B13": "토지7",
            "C13": "포항 남 지곡동",
            "D13": 919,
            "E13": 734,
            "F13": 23_488_000,
            "B14": "토지8",
            "C14": "포항 남 지곡동",
            "D14": "산138",
            "E14": 37_290,
            "F14": 829_148_580,
        },
        preserve_formula=True,
    )
    return _override_record(sheet_name, True, "ok", updated)


def _write_ordinary_asset_sheet_overrides(wb: Workbook, version: BudgetVersion) -> dict[str, Any]:
    sheet_name = _find_sheet_name(wb, [_SHEET_ORDINARY_ASSET])
    if not sheet_name:
        return _override_record(_SHEET_ORDINARY_ASSET, False, "sheet_not_found", 0)
    ws = wb[sheet_name]

    updated = 0
    updated += _set_cell_ref(ws, "F3", f"<기준일 : {version.year - 1}. 12. 31 >", preserve_formula=False)
    updated += _apply_values(
        ws,
        {
            "E5": 5_441_863_488,
            "F5": 5_441_863_488,
            "E6": 25_416_589_448,
            "F6": 25_416_589_448,
            "E8": 2_009_645_325,
            "F8": 1_005_679_095,
            "E10": 51_012_376,
            "F10": 3_727_818,
            "E11": 282_487_776,
            "F11": 46_953_592,
            "E13": 3_108_025_408,
            "F13": 3_108_025_408,
            "E14": 45_187_867_845,
            "F14": 27_170_660_679,
            "E15": 7_145_191_760,
            "F15": 356_781_116,
            "E16": 10_269_138_536,
            "F16": 206_326_444,
            "E17": 1_196_260_304,
            "F17": 155_073_518,
            "E18": 941_795_911,
            "F18": 941_795_911,
        },
        preserve_formula=True,
    )
    return _override_record(sheet_name, True, "ok", updated)


def _write_labor_statement_overrides(wb: Workbook, version: BudgetVersion) -> dict[str, Any]:
    sheet_name = _find_sheet_name(wb, [_SHEET_LABOR_STATEMENT, f"{_SHEET_LABOR_STATEMENT} "])
    if not sheet_name:
        return _override_record(_SHEET_LABOR_STATEMENT, False, "sheet_not_found", 0)
    ws = wb[sheet_name]

    updated = 0
    updated += _sync_reference_date_text(ws, version.year)
    updated += _apply_values(
        ws,
        {
            "E11": 1,
            "F6": 120_315,
            "G6": 114_586,
            "F7": 0,
            "G7": 0,
            "F8": 16_844,
            "G8": 16_523,
            "F9": 11_513,
            "G9": 11_294,
            "F10": 14_073,
            "G10": 10_615,
            "E17": 99,
            "G12": 4_639_948,
            "G13": 1_389_630,
            "G14": 534_385,
            "G15": 540_285,
            "G16": 693_812,
            "G17": 7_798_060,
            "G18": 7_170_495,
            "G19": 780_583,
        },
        preserve_formula=True,
    )
    return _override_record(sheet_name, True, "ok", updated)


def _run_safe_override(sheet_name: str, func) -> dict[str, Any]:
    try:
        result = func()
    except Exception as exc:  # pragma: no cover - defensive for template variance
        logger.warning("template override failed for %s: %s", sheet_name, exc, exc_info=True)
        return _override_record(sheet_name, False, f"error:{type(exc).__name__}", 0)
    if isinstance(result, dict):
        return result
    return _override_record(sheet_name, False, "invalid_override_result", 0)


def _write_template_overrides(wb: Workbook, version: BudgetVersion, rows: list[_EntryRow]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    records.append(_run_safe_override(
        _SHEET_MATRIX,
        lambda: _write_matrix_sheet_overrides(wb, rows),
    ))
    records.append(_run_safe_override(
        _SHEET_INCOME_TOTAL,
        lambda: _write_total_sheet_overrides(
            wb,
            rows,
            sheet_name=_SHEET_INCOME_TOTAL,
            subject_type="income",
            label="\uc218\uc785\uacc4",
        )
    ))
    records.append(_run_safe_override(
        _SHEET_EXPENSE_TOTAL,
        lambda: _write_total_sheet_overrides(
            wb,
            rows,
            sheet_name=_SHEET_EXPENSE_TOTAL,
            subject_type="expense",
            label="\uc9c0\ucd9c\uacc4",
        )
    ))
    records.append(_run_safe_override(
        _SHEET_BASIC_ASSET,
        lambda: _write_basic_asset_sheet_overrides(wb, version),
    ))
    records.append(_run_safe_override(
        _SHEET_ORDINARY_ASSET,
        lambda: _write_ordinary_asset_sheet_overrides(wb, version),
    ))
    records.append(_run_safe_override(
        _SHEET_LABOR_STATEMENT,
        lambda: _write_labor_statement_overrides(wb, version),
    ))
    return records


def _build_file_name(version: BudgetVersion) -> str:
    round_label = "\ubcf8\uc608\uc0b0" if int(version.round or 0) == 0 else f"{int(version.round)}\ucc28\ucd94\uacbd"
    return f"{version.year}_{round_label}_\uc608\uc0b0\uc11c.xlsx"


def _build_content_disposition(file_name: str) -> str:
    ascii_name = file_name.encode("ascii", errors="ignore").decode("ascii")
    ascii_name = "".join(ch if (ch.isalnum() or ch in ("-", "_", ".")) else "_" for ch in ascii_name)
    if not ascii_name:
        ascii_name = "budget_book.xlsx"
    elif not ascii_name.lower().endswith(".xlsx"):
        ascii_name = f"{ascii_name}.xlsx"
    else:
        stem = ascii_name[:-5]
        if not any(ch.isalpha() for ch in stem):
            compact = stem.strip("_") or "export"
            ascii_name = f"budget_book_{compact}.xlsx"
    encoded = quote(file_name)
    return f"attachment; filename={ascii_name}; filename*=UTF-8''{encoded}"


def build_budget_book_file(version: BudgetVersion):
    workbook, template_path = _open_workbook()
    rows = _collect_entry_rows(version)

    template_overrides = _write_template_overrides(workbook, version, rows)
    template_warnings = [
        item for item in template_overrides
        if (not item.get("applied")) or str(item.get("reason")) not in ("ok", "no_root_organizations")
    ]
    _write_seed_sheet(workbook, version, rows)
    _write_summary_sheet(workbook, version, rows, "income")
    _write_summary_sheet(workbook, version, rows, "expense")
    _write_deferred_sheet(workbook, version, len(rows))
    _write_asset_sample_sheet(workbook, version)
    _write_labor_sample_sheet(workbook, version)

    file_name = _build_file_name(version)
    disposition = _build_content_disposition(file_name)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    return (
        buffer.getvalue(),
        file_name,
        disposition,
        {
            "template_path": str(template_path) if template_path else "",
            "row_count": len(rows),
            "template_overrides": template_overrides,
            "template_warning_count": len(template_warnings),
            "template_warnings": template_warnings,
        },
    )
