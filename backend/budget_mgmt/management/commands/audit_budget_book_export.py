from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from budget_mgmt.models import BudgetVersion
from budget_mgmt.services.budget_book_export import build_budget_book_file


DEFAULT_SHEETS = [
    "예산목별조서",
    "수입(총괄)",
    "지출(총괄)",
    "기본재산명세서",
    "보통재산명세서",
    "총인건비명세서",
]

# Intentionally dynamic cell (synced by version year).
DEFAULT_ALLOWED_DIFFS = {
    ("보통재산명세서", "F3"),
}


def _find_sheet(wb, base_name: str):
    for name in wb.sheetnames:
        if str(name).strip() == base_name:
            return wb[name], name
    return None, ""


class Command(BaseCommand):
    help = "Audit exported budget book against template and report cell-level diffs."

    def add_arguments(self, parser):
        parser.add_argument("--version-id", type=int, required=True, help="BudgetVersion id")
        parser.add_argument(
            "--max-diffs-per-sheet",
            type=int,
            default=30,
            help="Max diff rows to emit for each sheet.",
        )
        parser.add_argument(
            "--output-json",
            type=str,
            default="",
            help="Optional JSON output path for audit report.",
        )
        parser.add_argument(
            "--strict",
            action="store_true",
            help="Raise CommandError when unexpected diffs or warnings exist.",
        )

    def handle(self, *args, **options):
        version_id = int(options["version_id"])
        max_diffs = max(1, int(options["max_diffs_per_sheet"] or 30))
        strict = bool(options.get("strict"))

        try:
            version = BudgetVersion.objects.get(id=version_id)
        except BudgetVersion.DoesNotExist as exc:
            raise CommandError(f"BudgetVersion not found: id={version_id}") from exc

        file_bytes, file_name, _disposition, meta = build_budget_book_file(version)
        template_path = str(meta.get("template_path") or "").strip()
        if not template_path:
            raise CommandError("Template path is empty. Set BUDGET_BOOK_TEMPLATE_PATH or default template file.")
        template_file = Path(template_path)
        if not template_file.exists():
            raise CommandError(f"Template file not found: {template_file}")

        report = {
            "version_id": version_id,
            "year": int(version.year),
            "round": int(version.round),
            "file_name": file_name,
            "template_path": str(template_file),
            "template_warning_count": int(meta.get("template_warning_count", 0) or 0),
            "template_warnings": meta.get("template_warnings") or [],
            "sheets": [],
        }

        wb_template = load_workbook(template_file, data_only=False)
        wb_output = load_workbook(BytesIO(file_bytes), data_only=False)

        try:
            for base_name in DEFAULT_SHEETS:
                ws_t, actual_t = _find_sheet(wb_template, base_name)
                ws_o, actual_o = _find_sheet(wb_output, base_name)

                sheet_report = {
                    "base_name": base_name,
                    "template_sheet": actual_t,
                    "output_sheet": actual_o,
                    "status": "ok",
                    "diff_count": 0,
                    "unexpected_diff_count": 0,
                    "diffs": [],
                }

                if ws_t is None or ws_o is None:
                    sheet_report["status"] = "missing_sheet"
                    report["sheets"].append(sheet_report)
                    continue

                max_row = max(int(ws_t.max_row or 0), int(ws_o.max_row or 0))
                max_col = max(int(ws_t.max_column or 0), int(ws_o.max_column or 0))
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        v_t = ws_t.cell(r, c).value
                        v_o = ws_o.cell(r, c).value
                        if v_t == v_o:
                            continue
                        coord = f"{get_column_letter(c)}{r}"
                        allowed = (base_name, coord) in DEFAULT_ALLOWED_DIFFS
                        sheet_report["diff_count"] += 1
                        if not allowed:
                            sheet_report["unexpected_diff_count"] += 1
                        if len(sheet_report["diffs"]) < max_diffs:
                            sheet_report["diffs"].append(
                                {
                                    "cell": coord,
                                    "template": v_t,
                                    "output": v_o,
                                    "allowed": allowed,
                                }
                            )

                if sheet_report["unexpected_diff_count"] > 0:
                    sheet_report["status"] = "unexpected_diffs"
                elif sheet_report["diff_count"] > 0:
                    sheet_report["status"] = "allowed_diffs_only"

                report["sheets"].append(sheet_report)
        finally:
            wb_template.close()
            wb_output.close()

        if options.get("output_json"):
            output_json = Path(str(options["output_json"])).expanduser()
            if not output_json.is_absolute():
                output_json = Path.cwd() / output_json
            output_json.parent.mkdir(parents=True, exist_ok=True)
            output_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
            self.stdout.write(f"Audit JSON written: {output_json}")

        self.stdout.write(self.style.SUCCESS("Budget book export audit completed."))
        self.stdout.write(f"version={report['version_id']} year={report['year']} round={report['round']}")
        self.stdout.write(f"template_warning_count={report['template_warning_count']}")
        for sheet in report["sheets"]:
            self.stdout.write(
                f"[{sheet['base_name']}] status={sheet['status']} "
                f"diff_count={sheet['diff_count']} unexpected={sheet['unexpected_diff_count']}"
            )

        if strict:
            has_warnings = report["template_warning_count"] > 0
            has_unexpected_diffs = any(s["unexpected_diff_count"] > 0 for s in report["sheets"])
            has_missing_sheet = any(s["status"] == "missing_sheet" for s in report["sheets"])
            if has_warnings or has_unexpected_diffs or has_missing_sheet:
                raise CommandError("Strict audit failed: warnings or unexpected diffs detected.")
