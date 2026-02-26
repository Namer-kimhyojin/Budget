from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from .models import Organization, BudgetSubject, BudgetEntry, BudgetDetail, BudgetTransfer, ApprovalLog, Notification, UserProfile, BudgetExecution, SpendingLimitRule, BudgetVersion, EntrustedProject, SubmissionComment, SupportingDocument
from .serializers import *
from .calculation import parse_calc_expression
from django.db import transaction, IntegrityError, DatabaseError
from django.db.models import Sum, Max, Q, F, Prefetch
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.authtoken.models import Token
from rest_framework.exceptions import ValidationError as DRFValidationError, PermissionDenied, APIException
from .erpnext_client import get_erpnext_client, ERPNextError
from .audit import write_audit_log
from pathlib import Path
import json
import re
import logging
import uuid

USERNAME_RE = re.compile(r'^[a-zA-Z0-9_.-]{4,50}$')
logger = logging.getLogger(__name__)

# Role normalization map from legacy role labels
_ROLE_NORMALIZE = {
    'REQUESTOR': 'STAFF',   # legacy REQUESTOR -> STAFF
    'REVIEWER':  'MANAGER', # legacy REVIEWER -> MANAGER
}

def _normalize_role(role):
    """Normalize legacy role labels to current role labels."""
    return _ROLE_NORMALIZE.get(role, role)


def _role(user_or_request):
    """Safely extract role string from request/user object."""
    user = getattr(user_or_request, 'user', user_or_request)
    profile = getattr(user, 'profile', None)
    role = getattr(profile, 'role', None)
    return _normalize_role(role)


class ConflictError(APIException):
    status_code = status.HTTP_409_CONFLICT
    default_detail = 'Resource was modified by another user.'
    default_code = 'conflict'


def _user_role(user_or_request):
    return _normalize_role(_role(user_or_request)) or 'STAFF'


def _is_admin(user_or_request):
    return _user_role(user_or_request) == 'ADMIN'


def _is_manager_or_admin(user_or_request):
    return _user_role(user_or_request) in ('MANAGER', 'ADMIN')


def _can_write_budget_data(user_or_request):
    return _user_role(user_or_request) in ('STAFF', 'MANAGER', 'ADMIN')


def _scope_org_ids_for_user(user_or_request):
    """
    Returns:
      - None: unrestricted scope (ADMIN)
      - set[int]: allowed organization ids
    """
    if _is_admin(user_or_request):
        return None
    user = getattr(user_or_request, 'user', user_or_request)
    profile = getattr(user, 'profile', None)
    if profile is None:
        return set()

    scope = set()
    if getattr(profile, 'team_id', None):
        scope.add(int(profile.team_id))
        return scope

    org_id = getattr(profile, 'organization_id', None)
    if org_id is None:
        return scope

    org = Organization.objects.filter(id=org_id).only('id', 'org_type', 'parent_id').first()
    if org is None:
        return scope
    if _is_team_organization(org):
        scope.add(int(org.id))
        return scope

    scope.add(int(org.id))
    scope.update(
        Organization.objects.filter(parent_id=org.id).values_list('id', flat=True)
    )
    return scope


def _org_in_scope(user_or_request, organization_id):
    allowed = _scope_org_ids_for_user(user_or_request)
    if allowed is None:
        return True
    if organization_id in (None, '', 0, '0'):
        return False
    try:
        org_id = int(organization_id)
    except (TypeError, ValueError):
        return False
    return org_id in allowed


def _scope_queryset_by_org(queryset, user_or_request, *, org_field='organization_id'):
    allowed = _scope_org_ids_for_user(user_or_request)
    if allowed is None:
        return queryset
    if not allowed:
        return queryset.none()
    return queryset.filter(**{f'{org_field}__in': list(allowed)})


def _require_roles_response(request, allowed_roles, *, message='No permission'):
    role = _user_role(request)
    if role in allowed_roles:
        return None
    return Response({'error': message}, status=status.HTTP_403_FORBIDDEN)


def _is_team_organization(org):
    if not org:
        return False
    return (str(org.org_type or '').lower() == 'team') or bool(org.parent_id)


def _resolve_user_org_team(organization_id, team_id, *, allow_auto_department=False):
    organization = None
    if organization_id not in (None, '', 0, '0'):
        organization = Organization.objects.filter(id=organization_id).first()
        if organization is None:
            return None, None, 'organization not found'
        if _is_team_organization(organization):
            return None, None, 'organization must be a department'

    team = None
    if team_id not in (None, '', 0, '0'):
        team = Organization.objects.select_related('parent').filter(id=team_id).first()
        if team is None:
            return None, None, 'team not found'
        if not _is_team_organization(team):
            return None, None, 'team must be a team organization'
        if team.parent_id is None:
            return None, None, 'team must have a parent department'
        if organization is None:
            if allow_auto_department:
                organization = team.parent
            else:
                return None, None, 'organization is required when team is set'
        elif team.parent_id != organization.id:
            return None, None, 'team does not belong to organization'

    return organization, team, None


def _serialize_user_context(user):
    """Serialize user context. Assumes profile exists or can be created once during signup."""
    profile = getattr(user, 'profile', None)
    if profile is None:
        default_role = 'ADMIN' if (getattr(user, 'is_superuser', False) or not UserProfile.objects.exists()) else 'STAFF'
        profile, created = UserProfile.objects.get_or_create(user=user, defaults={'role': default_role})
        if created:
            logger.warning('Created missing UserProfile for user id=%s username=%s role=%s', user.id, user.username, default_role)
    return {
        'user': {
            'id': user.id,
            'username': user.username,
            'name': user.first_name or user.username,
            'email': user.email,
        },
        'profile': {
            'role': profile.role,
            'organization': profile.organization_id,
            'organization_name': profile.organization.name if profile.organization else None,
            'team': profile.team_id,
            'team_name': profile.team.name if profile.team else None,
        },
    }


def _log_auth_event(request, *, action, user=None, reason='', status_code=None, metadata=None):
    write_audit_log(
        request=request,
        actor=user,
        log_type='AUTH',
        action=action,
        from_status='AUTH',
        to_status=action,
        reason=reason,
        resource_type='auth',
        resource_id=getattr(user, 'id', None),
        status_code=status_code,
        metadata=metadata or {},
    )


class AuthSignUpView(APIView):
    permission_classes = [permissions.AllowAny]

    @transaction.atomic
    def post(self, request):
        username = (request.data.get('username') or '').strip()
        password = request.data.get('password') or ''
        name = (request.data.get('name') or '').strip()
        email = (request.data.get('email') or '').strip()
        organization_id = request.data.get('organization')
        team_id = request.data.get('team')

        if not username or not password:
            return Response({'error': 'username and password are required'}, status=status.HTTP_400_BAD_REQUEST)
        if not USERNAME_RE.match(username):
            return Response({'error': 'username must be 4-50 chars and contain only letters, numbers, _, -, .'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=username).exists():
            return Response({'error': 'username already exists'}, status=status.HTTP_400_BAD_REQUEST)
        if not email:
            return Response({'error': 'email is required'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(email__iexact=email).exists():
            return Response({'error': 'email already exists'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_password(password)
        except DjangoValidationError as exc:
            return Response({'error': 'invalid password', 'details': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        organization, team, org_error = _resolve_user_org_team(
            organization_id,
            team_id,
            allow_auto_department=True,
        )
        if org_error:
            return Response({'error': org_error}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=name,
            email=email,
        )
        default_role = 'ADMIN' if not UserProfile.objects.exists() else 'STAFF'
        UserProfile.objects.create(user=user, organization=organization, team=team, role=default_role)
        token, _ = Token.objects.get_or_create(user=user)
        _log_auth_event(
            request,
            action='SIGNUP',
            user=user,
            reason='signup success',
            status_code=status.HTTP_201_CREATED,
            metadata={'username': user.username, 'email': user.email},
        )
        return Response({'token': token.key, **_serialize_user_context(user)}, status=status.HTTP_201_CREATED)


class AuthLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        identifier = (request.data.get('username') or '').strip()
        password = request.data.get('password') or ''
        if not identifier or not password:
            _log_auth_event(
                request,
                action='LOGIN_FAILED',
                reason='missing credentials',
                status_code=status.HTTP_400_BAD_REQUEST,
                metadata={'identifier': identifier},
            )
            return Response({'error': 'username and password are required'}, status=status.HTTP_400_BAD_REQUEST)

        login_username = identifier
        if '@' in identifier:
            target = User.objects.filter(email__iexact=identifier).first()
            if target:
                login_username = target.username

        user = authenticate(username=login_username, password=password)
        if user is None:
            _log_auth_event(
                request,
                action='LOGIN_FAILED',
                reason='invalid credentials',
                status_code=status.HTTP_401_UNAUTHORIZED,
                metadata={'identifier': identifier},
            )
            return Response({'error': 'invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        Token.objects.filter(user=user).delete()
        token = Token.objects.create(user=user)
        _log_auth_event(
            request,
            action='LOGIN',
            user=user,
            reason='login success',
            status_code=status.HTTP_200_OK,
            metadata={'identifier': identifier},
        )
        return Response({'token': token.key, **_serialize_user_context(user)})


class AuthLogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        _log_auth_event(
            request,
            action='LOGOUT',
            user=request.user,
            reason='logout',
            status_code=status.HTTP_200_OK,
        )
        Token.objects.filter(user=request.user).delete()
        return Response({'status': 'logged_out'})


class AuthMeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return Response(_serialize_user_context(request.user))


class AuthWithdrawView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        password = request.data.get('password') or ''
        if not request.user.check_password(password):
            return Response({'error': 'password mismatch'}, status=status.HTTP_400_BAD_REQUEST)
        user = request.user
        Token.objects.filter(user=user).delete()
        user.delete()
        return Response({'status': 'withdrawn'})


class AuthFindIdView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        name = (request.data.get('name') or '').strip()
        email = (request.data.get('email') or '').strip()
        if not email:
            return Response({'error': 'email is required'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = User.objects.filter(email__iexact=email)
        if name:
            queryset = queryset.filter(first_name__iexact=name)

        def mask_username(value):
            if len(value) <= 2:
                return value[0] + '*'
            return value[:2] + ('*' * max(1, len(value) - 3)) + value[-1]

        hints = [mask_username(user.username) for user in queryset[:5]]
        return Response({'count': len(hints), 'username_hints': hints})


class AuthAssignRoleView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        admin_profile = getattr(request.user, 'profile', None)
        if getattr(admin_profile, 'role', None) != 'ADMIN':
            return Response({'error': 'admin role required'}, status=status.HTTP_403_FORBIDDEN)

        user_id = request.data.get('user_id')
        role = request.data.get('role')
        organization_id = request.data.get('organization')
        team_id = request.data.get('team')

        if role not in ('STAFF', 'MANAGER', 'ADMIN', 'ORG_VIEWER'):
            return Response({'error': 'invalid role'}, status=status.HTTP_400_BAD_REQUEST)

        target_user = User.objects.filter(id=user_id).first()
        if target_user is None:
            return Response({'error': 'user not found'}, status=status.HTTP_404_NOT_FOUND)

        profile, _ = UserProfile.objects.get_or_create(user=target_user, defaults={'role': 'STAFF'})
        profile.role = role
        organization, team, org_error = _resolve_user_org_team(
            organization_id,
            team_id,
            allow_auto_department=True,
        )
        if org_error:
            return Response({'error': org_error}, status=status.HTTP_400_BAD_REQUEST)
        profile.organization = organization
        profile.team = team
        profile.save()
        write_audit_log(
            request=request,
            actor=request.user,
            log_type='CRUD',
            action='UPDATE',
            from_status='API',
            to_status='UPDATE',
            reason='assign role',
            resource_type='user',
            resource_id=target_user.id,
            status_code=status.HTTP_200_OK,
            metadata={
                'target_username': target_user.username,
                'role': role,
                'organization': organization.id if organization else None,
                'team': team.id if team else None,
            },
        )
        return Response({'status': 'updated', **_serialize_user_context(target_user)})


class AuthAdminUsersView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _check_admin(self, request):
        profile = getattr(request.user, 'profile', None)
        return getattr(profile, 'role', None) == 'ADMIN'

    def get(self, request):
        if not self._check_admin(request):
            return Response({'error': 'admin role required'}, status=status.HTTP_403_FORBIDDEN)
        users = User.objects.select_related('profile__organization', 'profile__team').order_by('id')
        payload = []
        for user in users:
            profile = getattr(user, 'profile', None)
            if not profile:
                # Skip when profile is missing (rare in normal flow)
                continue
            payload.append({
                'id': user.id,
                'username': user.username,
                'name': user.first_name,
                'email': user.email,
                'is_active': user.is_active,
                'is_superuser': user.is_superuser,
                'role': profile.role,
                'organization': profile.organization_id,
                'organization_name': profile.organization.name if profile.organization else None,
                'team': profile.team_id,
                'team_name': profile.team.name if profile.team else None,
                'last_login': user.last_login,
                'date_joined': user.date_joined,
            })
        return Response(payload)

    @transaction.atomic
    def post(self, request):
        if not self._check_admin(request):
            return Response({'error': 'admin role required'}, status=status.HTTP_403_FORBIDDEN)

        username = (request.data.get('username') or '').strip()
        password = request.data.get('password') or ''
        name = (request.data.get('name') or '').strip()
        email = (request.data.get('email') or '').strip()
        role = request.data.get('role') or 'STAFF'
        organization_id = request.data.get('organization')
        team_id = request.data.get('team')

        if not username or not password:
            return Response({'error': 'username and password are required'}, status=status.HTTP_400_BAD_REQUEST)
        if not USERNAME_RE.match(username):
            return Response({'error': 'invalid username format'}, status=status.HTTP_400_BAD_REQUEST)
        if role not in ('STAFF', 'MANAGER', 'ADMIN', 'ORG_VIEWER'):
            return Response({'error': 'invalid role'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=username).exists():
            return Response({'error': 'username already exists'}, status=status.HTTP_400_BAD_REQUEST)
        if email and User.objects.filter(email__iexact=email).exists():
            return Response({'error': 'email already exists'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_password(password)
        except DjangoValidationError as exc:
            return Response({'error': 'invalid password', 'details': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        organization, team, org_error = _resolve_user_org_team(
            organization_id,
            team_id,
            allow_auto_department=True,
        )
        if org_error:
            return Response({'error': org_error}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=name,
            email=email,
            is_active=True,
        )
        UserProfile.objects.create(user=user, role=role, organization=organization, team=team)
        write_audit_log(
            request=request,
            actor=request.user,
            log_type='CRUD',
            action='CREATE',
            from_status='API',
            to_status='CREATE',
            reason='admin created user',
            resource_type='user',
            resource_id=user.id,
            status_code=status.HTTP_201_CREATED,
            metadata={
                'target_username': user.username,
                'role': role,
                'organization': organization.id if organization else None,
                'team': team.id if team else None,
            },
        )
        return Response({'status': 'created', **_serialize_user_context(user)}, status=status.HTTP_201_CREATED)


class AuthAdminUserDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _check_admin(self, request):
        profile = getattr(request.user, 'profile', None)
        return getattr(profile, 'role', None) == 'ADMIN'

    @transaction.atomic
    def patch(self, request, user_id):
        if not self._check_admin(request):
            return Response({'error': 'admin role required'}, status=status.HTTP_403_FORBIDDEN)

        user = User.objects.filter(id=user_id).first()
        if user is None:
            return Response({'error': 'user not found'}, status=status.HTTP_404_NOT_FOUND)

        username = request.data.get('username')
        name = request.data.get('name')
        email = request.data.get('email')
        is_active = request.data.get('is_active')
        role = request.data.get('role')
        organization_id = request.data.get('organization')
        team_id = request.data.get('team')
        reset_password = request.data.get('reset_password')

        if username is not None:
            username = username.strip()
            if not USERNAME_RE.match(username):
                return Response({'error': 'invalid username format'}, status=status.HTTP_400_BAD_REQUEST)
            if User.objects.exclude(id=user.id).filter(username=username).exists():
                return Response({'error': 'username already exists'}, status=status.HTTP_400_BAD_REQUEST)
            user.username = username
        if name is not None:
            user.first_name = name.strip()
        if email is not None:
            email = email.strip()
            if email and User.objects.exclude(id=user.id).filter(email__iexact=email).exists():
                return Response({'error': 'email already exists'}, status=status.HTTP_400_BAD_REQUEST)
            user.email = email
        if is_active is not None:
            user.is_active = bool(is_active)
        if reset_password:
            try:
                validate_password(reset_password, user=user)
            except DjangoValidationError as exc:
                return Response({'error': 'invalid password', 'details': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)
            user.set_password(reset_password)
            Token.objects.filter(user=user).delete()
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user, defaults={'role': 'STAFF'})
        if role is not None:
            if role not in ('STAFF', 'MANAGER', 'ADMIN', 'ORG_VIEWER'):
                return Response({'error': 'invalid role'}, status=status.HTTP_400_BAD_REQUEST)
            profile.role = role
        if 'organization' in request.data or 'team' in request.data:
            candidate_org_id = organization_id if 'organization' in request.data else profile.organization_id
            candidate_team_id = team_id if 'team' in request.data else profile.team_id

            # Keep org/team consistent even for partial updates.
            if 'organization' in request.data and 'team' not in request.data:
                candidate_team_id = None
            if 'team' in request.data and 'organization' not in request.data and candidate_team_id not in (None, '', 0, '0'):
                candidate_org_id = None

            organization, team, org_error = _resolve_user_org_team(
                candidate_org_id,
                candidate_team_id,
                allow_auto_department=True,
            )
            if org_error:
                return Response({'error': org_error}, status=status.HTTP_400_BAD_REQUEST)
            profile.organization = organization
            profile.team = team

        profile.save()
        write_audit_log(
            request=request,
            actor=request.user,
            log_type='CRUD',
            action='UPDATE',
            from_status='API',
            to_status='UPDATE',
            reason='admin updated user',
            resource_type='user',
            resource_id=user.id,
            status_code=status.HTTP_200_OK,
            metadata={
                'target_username': user.username,
                'fields': sorted(list(request.data.keys())),
            },
        )

        return Response({'status': 'updated', **_serialize_user_context(user)})

    @transaction.atomic
    def delete(self, request, user_id):
        if not self._check_admin(request):
            return Response({'error': 'admin role required'}, status=status.HTTP_403_FORBIDDEN)

        user = User.objects.filter(id=user_id).first()
        if user is None:
            return Response({'error': 'user not found'}, status=status.HTTP_404_NOT_FOUND)
        if user.id == request.user.id:
            return Response({'error': 'cannot delete yourself'}, status=status.HTTP_400_BAD_REQUEST)
        target_username = user.username
        Token.objects.filter(user=user).delete()
        user.delete()
        write_audit_log(
            request=request,
            actor=request.user,
            log_type='CRUD',
            action='DELETE',
            from_status='API',
            to_status='DELETE',
            reason='admin deleted user',
            resource_type='user',
            resource_id=user_id,
            status_code=status.HTTP_200_OK,
            metadata={'target_username': target_username},
        )
        return Response({'status': 'deleted'})


class AuthChangePasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        current_password = request.data.get('current_password') or ''
        new_password = request.data.get('new_password') or ''
        if not request.user.check_password(current_password):
            return Response({'error': 'current password mismatch'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            validate_password(new_password, user=request.user)
        except DjangoValidationError as exc:
            return Response({'error': 'invalid new password', 'details': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        request.user.set_password(new_password)
        request.user.save()
        Token.objects.filter(user=request.user).delete()
        token = Token.objects.create(user=request.user)
        return Response({'status': 'password_changed', 'token': token.key})


class AuthPasswordPolicyView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        return Response({
            'username_rule': '4-50 chars, letters/numbers/_/./-',
            'password_rule': 'Use at least 8 chars; avoid common/numeric-only passwords.',
            'lock_policy': None,
            'login_supports': ['username', 'email'],
        })

class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.select_related('parent').all().order_by('sort_order', 'id')
    serializer_class = OrganizationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_admin(self, request):
        denied = _require_roles_response(
            request,
            {'ADMIN'},
            message='Organization management is ADMIN-only.',
        )
        if denied is not None:
            return denied
        return None

    def create(self, request, *args, **kwargs):
        denied = self._ensure_admin(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_admin(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._ensure_admin(request)
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Organization delete is ADMIN-only. Reject when child orgs or linked budget data exist."""
        denied = self._ensure_admin(request)
        if denied is not None:
            return denied
        org = self.get_object()
        # Check child organizations first
        if Organization.objects.filter(parent=org).exists():
            return Response({'error': 'Cannot delete organization with child organizations. Delete children first.'}, status=status.HTTP_409_CONFLICT)
        # Check linked budget entries
        if BudgetEntry.objects.filter(organization=org).exists():
            return Response({'error': 'Cannot delete organization with linked budget data.'}, status=status.HTTP_409_CONFLICT)
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        org_type = self.request.query_params.get('org_type')
        parent = self.request.query_params.get('parent')
        keyword = self.request.query_params.get('q')

        if org_type:
            queryset = queryset.filter(org_type=org_type)
        if parent is not None:
            if str(parent).lower() in ('null', 'none', ''):
                queryset = queryset.filter(parent__isnull=True)
            elif str(parent).isdigit():
                queryset = queryset.filter(parent_id=parent)
        if keyword:
            queryset = queryset.filter(Q(name__icontains=keyword) | Q(code__icontains=keyword))
        return queryset

    @action(detail=False, methods=['post'], url_path='reorder')
    @transaction.atomic
    def reorder(self, request):
        """
        ordered_ids: [id1, id2, id3, ...] in desired order
        Update sort_order of provided ids to 0,1,2...
        """
        denied = self._ensure_admin(request)
        if denied is not None:
            return denied
        ordered_ids = request.data.get('ordered_ids', [])
        if not ordered_ids or not isinstance(ordered_ids, list):
            return Response({'error': 'ordered_ids array is required.'}, status=status.HTTP_400_BAD_REQUEST)
        objs = Organization.objects.filter(id__in=ordered_ids)
        id_map = {o.id: o for o in objs}
        updated = []
        for idx, oid in enumerate(ordered_ids):
            obj = id_map.get(int(oid))
            if obj and obj.sort_order != idx:
                obj.sort_order = idx
                updated.append(obj)
        if updated:
            Organization.objects.bulk_update(updated, ['sort_order'])
        return Response({'updated': len(updated)})

class BudgetSubjectViewSet(viewsets.ModelViewSet):
    queryset = BudgetSubject.objects.select_related('parent').all().order_by('sort_order', 'id')
    serializer_class = BudgetSubjectSerializer

    _CODE_RE = re.compile(r'^[0-9A-Z]{4}$')

    def _ensure_subject_editor(self, request):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Subject management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return None

    def create(self, request, *args, **kwargs):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def _norm_code(self, code):
        return (str(code or '').strip().upper())

    def _token_is_valid(self, token):
        return bool(token) and token.isalnum()

    def _validate_code_rule(self, state, by_id):
        code = self._norm_code(state['code'])
        level = int(state['level'])
        parent_id = state.get('parent')
        if not self._CODE_RE.match(code):
            return f"[{state['id']}] Code must be 4 alphanumeric characters."

        if level == 1:
            if code[1:] != '000' or not self._token_is_valid(code[0]):
                return f"[{state['id']}] Level-1 code must match X000 format. (X: 0-9/A-Z)"
            return None

        if parent_id is None:
            return f"[{state['id']}] Level {level} item requires a parent."
        parent = by_id.get(parent_id)
        if parent is None:
            return f"[{state['id']}] Parent item ({parent_id}) was not found."
        parent_code = self._norm_code(parent['code'])

        if level == 2:
            if code[2:] != '00' or not self._token_is_valid(code[1]):
                return f"[{state['id']}] Level-2 code must match XX00 format. (2nd char: 0-9/A-Z)"
            if code[0] != parent_code[0]:
                return f"[{state['id']}] First char of level-2 code must match parent level-1 code."
            return None

        if level == 3:
            if code[3] != '0' or not self._token_is_valid(code[2]):
                return f"[{state['id']}] Level-3 code must match XXX0 format. (3rd char: 0-9/A-Z)"
            if code[:2] != parent_code[:2]:
                return f"[{state['id']}] First two chars of level-3 code must match parent level-2 code."
            return None

        if level == 4:
            if not self._token_is_valid(code[3]):
                return f"[{state['id']}] Last char of level-4 code must be 0-9/A-Z."
            if code[:3] != parent_code[:3]:
                return f"[{state['id']}] First three chars of level-4 code must match parent level-3 code."
            return None

        return f"[{state['id']}] level must be in range 1..4."

    def _default_subject_json_path(self):
        return Path(settings.BASE_DIR) / 'data' / 'budget_account_system_v1.json'

    def _load_default_subject_records(self, subject_type):
        target_types = {'income', 'expense'} if subject_type == 'all' else {subject_type}
        file_path = self._default_subject_json_path()
        if not file_path.exists():
            raise FileNotFoundError(str(file_path))

        with file_path.open('r', encoding='utf-8') as f:
            payload = json.load(f)

        system = payload.get('budget_account_system', payload)
        income_root = system.get('income_budget') or {}
        expense_root = system.get('expense_budget') or {}

        records = []

        def walk(node, node_subject_type, level, parent_code, sort_order):
            if node_subject_type not in target_types:
                return
            code = str(node.get('code') or '').strip()
            name = str(node.get('name') or '').strip()
            description = str(node.get('description') or '').strip()
            if not code or not name:
                return
            if level < 1 or level > 4:
                return

            records.append({
                'code': code,
                'name': name,
                'description': description,
                'level': level,
                'parent_code': parent_code,
                'subject_type': node_subject_type,
                'sort_order': sort_order,
            })

            children = node.get('children') or []
            for idx, child in enumerate(children):
                walk(child, node_subject_type, level + 1, code, idx)

        for idx, child in enumerate(income_root.get('children') or []):
            walk(child, 'income', 1, None, idx)
        for idx, child in enumerate(expense_root.get('children') or []):
            walk(child, 'expense', 1, None, idx)

        return records

    @action(detail=False, methods=['post'], url_path='restore-defaults')
    @transaction.atomic
    def restore_defaults(self, request):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        subject_type = str(request.data.get('subject_type') or '').strip().lower()
        if subject_type not in ('income', 'expense', 'all'):
            return Response({'error': 'subject_type must be one of income, expense, all'}, status=status.HTTP_400_BAD_REQUEST)

        force = str(request.data.get('force') or '').strip().lower() in ('1', 'true', 'yes', 'y')
        target_types = {'income', 'expense'} if subject_type == 'all' else {subject_type}

        try:
            records = self._load_default_subject_records(subject_type)
        except FileNotFoundError as e:
            return Response({'error': f'default subject file not found: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except json.JSONDecodeError as e:
            return Response({'error': f'default subject json parse failed: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if not records:
            return Response({'error': 'no default records found for target type'}, status=status.HTTP_400_BAD_REQUEST)

        incoming_codes = {r['code'] for r in records}

        created = 0
        updated = 0
        deleted = 0
        skipped = 0

        if force:
            BudgetEntry.objects.filter(subject__subject_type__in=target_types).delete()
        protected_subject_ids = set(
            BudgetEntry.objects.filter(subject__subject_type__in=target_types)
            .values_list('subject_id', flat=True)
        )

        records_sorted = sorted(records, key=lambda rec: (rec['level'], rec['sort_order']))
        code_to_obj = {obj.code: obj for obj in BudgetSubject.objects.all().select_related('parent')}

        for rec in records_sorted:
            parent_obj = code_to_obj.get(rec['parent_code']) if rec['parent_code'] else None
            defaults = {
                'name': rec['name'],
                'description': rec['description'],
                'level': rec['level'],
                'parent': parent_obj,
                'subject_type': rec['subject_type'],
                'sort_order': rec['sort_order'],
            }
            obj, was_created = BudgetSubject.objects.update_or_create(code=rec['code'], defaults=defaults)
            code_to_obj[rec['code']] = obj
            if was_created:
                created += 1
            else:
                updated += 1

        stale_subjects = list(
            BudgetSubject.objects
            .filter(subject_type__in=target_types)
            .exclude(code__in=incoming_codes)
            .order_by('-level', '-id')
        )
        for subject in stale_subjects:
            if subject.id in protected_subject_ids:
                skipped += 1
                continue
            try:
                subject.delete()
                deleted += 1
            except Exception:
                skipped += 1

        return Response({
            'status': 'ok',
            'subject_type': subject_type,
            'force': force,
            'created': created,
            'updated': updated,
            'deleted': deleted,
            'skipped': skipped,
            'total': BudgetSubject.objects.filter(subject_type__in=target_types).count(),
        })

    @action(detail=False, methods=['post'], url_path='bulk-update-tree')
    @transaction.atomic
    def bulk_update_tree(self, request):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        updates = request.data.get('updates')
        if not isinstance(updates, list) or len(updates) == 0:
            return Response({'error': 'updates array is required.'}, status=status.HTTP_400_BAD_REQUEST)

        subjects = BudgetSubject.objects.select_related('parent').all()
        subject_by_id_obj = {s.id: s for s in subjects}
        final_by_id = {
            s.id: {
                'id': s.id,
                'code': self._norm_code(s.code),
                'name': s.name,
                'level': int(s.level),
                'parent': s.parent_id,
                'subject_type': s.subject_type,
            }
            for s in subjects
        }

        target_ids = []
        for row in updates:
            if not isinstance(row, dict):
                return Response({'error': 'Each update item must be an object.'}, status=status.HTTP_400_BAD_REQUEST)
            if 'id' not in row:
                return Response({'error': 'Each update item requires id.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                sid = int(row.get('id'))
            except (TypeError, ValueError):
                return Response({'error': f'invalid id: {row.get("id")}'}, status=status.HTTP_400_BAD_REQUEST)
            if sid not in final_by_id:
                return Response({'error': f'Unknown subject id: {sid}'}, status=status.HTTP_400_BAD_REQUEST)
            target_ids.append(sid)

            current = final_by_id[sid]
            if 'code' in row:
                code = self._norm_code(row.get('code'))
                if not code:
                    return Response({'error': f'[{sid}] code cannot be empty.'}, status=status.HTTP_400_BAD_REQUEST)
                if len(code) > 20:
                    return Response({'error': f'[{sid}] code cannot exceed 20 chars.'}, status=status.HTTP_400_BAD_REQUEST)
                current['code'] = code
            if 'name' in row:
                name = str(row.get('name') or '').strip()
                if not name:
                    return Response({'error': f'[{sid}] name cannot be empty.'}, status=status.HTTP_400_BAD_REQUEST)
                current['name'] = name
            if 'level' in row:
                try:
                    current['level'] = int(row.get('level'))
                except (TypeError, ValueError):
                    return Response({'error': f'[{sid}] invalid level value.'}, status=status.HTTP_400_BAD_REQUEST)
            if 'parent' in row:
                parent_value = row.get('parent')
                if parent_value in (None, '', 'null', 'None'):
                    current['parent'] = None
                else:
                    try:
                        current['parent'] = int(parent_value)
                    except (TypeError, ValueError):
                        return Response({'error': f'[{sid}] invalid parent value.'}, status=status.HTTP_400_BAD_REQUEST)

        # Parent reference + level/type consistency
        for sid, state in final_by_id.items():
            parent_id = state.get('parent')
            if parent_id is None:
                if int(state['level']) != 1:
                    return Response({'error': f'[{sid}] level must be 1 when parent is null.'}, status=status.HTTP_400_BAD_REQUEST)
                continue
            parent = final_by_id.get(parent_id)
            if parent is None:
                return Response({'error': f'[{sid}] parent item ({parent_id}) does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
            if parent_id == sid:
                return Response({'error': f'[{sid}] cannot set self as parent.'}, status=status.HTTP_400_BAD_REQUEST)
            expected_level = int(parent['level']) + 1
            if int(state['level']) != expected_level:
                return Response({'error': f'[{sid}] level must be parent level + 1.'}, status=status.HTTP_400_BAD_REQUEST)
            if state['subject_type'] != parent['subject_type']:
                return Response({'error': f'[{sid}] subject_type must match parent.'}, status=status.HTTP_400_BAD_REQUEST)
            if int(state['level']) < 1 or int(state['level']) > 4:
                return Response({'error': f'[{sid}] level must be in range 1..4.'}, status=status.HTTP_400_BAD_REQUEST)

        # Cycle detection on final graph
        for sid in final_by_id.keys():
            seen = set()
            cursor = sid
            while cursor is not None:
                if cursor in seen:
                    return Response({'error': f'Cycle detected. (subject id: {sid})'}, status=status.HTTP_400_BAD_REQUEST)
                seen.add(cursor)
                cursor_state = final_by_id.get(cursor)
                cursor = cursor_state.get('parent') if cursor_state else None

        # Global unique code check
        code_owner = {}
        for sid, state in final_by_id.items():
            code = self._norm_code(state['code'])
            owner = code_owner.get(code)
            if owner is not None and owner != sid:
                return Response({'error': f'Duplicate code: {code}'}, status=status.HTTP_400_BAD_REQUEST)
            code_owner[code] = sid

        # Code format rule validation
        for sid, state in final_by_id.items():
            msg = self._validate_code_rule(state, final_by_id)
            if msg:
                return Response({'error': msg}, status=status.HTTP_400_BAD_REQUEST)

        changed_ids = []
        code_changed_ids = []
        for sid, state in final_by_id.items():
            origin = subject_by_id_obj[sid]
            has_change = (
                self._norm_code(origin.code) != self._norm_code(state['code']) or
                str(origin.name) != str(state['name']) or
                int(origin.level) != int(state['level']) or
                (origin.parent_id != state['parent'])
            )
            if has_change:
                changed_ids.append(sid)
                if self._norm_code(origin.code) != self._norm_code(state['code']):
                    code_changed_ids.append(sid)

        if not changed_ids:
            return Response({'updated_count': 0, 'subjects': []}, status=status.HTTP_200_OK)

        # Temporary code detour to avoid unique collisions (swap, rotate, etc.)
        tmp_prefix = f"TMP{uuid.uuid4().hex[:8]}_"
        for idx, sid in enumerate(code_changed_ids):
            obj = subject_by_id_obj[sid]
            obj.code = f"{tmp_prefix}{idx}"
            obj.save(update_fields=['code'])

        updated_subjects = []
        for sid in changed_ids:
            obj = subject_by_id_obj[sid]
            state = final_by_id[sid]
            obj.code = self._norm_code(state['code'])
            obj.name = state['name']
            obj.level = int(state['level'])
            obj.parent_id = state['parent']
            obj.save(update_fields=['code', 'name', 'level', 'parent'])
            updated_subjects.append(obj)

        payload = BudgetSubjectSerializer(updated_subjects, many=True).data
        return Response({'updated_count': len(updated_subjects), 'subjects': payload}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'], url_path='force-delete')
    @transaction.atomic
    def force_delete(self, request, pk=None):
        """
        Delete subject and all descendants.
        If any linked BudgetEntry exists, deletion is rejected.
        BudgetEntry rows are not deleted by this API.
        """
        role = _normalize_role(getattr(getattr(request.user, 'profile', None), 'role', None))
        if role not in ('MANAGER', 'ADMIN'):
            return Response({'error': 'Subject delete is allowed only for MANAGER or ADMIN.'}, status=status.HTTP_403_FORBIDDEN)
        subject = self.get_object()

        # 1) Collect target + descendant ids (iterative BFS, avoids recursion limit)
        target_ids = []
        queue = [subject.id]
        while queue:
            current_id = queue.pop(0)
            target_ids.append(current_id)
            children_ids = list(
                BudgetSubject.objects.filter(parent_id=current_id).values_list('id', flat=True)
            )
            queue.extend(children_ids)

        # 2) Reject when linked BudgetEntry exists
        from .models import BudgetEntry
        linked_entry_count = BudgetEntry.objects.filter(subject_id__in=target_ids).count()
        if linked_entry_count > 0:
            return Response(
                {
                    'error': f'Cannot delete: linked budget entries exist ({linked_entry_count}). '
                             f'Remove linked entries first and retry.',
                    'linked_entry_count': linked_entry_count,
                },
                status=status.HTTP_409_CONFLICT,
            )

        # 3) Delete only when no linked entries
        subject.delete()

        return Response({
            'status': 'ok',
            'deleted_subjects_count': len(target_ids),
            'deleted_entries_count': 0,
            'message': f'Deleted {len(target_ids)} subject(s).',
        })

    def destroy(self, request, *args, **kwargs):
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        from django.db.models.deletion import ProtectedError
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError as e:
            return Response({
                'error': 'Cannot delete: this subject is still in use.',
                'details': 'Linked budget data exists. Reassign or remove linked data first, then retry.'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='reorder')
    @transaction.atomic
    def reorder(self, request):
        """
        ordered_ids: [id1, id2, ...] in desired order
        Update sort_order to 0,1,2...
        """
        denied = self._ensure_subject_editor(request)
        if denied is not None:
            return denied
        ordered_ids = request.data.get('ordered_ids', [])
        if not ordered_ids or not isinstance(ordered_ids, list):
            return Response({'error': 'ordered_ids array is required.'}, status=status.HTTP_400_BAD_REQUEST)
        objs = BudgetSubject.objects.filter(id__in=ordered_ids)
        id_map = {o.id: o for o in objs}
        updated = []
        for idx, oid in enumerate(ordered_ids):
            obj = id_map.get(int(oid))
            if obj and obj.sort_order != idx:
                obj.sort_order = idx
                updated.append(obj)
        if updated:
            BudgetSubject.objects.bulk_update(updated, ['sort_order'])
        return Response({'updated': len(updated)})

class EntrustedProjectViewSet(viewsets.ModelViewSet):
    queryset = EntrustedProject.objects.select_related('organization', 'source_project').all().order_by('-year', 'name')
    serializer_class = EntrustedProjectSerializer

    def _ensure_project_editor(self, request):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Project management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return None

    def _generate_code(self):
        """Generate project code with low collision risk."""
        import time, random, string
        ts = hex(int(time.time() * 1000))[2:].upper()
        rand = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
        return f"EP_{ts}_{rand}"

    def get_queryset(self):
        queryset = _scope_queryset_by_org(super().get_queryset(), self.request, org_field='organization_id')
        org_id = self.request.query_params.get('org_id')
        year = self.request.query_params.get('year')
        keyword = self.request.query_params.get('q')
        status_param = self.request.query_params.get('status')
        if org_id:
            if str(org_id).isdigit():
                queryset = queryset.filter(organization_id=org_id)
            else:
                queryset = queryset.filter(organization__code=org_id)
        if year:
            queryset = queryset.filter(year=year)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if keyword:
            queryset = queryset.filter(Q(name__icontains=keyword))
        return queryset

    def create(self, request, *args, **kwargs):
        denied = self._ensure_project_editor(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_project_editor(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._ensure_project_editor(request)
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def perform_create(self, serializer):
        """Create project with auto-generated code."""
        organization = serializer.validated_data.get('organization')
        if not _org_in_scope(self.request, getattr(organization, 'id', None)):
            raise PermissionDenied('No permission to create project for this organization.')
        code = self._generate_code()
        # Retry on code collision
        while EntrustedProject.objects.filter(code=code).exists():
            code = self._generate_code()
        serializer.save(code=code)

    def destroy(self, request, *args, **kwargs):
        """Reject project delete when linked budget entries exist."""
        denied = self._ensure_project_editor(request)
        if denied is not None:
            return denied
        project = self.get_object()
        entry_count = BudgetEntry.objects.filter(entrusted_project=project).count()
        if entry_count > 0:
            return Response(
                {
                    'error': f'Cannot delete project: linked budget entries exist ({entry_count}). Clean up linked entries first.',
                    'entry_count': entry_count,
                    'can_force': True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        # Guard rail: entrusted project deletion must not alter budget subject hierarchy.
        subject_ids_before = set(BudgetSubject.objects.values_list('id', flat=True))
        try:
            with transaction.atomic():
                response = super().destroy(request, *args, **kwargs)
                subject_ids_after = set(BudgetSubject.objects.values_list('id', flat=True))
                if subject_ids_before != subject_ids_after:
                    removed_subject_count = len(subject_ids_before - subject_ids_after)
                    raise DRFValidationError({
                        'error': '수탁사업 삭제 중 예산 계정 체제 변경이 감지되어 삭제를 취소했습니다.',
                        'details': {'removed_subject_count': removed_subject_count},
                    })
                return response
        except DRFValidationError as exc:
            logger.error(
                'EntrustedProject delete rollback due to budget subject drift: project_id=%s detail=%s',
                project.id,
                exc.detail,
            )
            return Response(exc.detail, status=status.HTTP_409_CONFLICT)

    @action(detail=True, methods=['delete'], url_path='force-delete')
    @transaction.atomic
    def force_delete_project(self, request, pk=None):  # noqa: ARG002
        """Force-delete a project and all linked budget entries/details. ADMIN-only."""
        role = _normalize_role(_role(request))
        if role != 'ADMIN':
            return Response({'error': 'Project force-delete is ADMIN-only.'}, status=status.HTTP_403_FORBIDDEN)
        project = self.get_object()
        entry_count = BudgetEntry.objects.filter(entrusted_project=project).count()
        BudgetEntry.objects.filter(entrusted_project=project).delete()
        project.delete()
        return Response({'status': 'force_deleted', 'deleted_entries': entry_count})

    @action(detail=True, methods=['post'], url_path='clone')
    def clone(self, request, pk=None):
        """Clone entrusted project to another year/organization."""
        denied = self._ensure_project_editor(request)
        if denied is not None:
            return denied
        source = self.get_object()
        new_year = request.data.get('year')
        new_name = (request.data.get('name') or '').strip()
        new_org_id = request.data.get('organization') or source.organization_id
        copy_entries = request.data.get('copy_entries', True)

        if not new_year or not new_name:
            return Response({'error': 'year and name are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            new_year = int(new_year)
        except (TypeError, ValueError):
            return Response({'error': 'Invalid year value.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            org = Organization.objects.get(pk=new_org_id)
        except Organization.DoesNotExist:
            return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)
        if not _org_in_scope(request, org.id):
            return Response({'error': 'No permission for target organization.'}, status=status.HTTP_403_FORBIDDEN)

        code = self._generate_code()
        while EntrustedProject.objects.filter(code=code).exists():
            code = self._generate_code()

        with transaction.atomic():
            new_proj = EntrustedProject.objects.create(
                organization=org,
                year=new_year,
                code=code,
                name=new_name,
                status='PLANNED',
                source_project=source,
            )

            if copy_entries:
                # Clone BudgetEntry rows from source project (keep amount/formula, no history)
                for entry in source.entries.select_related('subject').prefetch_related('details').all():
                    new_entry = BudgetEntry.objects.create(
                        subject=entry.subject,
                        organization=org,
                        entrusted_project=new_proj,
                        year=new_year,
                        status='DRAFT',
                        last_year_amount=entry.total_amount,  # keep source amount as last-year baseline
                        budget_category='ORIGINAL',
                        supplemental_round=0,
                        carryover_type='NONE',
                    )
                    for detail in entry.details.all():
                        BudgetDetail.objects.create(
                            entry=new_entry,
                            name=detail.name,
                            price=detail.price,
                            qty=detail.qty,
                            freq=detail.freq,
                            currency_unit=detail.currency_unit,
                            unit=detail.unit,
                            freq_unit=detail.freq_unit,
                            sort_order=detail.sort_order,
                            sub_label=detail.sub_label,
                            source=detail.source,
                            organization=detail.organization,
                        )

        serializer = self.get_serializer(new_proj)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class BudgetEntryViewSet(viewsets.ModelViewSet):
    queryset = BudgetEntry.objects.all()
    serializer_class = BudgetEntrySerializer

    def _role(self, user):
        """Return normalized role value."""
        profile = getattr(user, 'profile', None)
        role = getattr(profile, 'role', None)
        return _normalize_role(role)

    def create(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'STAFF', 'MANAGER', 'ADMIN'},
            message='No permission to create budget entry.',
        )
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'STAFF', 'MANAGER', 'ADMIN'},
            message='No permission to update budget entry.',
        )
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'STAFF', 'MANAGER', 'ADMIN'},
            message='No permission to update budget entry.',
        )
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def perform_create(self, serializer):
        organization = serializer.validated_data.get('organization')
        if not _org_in_scope(self.request, getattr(organization, 'id', None)):
            raise PermissionDenied('No permission to create entry for this organization.')
        serializer.save()

    def perform_update(self, serializer):
        instance = serializer.instance
        if not _org_in_scope(self.request, getattr(instance, 'organization_id', None)):
            raise PermissionDenied('No permission to update this entry.')
        next_org = serializer.validated_data.get('organization')
        if next_org is not None and not _org_in_scope(self.request, getattr(next_org, 'id', None)):
            raise PermissionDenied('No permission to move entry to this organization.')
        serializer.save()

    def get_queryset(self):
        prefetch_logs = Prefetch(
            'approval_logs',
            queryset=ApprovalLog.objects.select_related('actor').order_by('-created_at'),
        )
        queryset = (
            BudgetEntry.objects
            .select_related('organization', 'subject', 'entrusted_project')
            .prefetch_related('details', 'details__organization', prefetch_logs, 'comments')
            .order_by('id')
        )
        queryset = _scope_queryset_by_org(queryset, self.request, org_field='organization_id')
        org_id = self.request.query_params.get('org_id')
        year = self.request.query_params.get('year')
        entrusted_project_id = self.request.query_params.get('entrusted_project_id')
        keyword = self.request.query_params.get('q')
        if org_id:
            if str(org_id).isdigit():
                queryset = queryset.filter(organization_id=org_id)
            else:
                queryset = queryset.filter(organization__code=org_id)
        if year:
            queryset = queryset.filter(year=year)
        if entrusted_project_id:
            queryset = queryset.filter(entrusted_project_id=entrusted_project_id)
        if keyword:
            queryset = queryset.filter(
                Q(subject__name__icontains=keyword) |
                Q(subject__code__icontains=keyword) |
                Q(entrusted_project__name__icontains=keyword) |
                Q(entrusted_project__code__icontains=keyword)
            )
        
        round_param = self.request.query_params.get('round') or self.request.query_params.get('supplemental_round')
        if round_param is not None:
             queryset = queryset.filter(supplemental_round=round_param)
        
        return queryset

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def submit(self, request, pk=None):
        # STAFF owner or ADMIN can submit (DRAFT -> PENDING)
        role = self._role(request.user)
        if role not in ('STAFF', 'ADMIN'):
            return Response({'error': 'Not permitted'}, status=status.HTTP_403_FORBIDDEN)
        entry = self.get_object()
        if entry.status != 'DRAFT':
            return Response({'error': 'Only DRAFT entries can be submitted'}, status=status.HTTP_400_BAD_REQUEST)
        old_status = entry.status
        entry.status = 'PENDING'
        entry.save(update_fields=['status'])
        ApprovalLog.objects.create(entry=entry, from_status=old_status, to_status='PENDING', actor=request.user)
        return Response({'status': 'submitted'})

    def destroy(self, request, *args, **kwargs):
        """Delete budget entry only when version is editable and role is STAFF+."""
        entry = self.get_object()
        role = self._role(request.user)
        if role not in ('STAFF', 'MANAGER', 'ADMIN'):
            return Response({'error': 'No permission to delete.'}, status=status.HTTP_403_FORBIDDEN)
        # Reject delete when related version is CLOSED
        version = BudgetVersion.objects.filter(year=entry.year, round=entry.supplemental_round).first()
        if version and version.computed_status in ('CLOSED', 'EXPIRED'):
            return Response({'error': 'Cannot delete entries in a closed round.'}, status=status.HTTP_409_CONFLICT)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def approve(self, request, pk=None):
        entry = self.get_object()
        old_status = entry.status
        role = self._role(request.user)
        # PENDING -> REVIEWING by MANAGER/ADMIN
        if entry.status == 'PENDING':
            if role not in ('MANAGER', 'ADMIN'):
                return Response({'error': 'Not permitted'}, status=status.HTTP_403_FORBIDDEN)
            new_status = 'REVIEWING'
        # REVIEWING -> FINALIZED by MANAGER/ADMIN
        elif entry.status == 'REVIEWING':
            if role not in ('MANAGER', 'ADMIN'):
                return Response({'error': 'Not permitted'}, status=status.HTTP_403_FORBIDDEN)
            new_status = 'FINALIZED'
        else:
            return Response({'error': 'Cannot approve in current status'}, status=status.HTTP_400_BAD_REQUEST)

        entry.status = new_status
        entry.save(update_fields=['status'])
        ApprovalLog.objects.create(entry=entry, from_status=old_status, to_status=new_status, actor=request.user)
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def reject(self, request, pk=None):
        entry = self.get_object()
        reason = request.data.get('reason')
        old_status = entry.status
        role = self._role(request.user)
        # Reject request: MANAGER/ADMIN only
        if entry.status in ('PENDING', 'REVIEWING') and role not in ('MANAGER', 'ADMIN'):
            return Response({'error': 'Not permitted'}, status=status.HTTP_403_FORBIDDEN)
        entry.status = 'DRAFT'
        entry.save(update_fields=['status'])
        ApprovalLog.objects.create(entry=entry, from_status=old_status, to_status='DRAFT', actor=request.user, reason=reason)
        return Response({'status': 'rejected'})

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def reopen(self, request, pk=None):
        entry = self.get_object()
        role = self._role(request.user)
        if role != 'ADMIN':
            return Response({'error': 'Not permitted'}, status=status.HTTP_403_FORBIDDEN)
        if entry.status != 'FINALIZED':
            return Response({'error': 'Only finalized entries can be reopened'}, status=status.HTTP_400_BAD_REQUEST)

        to_status = request.data.get('to_status') or 'DRAFT'
        if to_status not in ('DRAFT', 'REVIEWING'):
            return Response({'error': 'Invalid to_status'}, status=status.HTTP_400_BAD_REQUEST)

        reason = request.data.get('reason') or 'reopen from finalized'
        old_status = entry.status
        entry.status = to_status
        entry.save(update_fields=['status'])
        ApprovalLog.objects.create(entry=entry, from_status=old_status, to_status=to_status, actor=request.user, reason=reason)
        return Response({'status': 'reopened', 'to_status': to_status})

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def recall(self, request, pk=None):
        """Recall approved/reviewing/finalized entry back to DRAFT."""
        entry = self.get_object()
        role = self._role(request.user)

        # Permission: STAFF owner or ADMIN
        if role not in ('STAFF', 'ADMIN'):
            return Response({'error': 'No permission to recall.'}, status=status.HTTP_403_FORBIDDEN)

        # No-op for DRAFT
        if entry.status == 'DRAFT':
            return Response({'error': 'Entry is already in DRAFT.'}, status=status.HTTP_400_BAD_REQUEST)

        # Recall is allowed only from PENDING/REVIEWING/FINALIZED
        if entry.status not in ('PENDING', 'REVIEWING', 'FINALIZED'):
            return Response({'error': 'Current status cannot be recalled.'}, status=status.HTTP_400_BAD_REQUEST)

        reason = request.data.get('reason') or 'Entry recall'
        old_status = entry.status
        entry.status = 'DRAFT'
        entry.save(update_fields=['status'])

        ApprovalLog.objects.create(
            entry=entry,
            from_status=old_status,
            to_status='DRAFT',
            actor=request.user,
            reason=f'[Recall] {reason}'
        )

        # Optional notification
        if entry.organization and entry.organization.userprofile_set.exists():
            Notification.objects.create(
                user=entry.organization.userprofile_set.first().user,
                message=f'Entry recalled: {entry.subject.name} ({old_status} -> DRAFT)'
            )

        return Response({
            'status': 'recalled',
            'from_status': old_status,
            'to_status': 'DRAFT',
            'message': 'Entry was recalled successfully.'
        })

    @action(detail=True, methods=['post'], url_path='note')
    def note(self, request, pk=None):
        """
        Write a note log without changing status.
        """
        entry = self.get_object()
        role = self._role(request.user)
        if role not in ('STAFF', 'MANAGER', 'ADMIN'):
            return Response({'error': 'No permission to write note log.'}, status=status.HTTP_403_FORBIDDEN)

        reason = str(request.data.get('reason') or '').strip()
        if not reason:
            return Response({'error': 'reason is required'}, status=status.HTTP_400_BAD_REQUEST)

        ApprovalLog.objects.create(
            entry=entry,
            from_status=entry.status,
            to_status=entry.status,
            actor=request.user,
            reason=reason,
        )
        return Response({'status': 'logged'})

    @action(detail=False, methods=['post'], url_path='workflow')
    def workflow(self, request):
        action_name = request.data.get('action')
        org_id = request.data.get('org_id')
        year = request.data.get('year')
        round_no = request.data.get('round')
        reason = str(request.data.get('reason') or '').strip()
        entry_ids = request.data.get('entry_ids')  # optional subset processing

        if action_name not in ('submit', 'approve', 'reject', 'reopen'):
            return Response({'error': 'Invalid action.'}, status=status.HTTP_400_BAD_REQUEST)
        if not org_id or year is None or round_no is None:
            return Response({'error': 'org_id, year, and round are required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not _org_in_scope(request, org_id):
            return Response({'error': 'No permission for this organization.'}, status=status.HTTP_403_FORBIDDEN)

        # If entry_ids exists process subset, otherwise process full org scope
        if entry_ids and isinstance(entry_ids, list) and len(entry_ids) > 0:
            queryset = BudgetEntry.objects.filter(
                id__in=entry_ids,
                year=year,
                supplemental_round=round_no,
            )
        else:
            queryset = BudgetEntry.objects.filter(
                organization_id=org_id,
                year=year,
                supplemental_round=round_no,
            )
        queryset = _scope_queryset_by_org(queryset, request, org_field='organization_id')

        entries_data = list(queryset.values('id', 'status'))
        if not entries_data:
            return Response({'error': 'No budget entries found for the given condition.'}, status=status.HTTP_400_BAD_REQUEST)

        role = self._role(request.user)
        action_labels = {
            'submit': 'Submit',
            'approve': 'Approve',
            'reject': 'Reject',
            'reopen': 'Reopen',
        }

        # Permission and status check
        if action_name == 'submit':
            # STAFF/ADMIN can submit DRAFT -> PENDING
            if role not in ('STAFF', 'ADMIN'):
                return Response({'error': 'No permission to submit. (STAFF/ADMIN only)'}, status=status.HTTP_403_FORBIDDEN)
            from_status = 'DRAFT'
            to_status = 'PENDING'
        elif action_name == 'approve':
            statuses = {entry['status'] for entry in entries_data}
            if statuses == {'PENDING'} or 'PENDING' in statuses:
                # PENDING -> REVIEWING by MANAGER/ADMIN
                if role not in ('MANAGER', 'ADMIN'):
                    return Response({'error': 'No permission to review. (MANAGER/ADMIN only)'}, status=status.HTTP_403_FORBIDDEN)
                from_status = 'PENDING'
                to_status = 'REVIEWING'
            elif statuses == {'REVIEWING'} or 'REVIEWING' in statuses:
                # REVIEWING -> FINALIZED by MANAGER/ADMIN
                if role not in ('MANAGER', 'ADMIN'):
                    return Response({'error': 'No permission for final approval. (MANAGER/ADMIN only)'}, status=status.HTTP_403_FORBIDDEN)
                from_status = 'REVIEWING'
                to_status = 'FINALIZED'
            else:
                return Response({'error': 'No entries are in approvable status.', 'statuses': list(statuses)}, status=status.HTTP_400_BAD_REQUEST)
        elif action_name == 'reject':
            statuses = {entry['status'] for entry in entries_data}
            if 'PENDING' in statuses or 'REVIEWING' in statuses:
                # Reject request by MANAGER/ADMIN only
                if role not in ('MANAGER', 'ADMIN'):
                    return Response({'error': 'No permission to reject. (MANAGER/ADMIN only)'}, status=status.HTTP_403_FORBIDDEN)
                from_status = 'PENDING' if 'PENDING' in statuses else 'REVIEWING'
            else:
                return Response({'error': 'No entries are rejectable.'}, status=status.HTTP_400_BAD_REQUEST)
            to_status = 'DRAFT'
        else:  # reopen
            if role != 'ADMIN':
                return Response({'error': 'No permission to reopen. (ADMIN only)'}, status=status.HTTP_403_FORBIDDEN)
            from_status = 'FINALIZED'
            to_status = request.data.get('to_status') or 'DRAFT'
            if to_status not in ('DRAFT', 'REVIEWING'):
                return Response({'error': 'Invalid target status.'}, status=status.HTTP_400_BAD_REQUEST)

        target_ids = [entry['id'] for entry in entries_data if entry['status'] == from_status]
        skipped = len(entries_data) - len(target_ids)

        if not target_ids:
            return Response({
                'error': f"No entries available for {action_labels[action_name]}. (current statuses: {', '.join({entry['status'] for entry in entries_data})})"
            }, status=status.HTTP_400_BAD_REQUEST)

        default_reason = reason or f'Department bulk {action_labels[action_name]}'

        with transaction.atomic():
            updated = BudgetEntry.objects.filter(id__in=target_ids).update(status=to_status)

            updated_entries = BudgetEntry.objects.filter(id__in=target_ids).values_list('id', flat=True)
            approval_logs = [
                ApprovalLog(
                    entry_id=entry_id,
                    from_status=from_status,
                    to_status=to_status,
                    actor=request.user,
                    reason=default_reason
                )
                for entry_id in updated_entries
            ]
            ApprovalLog.objects.bulk_create(approval_logs)

            try:
                org = Organization.objects.get(id=org_id)
                profiles = list(org.userprofile_set.values_list('user_id', flat=True))
                notifications = [
                    Notification(
                        user_id=user_id,
                        message=f'[{org.name}] Budget {action_labels[action_name]} processed: {updated} item(s) ({from_status} -> {to_status})'
                    )
                    for user_id in profiles
                ]
                Notification.objects.bulk_create(notifications)
            except Organization.DoesNotExist:
                pass

        return Response({
            'status': 'ok',
            'action': action_name,
            'from_status': from_status,
            'to_status': to_status,
            'updated_count': updated,
            'skipped_count': skipped,
            'message': f'{updated} item(s) {action_labels[action_name]} processed'
            + (f' ({skipped} skipped)' if skipped else '')
        })

class BudgetVersionViewSet(viewsets.ModelViewSet):
    queryset = BudgetVersion.objects.all().order_by('-year', '-round')
    serializer_class = BudgetVersionSerializer

    def _ensure_version_editor(self, request):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Version management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return None

    def create(self, request, *args, **kwargs):
        denied = self._ensure_version_editor(request)
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = self._ensure_version_editor(request)
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = self._ensure_version_editor(request)
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    @staticmethod
    def _as_int(value, default=None):
        try:
            if value is None or value == '':
                return default
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _detail_snapshot(detail):
        return {
            'detail_id': detail.id,
            'entry_id': detail.entry_id,
            'name': detail.name,
            'price': detail.price,
            'qty': detail.qty,
            'freq': detail.freq,
            'currency_unit': detail.currency_unit,
            'unit': detail.unit,
            'freq_unit': detail.freq_unit,
            'source': detail.source,
            'sub_label': detail.sub_label,
            'sort_order': detail.sort_order,
            'is_rate': detail.is_rate,
            'organization': detail.organization_id,
            'total_price': detail.total_price,
        }

    def _clone_entries_into_version(self, source_entries, target_version, *, budget_category):
        created_entries = 0
        detail_rows = []
        target_year = target_version.year
        project_mapping = {}
        for src in source_entries:
            src_detail_total = sum(detail.total_price for detail in src.details.all())
            src_total_amount = int(getattr(src, 'total_amount', 0) or 0)
            if src_total_amount == 0:
                src_total_amount = src_detail_total
            # Entrusted project mapping: create or map project for target year if missing
            target_proj = src.entrusted_project
            if target_proj and target_proj.year != target_year:
                if target_proj.id not in project_mapping:
                    existing = EntrustedProject.objects.filter(
                        organization=target_proj.organization, year=target_year, code=target_proj.code
                    ).first() or EntrustedProject.objects.filter(
                        organization=target_proj.organization, year=target_year, name=target_proj.name
                    ).first()

                    if existing:
                        project_mapping[target_proj.id] = existing
                    else:
                        project_mapping[target_proj.id] = EntrustedProject.objects.create(
                            organization=target_proj.organization,
                            year=target_year,
                            code=target_proj.code,
                            name=target_proj.name,
                            status='PLANNED',
                            source_project=target_proj
                        )
                target_proj = project_mapping[target_proj.id]

            new_entry = BudgetEntry.objects.create(
                subject=src.subject,
                organization=src.organization,
                entrusted_project=target_proj,
                year=target_year,
                supplemental_round=target_version.round,
                status='DRAFT',
                last_year_amount=src_total_amount,
                budget_category=budget_category,
                carryover_type=src.carryover_type,
                total_amount=src_total_amount,
                executed_amount=0,
                remaining_amount=src_total_amount,
            )
            created_entries += 1

            for detail in src.details.all():
                detail_rows.append(BudgetDetail(
                    entry=new_entry,
                    name=detail.name,
                    price=detail.price,
                    qty=detail.qty,
                    freq=detail.freq,
                    currency_unit=detail.currency_unit,
                    unit=detail.unit,
                    freq_unit=detail.freq_unit,
                    sub_label=detail.sub_label,
                    sort_order=detail.sort_order,
                    source=detail.source,
                    is_rate=detail.is_rate,
                    organization=detail.organization,
                    region_context=detail.region_context,
                    weather_context=detail.weather_context,
                    evidence_source_name=detail.evidence_source_name,
                    evidence_source_url=detail.evidence_source_url,
                    evidence_as_of=detail.evidence_as_of,
                    transfer_source_detail=detail,
                    before_snapshot=self._detail_snapshot(detail),
                ))

        created_details = len(detail_rows)
        if detail_rows:
            BudgetDetail.objects.bulk_create(detail_rows)

        return created_entries, created_details

    @action(detail=False, methods=['post'])
    def create_next_round(self, request):
        denied = self._ensure_version_editor(request)
        if denied is not None:
            return denied
        year = self._as_int(request.data.get('year'))
        round_name = request.data.get('name')
        if year is None:
            return Response({'error': 'year required'}, status=status.HTTP_400_BAD_REQUEST)

        latest = BudgetVersion.objects.filter(year=year).order_by('-round').first()
        next_round = (latest.round + 1) if latest else 0
        default_name = f"{year}년 {next_round}차 추경" if next_round > 0 else f"{year}년 본예산"
        name = round_name.strip() if isinstance(round_name, str) and round_name.strip() else default_name

        base_data_mode = str(request.data.get('base_data_mode') or '').upper().strip()
        creation_mode = str(request.data.get('creation_mode') or 'NEW').upper()
        source_version_id = self._as_int(
            request.data.get('source_version_id')
            or request.data.get('base_version_id')
            or request.data.get('baseline_version_id')
        )

        if base_data_mode in ('IMPORT_PREVIOUS', 'IMPORT', 'TRANSFER'):
            creation_mode = 'TRANSFER'
        elif base_data_mode in ('NEW', 'EMPTY'):
            creation_mode = 'NEW'

        if source_version_id and creation_mode != 'TRANSFER':
            creation_mode = 'TRANSFER'
        if creation_mode not in ('NEW', 'TRANSFER'):
            creation_mode = 'NEW'

        source_version = None
        source_entries = None
        clone_budget_category = 'SUPPLEMENTAL' if next_round > 0 else 'ORIGINAL'

        if creation_mode == 'TRANSFER':
            if source_version_id is None:
                return Response({'error': 'source_version_id is required for transfer mode'}, status=status.HTTP_400_BAD_REQUEST)
            source_version = BudgetVersion.objects.filter(id=source_version_id).first()
            if source_version is None:
                return Response({'error': 'source version not found'}, status=status.HTTP_404_NOT_FOUND)
            source_entries = (
                BudgetEntry.objects
                .filter(year=source_version.year, supplemental_round=source_version.round)
                .select_related('subject', 'organization', 'entrusted_project')
                .prefetch_related('details')
            )
            if not source_entries.exists():
                return Response({'error': 'source version has no entries'}, status=status.HTTP_404_NOT_FOUND)
        elif next_round > 0 and latest:
            source_version = latest
            source_entries = (
                BudgetEntry.objects
                .filter(year=year, supplemental_round=latest.round, status='FINALIZED')
                .select_related('subject', 'organization', 'entrusted_project')
                .prefetch_related('details')
            )

        version, created = BudgetVersion.objects.get_or_create(year=year, round=next_round, defaults={
            'name': name,
            'status': 'DRAFT',
            'start_date': request.data.get('start_date') or None,
            'end_date': request.data.get('end_date') or None,
            'guidelines': request.data.get('guidelines') or '',
            'guidelines_file': request.FILES.get('guidelines_file') or None,
            'creation_mode': 'TRANSFER' if (creation_mode == 'TRANSFER' and source_version) else 'NEW',
            'source_version': source_version if creation_mode == 'TRANSFER' else None,
        })

        cloned_count = 0
        cloned_detail_count = 0
        if not created:
            existing_entry_count = BudgetEntry.objects.filter(
                year=version.year,
                supplemental_round=version.round,
            ).count()

            should_clone_into_existing = (
                creation_mode == 'TRANSFER'
                and source_version is not None
                and source_entries is not None
                and source_entries.exists()
                and existing_entry_count == 0
            )
            if should_clone_into_existing:
                with transaction.atomic():
                    version.creation_mode = 'TRANSFER'
                    version.source_version = source_version
                    version.save(update_fields=['creation_mode', 'source_version'])
                    cloned_count, cloned_detail_count = self._clone_entries_into_version(
                        source_entries=source_entries,
                        target_version=version,
                        budget_category=clone_budget_category,
                    )
                existing_entry_count = BudgetEntry.objects.filter(
                    year=version.year,
                    supplemental_round=version.round,
                ).count()

            payload = BudgetVersionSerializer(version).data
            payload['cloned_count'] = cloned_count
            payload['cloned_detail_count'] = cloned_detail_count
            payload['already_exists'] = True
            payload['existing_entry_count'] = existing_entry_count
            return Response(payload, status=status.HTTP_200_OK)

        with transaction.atomic():
            version.name = name
            version.start_date = request.data.get('start_date') or None
            version.end_date = request.data.get('end_date') or None
            version.guidelines = request.data.get('guidelines') or ''
            if 'guidelines_file' in request.FILES:
                version.guidelines_file = request.FILES['guidelines_file']
            if creation_mode == 'TRANSFER' and source_version:
                version.creation_mode = 'TRANSFER'
                version.source_version = source_version
            else:
                version.creation_mode = 'NEW'
                version.source_version = None
            version.save()

            if source_entries is not None and source_entries.exists():
                cloned_count, cloned_detail_count = self._clone_entries_into_version(
                    source_entries=source_entries,
                    target_version=version,
                    budget_category=clone_budget_category,
                )

        payload = BudgetVersionSerializer(version).data
        payload['cloned_count'] = cloned_count
        payload['cloned_detail_count'] = cloned_detail_count
        return Response(payload, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='export-department-budget')
    def export_department_budget(self, request, pk=None):
        version = self.get_object()
        org_id = request.GET.get('org_id')
        
        if not org_id:
             return Response({'error': 'org_id is required'}, status=status.HTTP_400_BAD_REQUEST)
             
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from django.http import HttpResponse
        from io import BytesIO
        from urllib.parse import quote
        
        org = Organization.objects.filter(id=org_id).first()
        if not org:
            return Response({'error': 'Organization not found'}, status=status.HTTP_404_NOT_FOUND)
        if not _org_in_scope(request, org.id):
            return Response({'error': 'No permission for this organization.'}, status=status.HTTP_403_FORBIDDEN)
        
        org_name = org.name
        org_ids = [org.id] + list(Organization.objects.filter(parent_id=org.id).values_list('id', flat=True))
        
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
            
        # Re-usable styles
        fill_header = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
        font_title = Font(name='맑은 고딕', size=20, bold=False)
        font_header = Font(name='맑은 고딕', size=11, bold=True)
        font_normal = Font(name='맑은 고딕', size=11, bold=False)
        font_bold = Font(name='맑은 고딕', size=11, bold=True)
        
        bd_thin = Side(style='thin')
        border_all_thin = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thin)
        # For columns without right border (e.g. Col H in sample)
        border_no_right = Border(left=bd_thin, right=None, top=bd_thin, bottom=bd_thin)
        # Empty cells merging visually
        border_blank = Border(left=None, right=None, top=bd_thin, bottom=bd_thin)
        
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=False)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=False)

        def create_budget_sheet(subject_type, is_expense):
            title = f"{'지출' if is_expense else '수입'}({org_name})"
            if len(org_name) >= 2 and (org_name.endswith('실') or org_name.endswith('본부')):
                abbrev = org_name[:2]
                title = f"{'지출' if is_expense else '수입'}({abbrev})"
                
            ws = wb.create_sheet(title=title)
            
            # Row 1: Title
            head_title = f"{org_name} {'지출' if is_expense else '수입'}예산서"
            ws.cell(row=1, column=1, value=head_title)
            ws.cell(row=1, column=1).font = font_title
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
            
            # Row 3: Headers
            headers = ["장", "관", "항", "목", "예산액", "전년도 예산액", "증감액", "산출내역", "", "", "", "", "", ""]
            for col_idx, h_val in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=h_val)
                if h_val:
                    cell.fill = fill_header
                    cell.font = font_header
                    if col_idx == 7:
                        cell.alignment = align_center_wrap
                    else:
                        cell.alignment = align_center
                
                if col_idx <= 7:
                    cell.border = border_all_thin
                elif col_idx == 8:
                    cell.border = border_no_right
                else:
                    cell.border = border_blank

            entries = BudgetEntry.objects.filter(
                year=version.year,
                supplemental_round=version.round,
                organization_id__in=org_ids,
                subject__subject_type=subject_type
            ).select_related(
                'subject', 'subject__parent', 'subject__parent__parent', 
                'subject__parent__parent__parent', 'organization'
            ).prefetch_related('details').order_by('subject__code', 'organization__id')

            row_num = 4
            
            total_current = sum((entry.total_amount for entry in entries))
            total_previous = sum((entry.last_year_amount for entry in entries))
            
            # "수입계" / "지출계" row at row 4
            ws.row_dimensions[row_num].height = 30
            cell = ws.cell(row=row_num, column=1, value=f"{'지출' if is_expense else '수입'}계")
            cell.font = font_bold
            cell.alignment = align_center
            
            for c in range(1, 9):
                ws.cell(row=row_num, column=c).border = border_all_thin if c <= 7 else border_no_right
            for c in range(9, 15):
                ws.cell(row=row_num, column=c).border = border_blank
                
            ws.cell(row=row_num, column=5, value=total_current).number_format = '#,##0'
            ws.cell(row=row_num, column=6, value=total_previous).number_format = '#,##0'
            ws.cell(row=row_num, column=7, value=f"=E{row_num}-F{row_num}").number_format = '#,##0'
            
            row_num += 1

            # Grouping Logic to print hierarchy rows appropriately.
            class Node:
                def __init__(self, name):
                    self.name = name
                    self.current = 0
                    self.previous = 0
                    self.children = {}
                    self.entries = []
            
            tree = Node("Root")
            for entry in entries:
                subject = entry.subject
                by_level = {}
                cursor = subject
                while cursor:
                    lvl = int(getattr(cursor, 'level', 0) or 0)
                    if lvl > 0: by_level[lvl] = cursor.name
                    cursor = getattr(cursor, 'parent', None)
                
                jang = by_level.get(1, "") or ""
                gwan = by_level.get(2, "") or ""
                hang = by_level.get(3, "") or ""
                mok = by_level.get(4, "") or ""
                
                if jang not in tree.children: tree.children[jang] = Node(jang)
                if gwan not in tree.children[jang].children: tree.children[jang].children[gwan] = Node(gwan)
                if hang not in tree.children[jang].children[gwan].children: tree.children[jang].children[gwan].children[hang] = Node(hang)
                if mok not in tree.children[jang].children[gwan].children[hang].children: tree.children[jang].children[gwan].children[hang].children[mok] = Node(mok)
                
                mok_node = tree.children[jang].children[gwan].children[hang].children[mok]
                mok_node.entries.append(entry)
                
                c_amt = int(entry.total_amount or 0)
                p_amt = int(entry.last_year_amount or 0)
                
                tree.children[jang].current += c_amt
                tree.children[jang].previous += p_amt
                tree.children[jang].children[gwan].current += c_amt
                tree.children[jang].children[gwan].previous += p_amt
                tree.children[jang].children[gwan].children[hang].current += c_amt
                tree.children[jang].children[gwan].children[hang].previous += p_amt
                mok_node.current += c_amt
                mok_node.previous += p_amt

            def write_styled_row(r, c1, c2, c3, c4, c5_val, c6_val, is_mok=False):
                ws.row_dimensions[r].height = 30
                vals = [c1, c2, c3, c4]
                for i, v in enumerate(vals, start=1):
                    fc = ws.cell(row=r, column=i, value=v)
                    fc.font = font_normal
                    fc.alignment = align_center
                    fc.border = border_all_thin
                
                c5 = ws.cell(row=r, column=5, value=c5_val)
                c6 = ws.cell(row=r, column=6, value=c6_val)
                c7 = ws.cell(row=r, column=7, value=f"=E{r}-F{r}")
                
                for c in (c5, c6, c7):
                    c.number_format = '#,##0'
                    c.border = border_all_thin
                    c.font = font_normal
                    
                c8 = ws.cell(row=r, column=8, value="")
                c8.border = border_no_right
                for c in range(9, 15):
                    ws.cell(row=r, column=c).border = border_blank
                    
            def write_detail_row(r, detail_text, is_first, price_val=None, text_val=None):
                ws.row_dimensions[r].height = 30
                for c in range(1, 8):
                    ws.cell(row=r, column=c, value="").border = border_all_thin
                    
                c8 = ws.cell(row=r, column=8, value=f"  - {detail_text}")
                c8.font = font_normal
                c8.alignment = align_left
                c8.border = border_no_right
                
                if price_val is not None:
                    c9 = ws.cell(row=r, column=9, value=price_val)
                    c9.number_format = '#,##0'
                    c9.font = font_normal
                    c9.alignment = align_right
                    if text_val:
                        ws.cell(row=r, column=10, value=text_val).font = font_normal
                        
                for c in range(9, 15):
                    ws.cell(row=r, column=c).border = border_blank

            for j_name, j_node in tree.children.items():
                if j_name:
                    write_styled_row(row_num, j_name, "", "", "", j_node.current, j_node.previous)
                    row_num += 1
                for g_name, g_node in j_node.children.items():
                    if g_name:
                        write_styled_row(row_num, "", g_name, "", "", g_node.current, g_node.previous)
                        row_num += 1
                    for h_name, h_node in g_node.children.items():
                        if h_name:
                            write_styled_row(row_num, "", "", h_name, "", h_node.current, h_node.previous)
                            row_num += 1
                        for m_name, m_node in h_node.children.items():
                            write_styled_row(row_num, "", "", "", m_name, m_node.current, m_node.previous, is_mok=True)
                            row_num += 1
                            
                            for entry in m_node.entries:
                                details = list(entry.details.all())
                                if not details:
                                    if getattr(entry, 'remark', None):
                                        write_detail_row(row_num, f"[{org_name}] {entry.remark}", True, entry.total_amount, "원")
                                    else:
                                        write_detail_row(row_num, f"[{org_name}] 산출내역 없음", True, entry.total_amount, "원")
                                    row_num += 1
                                else:
                                    for idx, detail in enumerate(details):
                                        write_detail_row(row_num, detail.name, idx==0, detail.total_price or (detail.price * detail.qty * detail.freq), detail.currency_unit or "원")
                                        row_num += 1

            # Exact Column Widths parsed from sample (Expense) for both
            ws.column_dimensions['A'].width = 10.625 if not is_expense else 40.625
            ws.column_dimensions['B'].width = 46.625 if not is_expense else 20.625
            ws.column_dimensions['C'].width = 14.625
            ws.column_dimensions['D'].width = 17.625
            ws.column_dimensions['E'].width = 14.625
            ws.column_dimensions['F'].width = 14.625
            ws.column_dimensions['G'].width = 14.625
            ws.column_dimensions['H'].width = 50.625 if not is_expense else 55.625
            ws.column_dimensions['I'].width = 15.625 if not is_expense else 16.625
            ws.column_dimensions['J'].width = 16.625 if not is_expense else 4.625

        create_budget_sheet('income', False)
        create_budget_sheet('expense', True)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_name = f"{org_name}_부서예산서_{version.year}.xlsx"
        encoded_name = quote(file_name)
        
        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_name}"
        return response


    def destroy(self, request, *args, **kwargs):
        """Version delete is ADMIN-only. Reject when budget data exists."""
        if _normalize_role(_role(request)) != 'ADMIN':
            return Response({'error': 'Version delete is ADMIN-only.'}, status=status.HTTP_403_FORBIDDEN)
        version = self.get_object()
        entry_count = BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).count()
        if entry_count > 0:
            return Response(
                {
                    'error': f'Cannot delete version: linked budget entries exist ({entry_count}). Clean up entries first.',
                    'entry_count': entry_count,
                    'can_force': True,
                },
                status=status.HTTP_409_CONFLICT,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['delete'], url_path='force-delete')
    @transaction.atomic
    def force_delete_version(self, request, pk=None):  # noqa: ARG002
        """Force-delete a version and all linked budget entries/details. ADMIN-only."""
        if _normalize_role(_role(request)) != 'ADMIN':
            return Response({'error': 'Version force-delete is ADMIN-only.'}, status=status.HTTP_403_FORBIDDEN)
        version = self.get_object()
        entry_count = BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).count()
        BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).delete()
        version.delete()
        return Response({'status': 'force_deleted', 'deleted_entries': entry_count})

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        if _normalize_role(_role(request)) not in ('MANAGER', 'ADMIN'):
            return Response({'error': 'No permission to close round.'}, status=status.HTTP_403_FORBIDDEN)
        version = self.get_object()
        version.status = 'CLOSED'
        version.save()
        return Response({
            'status': 'closed',
            'version': BudgetVersionSerializer(version).data,
        })

    @action(detail=True, methods=['post'])
    def reopen(self, request, pk=None):
        if _normalize_role(_role(request)) not in ('MANAGER', 'ADMIN'):
            return Response({'error': 'No permission to reopen round.'}, status=status.HTTP_403_FORBIDDEN)
        version = self.get_object()
        version.status = 'PENDING'
        version.save()
        return Response({
            'status': 'reopened',
            'version': BudgetVersionSerializer(version).data,
        })

    @action(detail=True, methods=['post'])
    def clone_from_previous(self, request, pk=None):
        denied = self._ensure_version_editor(request)
        if denied is not None:
            return denied
        target_version = self.get_object()
        source_year = self._as_int(request.data.get('source_year'))
        source_round = self._as_int(request.data.get('source_round'), default=0)

        if source_year is None:
            return Response({'error': 'source_year is required'}, status=status.HTTP_400_BAD_REQUEST)

        if target_version.computed_status != 'DRAFT':
            return Response({'error': 'Target version must be DRAFT'}, status=status.HTTP_400_BAD_REQUEST)

        if BudgetEntry.objects.filter(year=target_version.year, supplemental_round=target_version.round).exists():
            return Response({'error': 'Target version already has data. Please clear it first.'}, status=status.HTTP_400_BAD_REQUEST)

        source_entries = (
            BudgetEntry.objects
            .filter(year=source_year, supplemental_round=source_round)
            .select_related('subject', 'organization', 'entrusted_project')
            .prefetch_related('details')
        )
        if not source_entries.exists():
            return Response({'error': 'Source version has no data.'}, status=status.HTTP_404_NOT_FOUND)

        source_version = BudgetVersion.objects.filter(year=source_year, round=source_round).first()
        budget_category = 'SUPPLEMENTAL' if target_version.round > 0 else 'ORIGINAL'

        with transaction.atomic():
            cloned_count, cloned_detail_count = self._clone_entries_into_version(
                source_entries=source_entries,
                target_version=target_version,
                budget_category=budget_category,
            )
            target_version.creation_mode = 'TRANSFER'
            target_version.source_version = source_version
            target_version.save(update_fields=['creation_mode', 'source_version'])

        return Response({
            'status': 'ok',
            'cloned_count': cloned_count,
            'cloned_detail_count': cloned_detail_count,
        })

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    @transaction.atomic
    def bulk_delete(self, request):
        """Bulk delete versions. ADMIN-only."""
        if _normalize_role(_role(request)) != 'ADMIN':
            return Response({'error': 'Version bulk-delete is ADMIN-only.'}, status=status.HTTP_403_FORBIDDEN)
        
        ids = request.data.get('ids', [])
        if not ids or not isinstance(ids, list):
            return Response({'error': 'ids array is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        force = request.data.get('force', False)
        
        deleted_ids = []
        errors = []
        
        for v_id in ids:
            try:
                version = BudgetVersion.objects.get(id=v_id)
                entry_count = BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).count()
                
                if entry_count > 0 and not force:
                    errors.append({'id': v_id, 'name': version.name, 'error': 'Linked entries exist', 'entry_count': entry_count})
                    continue
                
                if force:
                    BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).delete()
                
                version.delete()
                deleted_ids.append(v_id)
            except BudgetVersion.DoesNotExist:
                errors.append({'id': v_id, 'error': 'Not found'})
            except Exception as e:
                errors.append({'id': v_id, 'error': str(e)})

        return Response({
            'deleted_ids': deleted_ids,
            'errors': errors
        })

    @action(detail=False, methods=['post'], url_path='bulk-update-status')
    @transaction.atomic
    def bulk_update_status(self, request):
        """Bulk update version statuses. MANAGER/ADMIN only."""
        if _normalize_role(_role(request)) not in ('MANAGER', 'ADMIN'):
            return Response({'error': 'No permission for bulk status update.'}, status=status.HTTP_403_FORBIDDEN)
        
        ids = request.data.get('ids', [])
        status_val = request.data.get('status')
        
        if not ids or not isinstance(ids, list):
            return Response({'error': 'ids array is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not status_val:
            return Response({'error': 'status is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        updated_ids = []
        errors = []
        
        for v_id in ids:
            try:
                version = BudgetVersion.objects.get(id=v_id)
                version.status = status_val
                if status_val == 'CONFIRMED':
                    version.confirmed_at = timezone.now()
                version.save()
                updated_ids.append(v_id)
            except BudgetVersion.DoesNotExist:
                errors.append({'id': v_id, 'error': 'Not found'})
            except Exception as e:
                errors.append({'id': v_id, 'error': str(e)})

        return Response({
            'updated_ids': updated_ids,
            'errors': errors
        })


    @action(detail=True, methods=['get'])
    def progress(self, request, pk=None):
        try:
             version = self.get_object()
        except BudgetVersion.DoesNotExist:
             return Response({'error': 'Version not found'}, status=status.HTTP_404_NOT_FOUND)

        allowed_org_ids = _scope_org_ids_for_user(request)

        # Aggregate stats by organization for this version
        # We need to know: Org Name, Status (aggregated), Entry Count, Total Amount
        
        # 1. Get all organizations (flat list or tree, simplified to flat for now)
        orgs_qs = Organization.objects.all()
        if allowed_org_ids is not None:
            orgs_qs = orgs_qs.filter(id__in=allowed_org_ids)
        orgs = orgs_qs.values('id', 'name', 'code', 'parent_id', 'org_type')
        org_map = {o['id']: o for o in orgs}

        # 2. Get entries for this version
        entries = BudgetEntry.objects.filter(year=version.year, supplemental_round=version.round).prefetch_related('details')
        if allowed_org_ids is not None:
            entries = entries.filter(organization_id__in=allowed_org_ids)

        stats = {}
        
        for entry in entries:
            org_id = entry.organization_id
            if org_id not in stats:
                stats[org_id] = {
                    'org_id': org_id,
                    'org_name': org_map.get(org_id, {}).get('name', 'Unknown'),
                    'parent_id': org_map.get(org_id, {}).get('parent_id'),
                    'org_type': org_map.get(org_id, {}).get('org_type'),
                    'total_entries': 0,
                    'total_amount': 0,
                    'status_counts': {'DRAFT': 0, 'PENDING': 0, 'REVIEWING': 0, 'FINALIZED': 0}
                }
            
            s = stats[org_id]
            s['total_entries'] += 1
            s['total_amount'] += entry.total_amount
            s['status_counts'][entry.status] = s['status_counts'].get(entry.status, 0) + 1

        # Calculate logical status for the department
        # Logic: If all FINALIZED -> Completed. If any REVIEWING/PENDING -> Submitted. Else -> Writing.
        for stat in stats.values():
            counts = stat['status_counts']
            total = stat['total_entries']
            
            if total == 0:
                stat['dept_status'] = 'EMPTY'
            elif counts['FINALIZED'] == total:
                stat['dept_status'] = 'COMPLETED'
            elif counts['REVIEWING'] > 0 or counts['PENDING'] > 0:
                 stat['dept_status'] = 'SUBMITTED' # Partial or full submission
            else:
                 stat['dept_status'] = 'WRITING'

        # Include orgs with no entries as "NOT_STARTED"
        result_list = []
        for org in orgs:
            if org['id'] in stats:
                result_list.append(stats[org['id']])
            else:
                result_list.append({
                    'org_id': org['id'],
                    'org_name': org['name'],
                    'parent_id': org.get('parent_id'),
                    'org_type': org.get('org_type'),
                    'total_entries': 0,
                    'total_amount': 0,
                    'status_counts': {'DRAFT': 0, 'PENDING': 0, 'REVIEWING': 0, 'FINALIZED': 0},
                    'dept_status': 'NOT_STARTED'
                })
        
        return Response(result_list)

class BudgetDetailViewSet(viewsets.ModelViewSet):
    queryset = BudgetDetail.objects.select_related('entry', 'organization', 'author', 'updated_by').order_by('id')
    serializer_class = BudgetDetailSerializer

    def get_queryset(self):
        queryset = BudgetDetail.objects.select_related('entry', 'organization', 'author', 'updated_by').order_by('id')
        queryset = _scope_queryset_by_org(queryset, self.request, org_field='entry__organization_id')
        entry_id = self.request.query_params.get('entry')
        if str(entry_id).isdigit():
            queryset = queryset.filter(entry_id=int(entry_id))
        return queryset

    def perform_create(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to create budget details.')
        entry = serializer.validated_data.get('entry')
        if not entry:
            raise ValueError('entry is required')
        if not _org_in_scope(self.request, getattr(entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this budget entry.')
        source = serializer.validated_data.get('source') or 'SELF'
        organization = serializer.validated_data.get('organization') or entry.organization
        if organization is not None and not _org_in_scope(self.request, getattr(organization, 'id', None)):
            raise PermissionDenied('No permission for this organization.')
        currency_unit = serializer.validated_data.get('currency_unit') or '원'
        unit = serializer.validated_data.get('unit') or '식'
        freq_unit = serializer.validated_data.get('freq_unit') or '회'
        sort_order = serializer.validated_data.get('sort_order')
        if sort_order is None and entry:
            try:
                max_order = entry.details.aggregate(max_order=Max('sort_order')).get('max_order')
                sort_order = (max_order if max_order is not None else -1) + 1
            except DatabaseError:
                sort_order = 0
        actor = self.request.user if getattr(self.request.user, 'is_authenticated', False) else None
        serializer.save(
            source=source, 
            organization=organization, 
            sort_order=sort_order or 0, 
            currency_unit=currency_unit, 
            unit=unit, 
            freq_unit=freq_unit,
            author=actor,
            updated_by=actor
        )

    def perform_update(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to update budget details.')
        instance = serializer.instance
        if not _org_in_scope(self.request, getattr(instance.entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this budget detail.')
        next_org = serializer.validated_data.get('organization')
        if next_org is not None and not _org_in_scope(self.request, getattr(next_org, 'id', None)):
            raise PermissionDenied('No permission for this organization.')

        expected_updated_at = serializer.validated_data.get('_updated_at')
        if expected_updated_at is None:
            header_value = self.request.headers.get('X-Detail-Updated-At')
            if header_value:
                expected_updated_at = parse_datetime(header_value)

        if expected_updated_at is not None and instance.updated_at is not None:
            actual_updated_at = instance.updated_at
            if timezone.is_naive(expected_updated_at):
                expected_updated_at = timezone.make_aware(expected_updated_at, timezone.get_current_timezone())
            if timezone.is_naive(actual_updated_at):
                actual_updated_at = timezone.make_aware(actual_updated_at, timezone.get_current_timezone())
            if abs((actual_updated_at - expected_updated_at).total_seconds()) > 0.001:
                raise ConflictError({
                    'error': 'Detail was modified by another user. Refresh and retry.',
                    'code': 'DETAIL_CONFLICT',
                    'detail_id': instance.id,
                    'server_updated_at': actual_updated_at.isoformat(),
                })

        actor = self.request.user if getattr(self.request.user, 'is_authenticated', False) else None
        serializer.save(updated_by=actor)

    def perform_destroy(self, instance):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to delete budget details.')
        if not _org_in_scope(self.request, getattr(instance.entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this budget detail.')
        instance.delete()

    def create(self, request, *args, **kwargs):
        trace_id = uuid.uuid4().hex[:12]
        incoming = request.data.copy()

        insert_after_raw = incoming.pop('insert_after_detail_id', None)
        if isinstance(insert_after_raw, (list, tuple)):
            insert_after_raw = insert_after_raw[0] if insert_after_raw else None

        if incoming.get('sort_order') in (None, ''):
            if insert_after_raw not in (None, ''):
                try:
                    insert_after_id = int(insert_after_raw)
                except (TypeError, ValueError):
                    return Response(
                        {'error': 'invalid insert_after_detail_id', 'trace_id': trace_id},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                raw_entry_id = incoming.get('entry')
                if isinstance(raw_entry_id, (list, tuple)):
                    raw_entry_id = raw_entry_id[0] if raw_entry_id else None
                requested_entry_id = None
                if raw_entry_id not in (None, ''):
                    try:
                        requested_entry_id = int(raw_entry_id)
                    except (TypeError, ValueError):
                        requested_entry_id = None

                try:
                    with transaction.atomic():
                        anchor_qs = BudgetDetail.objects.select_for_update().filter(id=insert_after_id)
                        if requested_entry_id is not None:
                            anchor_qs = anchor_qs.filter(entry_id=requested_entry_id)
                        anchor = anchor_qs.values('entry_id', 'sort_order').first()
                        if not anchor:
                            return Response(
                                {'error': 'anchor detail not found for insertion', 'trace_id': trace_id},
                                status=status.HTTP_400_BAD_REQUEST,
                            )

                        target_entry_id = int(anchor['entry_id'])
                        insert_order = int(anchor['sort_order'] or 0) + 1
                        BudgetDetail.objects.select_for_update().filter(
                            entry_id=target_entry_id,
                            sort_order__gte=insert_order,
                        ).update(sort_order=F('sort_order') + 1)

                        incoming['entry'] = target_entry_id
                        incoming['sort_order'] = insert_order
                except DatabaseError as exc:
                    logger.exception('detail.create.insert_after.database_error trace_id=%s', trace_id)
                    return Response(
                        {'error': 'detail insert ordering failed', 'details': str(exc), 'trace_id': trace_id},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )
            else:
                entry_id = incoming.get('entry')
                if isinstance(entry_id, (list, tuple)):
                    entry_id = entry_id[0] if entry_id else None
                computed_order = 0
                if entry_id:
                    try:
                        computed_order = (
                            BudgetDetail.objects.filter(entry_id=entry_id)
                            .aggregate(max_order=Max('sort_order'))
                            .get('max_order') or -1
                        ) + 1
                    except DatabaseError:
                        computed_order = 0
                incoming['sort_order'] = computed_order

        payload_snapshot = {
            'entry': incoming.get('entry'),
            'name': incoming.get('name'),
            'price': incoming.get('price'),
            'qty': incoming.get('qty'),
            'freq': incoming.get('freq'),
            'currency_unit': incoming.get('currency_unit'),
            'unit': incoming.get('unit'),
            'freq_unit': incoming.get('freq_unit'),
            'source': incoming.get('source'),
            'organization': incoming.get('organization'),
            'sort_order': incoming.get('sort_order'),
            'insert_after_detail_id': insert_after_raw,
        }
        logger.info('detail.create.request trace_id=%s user_id=%s payload=%s', trace_id, getattr(request.user, 'id', None), payload_snapshot)
        serializer = self.get_serializer(data=incoming)
        try:
            serializer.is_valid(raise_exception=True)
        except DRFValidationError as exc:
            logger.warning('detail.create.validation_failed trace_id=%s errors=%s', trace_id, exc.detail)
            return Response(
                {'error': 'detail create validation failed', 'details': exc.detail, 'trace_id': trace_id},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            self.perform_create(serializer)
        except PermissionDenied as exc:
            logger.warning('detail.create.permission_denied trace_id=%s errors=%s', trace_id, str(exc))
            return Response(
                {'error': str(exc), 'trace_id': trace_id},
                status=status.HTTP_403_FORBIDDEN
            )
        except IntegrityError as exc:
            logger.exception('detail.create.integrity_error trace_id=%s', trace_id)
            return Response(
                {'error': 'detail create failed', 'details': str(exc), 'trace_id': trace_id},
                status=status.HTTP_400_BAD_REQUEST
            )
        except DatabaseError as exc:
            logger.exception('detail.create.database_error trace_id=%s', trace_id)
            return Response(
                {'error': 'detail create database error', 'details': str(exc), 'trace_id': trace_id},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as exc:
            logger.exception('detail.create.unexpected_error trace_id=%s', trace_id)
            return Response(
                {'error': 'detail create unexpected error', 'details': f'{exc.__class__.__name__}: {exc}', 'trace_id': trace_id},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        headers = self.get_success_headers(serializer.data)
        logger.info('detail.create.success trace_id=%s detail_id=%s', trace_id, serializer.data.get('id'))
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['post'])
    def parse_expression(self, request):
        expression = request.data.get('expression', '')
        try:
            parsed = parse_calc_expression(expression)
            return Response({
                'price': parsed.price,
                'qty': parsed.qty,
                'freq': parsed.freq,
                'amount': parsed.amount,
                'normalized': parsed.normalized,
            })
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

class BudgetExecutionViewSet(viewsets.ModelViewSet):
    queryset = BudgetExecution.objects.select_related('entry', 'created_by').order_by('-id')
    serializer_class = BudgetExecutionSerializer

    def get_queryset(self):
        queryset = BudgetExecution.objects.select_related('entry', 'created_by').order_by('-id')
        return _scope_queryset_by_org(queryset, self.request, org_field='entry__organization_id')

    def perform_create(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to create execution history.')
        entry = serializer.validated_data.get('entry')
        if not _org_in_scope(self.request, getattr(entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this budget entry.')
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        if not _is_manager_or_admin(self.request):
            raise PermissionDenied('Only MANAGER/ADMIN can update execution history.')
        instance = serializer.instance
        if not _org_in_scope(self.request, getattr(instance.entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this execution history.')
        serializer.save()

    def perform_destroy(self, instance):
        if not _is_manager_or_admin(self.request):
            raise PermissionDenied('Only MANAGER/ADMIN can delete execution history.')
        if not _org_in_scope(self.request, getattr(instance.entry, 'organization_id', None)):
            raise PermissionDenied('No permission for this execution history.')
        instance.delete()

class SpendingLimitRuleViewSet(viewsets.ModelViewSet):
    queryset = SpendingLimitRule.objects.order_by('id')
    serializer_class = SpendingLimitRuleSerializer

    def create(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Rule management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Rule management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Rule management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Rule management is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return super().destroy(request, *args, **kwargs)

class BudgetTransferViewSet(viewsets.ModelViewSet):
    queryset = BudgetTransfer.objects.select_related('from_entry', 'to_entry', 'created_by', 'approved_by').order_by('-created_at')
    serializer_class = BudgetTransferSerializer

    def get_queryset(self):
        queryset = BudgetTransfer.objects.select_related('from_entry', 'to_entry', 'created_by', 'approved_by').order_by('-created_at')
        allowed = _scope_org_ids_for_user(self.request)
        if allowed is None:
            return queryset
        if not allowed:
            return queryset.none()
        return queryset.filter(
            Q(from_entry__organization_id__in=allowed) |
            Q(to_entry__organization_id__in=allowed)
        )

    def perform_create(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to create transfers.')
        from_entry = serializer.validated_data.get('from_entry')
        to_entry = serializer.validated_data.get('to_entry')
        if not _org_in_scope(self.request, getattr(from_entry, 'organization_id', None)):
            raise PermissionDenied('No permission for source entry organization.')
        if not _org_in_scope(self.request, getattr(to_entry, 'organization_id', None)):
            raise PermissionDenied('No permission for destination entry organization.')
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        if not _is_manager_or_admin(self.request):
            raise PermissionDenied('Only MANAGER/ADMIN can update transfers.')
        transfer = serializer.instance
        if not _org_in_scope(self.request, transfer.from_entry.organization_id):
            raise PermissionDenied('No permission for source entry organization.')
        if not _org_in_scope(self.request, transfer.to_entry.organization_id):
            raise PermissionDenied('No permission for destination entry organization.')
        serializer.save()

    def perform_destroy(self, instance):
        if not _is_manager_or_admin(self.request):
            raise PermissionDenied('Only MANAGER/ADMIN can delete transfers.')
        if not _org_in_scope(self.request, instance.from_entry.organization_id):
            raise PermissionDenied('No permission for source entry organization.')
        if not _org_in_scope(self.request, instance.to_entry.organization_id):
            raise PermissionDenied('No permission for destination entry organization.')
        instance.delete()

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Only MANAGER/ADMIN can approve transfers.',
        )
        if denied is not None:
            return denied
        transfer = self.get_object()
        if not _org_in_scope(request, transfer.from_entry.organization_id):
            return Response({'error': 'No permission for source entry organization.'}, status=status.HTTP_403_FORBIDDEN)
        if not _org_in_scope(request, transfer.to_entry.organization_id):
            return Response({'error': 'No permission for destination entry organization.'}, status=status.HTTP_403_FORBIDDEN)
        if transfer.status != 'PENDING':
            return Response({'error': 'Transfer already processed'}, status=status.HTTP_400_BAD_REQUEST)
        transfer.status = 'APPROVED'
        transfer.approved_by = request.user
        transfer.approved_at = timezone.now()
        transfer.reject_reason = None
        transfer.save()
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Only MANAGER/ADMIN can reject transfers.',
        )
        if denied is not None:
            return denied
        transfer = self.get_object()
        if not _org_in_scope(request, transfer.from_entry.organization_id):
            return Response({'error': 'No permission for source entry organization.'}, status=status.HTTP_403_FORBIDDEN)
        if not _org_in_scope(request, transfer.to_entry.organization_id):
            return Response({'error': 'No permission for destination entry organization.'}, status=status.HTTP_403_FORBIDDEN)
        if transfer.status != 'PENDING':
            return Response({'error': 'Transfer already processed'}, status=status.HTTP_400_BAD_REQUEST)
        transfer.status = 'REJECTED'
        transfer.reject_reason = request.data.get('reason')
        transfer.approved_by = request.user
        transfer.approved_at = timezone.now()
        transfer.save()
        return Response({'status': 'rejected'})

class UserProfileViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UserProfile.objects.select_related('user', 'organization').all().order_by('id')
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = UserProfile.objects.select_related('user', 'organization', 'team').all().order_by('id')
        if _is_admin(self.request):
            return queryset
        return queryset.filter(user=self.request.user)

    @action(detail=False, methods=['get'])
    def me(self, request):
        profile = getattr(request.user, 'profile', None)
        if profile:
            return Response(UserProfileSerializer(profile).data)
        return Response({'error': 'No profile found'}, status=status.HTTP_404_NOT_FOUND)

class ApprovalLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ApprovalLog.objects.select_related('actor', 'entry').all().order_by('-created_at')
    serializer_class = ApprovalLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    @staticmethod
    def _parse_int_list(raw_value):
        if raw_value is None:
            return []
        tokens = str(raw_value).split(',')
        values = []
        for token in tokens:
            token = token.strip()
            if token.isdigit():
                values.append(int(token))
        return values

    @staticmethod
    def _parse_csv_list(raw_value):
        if raw_value is None:
            return []
        return [token.strip() for token in str(raw_value).split(',') if token.strip()]

    def get_queryset(self):
        queryset = ApprovalLog.objects.select_related('actor', 'entry').order_by('-created_at')
        allowed = _scope_org_ids_for_user(self.request)
        if allowed is not None:
            user_id_str = str(getattr(self.request.user, 'id', ''))
            if allowed:
                queryset = queryset.filter(
                    Q(entry__organization_id__in=allowed) |
                    Q(actor=self.request.user) |
                    Q(resource_type='auth', resource_id=user_id_str)
                )
            else:
                queryset = queryset.filter(
                    Q(actor=self.request.user) |
                    Q(resource_type='auth', resource_id=user_id_str)
                )

        entry_id = self.request.query_params.get('entry')
        if str(entry_id).isdigit():
            queryset = queryset.filter(entry_id=int(entry_id))

        entry_ids = self._parse_int_list(self.request.query_params.get('entry_ids'))
        if entry_ids:
            queryset = queryset.filter(entry_id__in=entry_ids)

        year = self.request.query_params.get('year')
        if str(year).isdigit():
            queryset = queryset.filter(entry__year=int(year))

        round_no = self.request.query_params.get('round')
        if str(round_no).isdigit():
            queryset = queryset.filter(entry__supplemental_round=int(round_no))

        org_id = self.request.query_params.get('org_id')
        if str(org_id).isdigit():
            queryset = queryset.filter(entry__organization_id=int(org_id))

        org_ids = self._parse_int_list(self.request.query_params.get('org_ids'))
        if org_ids:
            queryset = queryset.filter(entry__organization_id__in=org_ids)

        log_types = self._parse_csv_list(self.request.query_params.get('log_type'))
        if log_types:
            queryset = queryset.filter(log_type__in=log_types)

        actions = self._parse_csv_list(self.request.query_params.get('action'))
        if actions:
            queryset = queryset.filter(action__in=actions)

        resource_type = self.request.query_params.get('resource_type')
        if resource_type:
            queryset = queryset.filter(resource_type=resource_type)

        resource_id = self.request.query_params.get('resource_id')
        if resource_id:
            queryset = queryset.filter(resource_id=str(resource_id))

        method = self.request.query_params.get('method')
        if method:
            queryset = queryset.filter(method__iexact=method)

        actor = self.request.query_params.get('actor')
        if actor:
            queryset = queryset.filter(
                Q(actor__username__icontains=actor) |
                Q(actor__first_name__icontains=actor)
            )

        status_value = self.request.query_params.get('status')
        if status_value:
            queryset = queryset.filter(
                Q(from_status__iexact=status_value) |
                Q(to_status__iexact=status_value)
            )

        status_code = self.request.query_params.get('status_code')
        if str(status_code).isdigit():
            queryset = queryset.filter(status_code=int(status_code))

        from_date = self.request.query_params.get('from_date') or self.request.query_params.get('from')
        if from_date:
            queryset = queryset.filter(created_at__date__gte=from_date)

        to_date = self.request.query_params.get('to_date') or self.request.query_params.get('to')
        if to_date:
            queryset = queryset.filter(created_at__date__lte=to_date)

        keyword = self.request.query_params.get('q')
        if keyword:
            queryset = queryset.filter(
                Q(reason__icontains=keyword) |
                Q(path__icontains=keyword) |
                Q(resource_type__icontains=keyword) |
                Q(resource_id__icontains=keyword) |
                Q(actor__username__icontains=keyword) |
                Q(actor__first_name__icontains=keyword)
            )

        return queryset

    def paginate_queryset(self, queryset):
        if self.request.query_params.get('entry_ids'):
            return None
        return super().paginate_queryset(queryset)

class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.select_related('user').order_by('-created_at')
    serializer_class = NotificationSerializer

    def get_queryset(self):
        queryset = Notification.objects.select_related('user').order_by('-created_at')
        if _is_admin(self.request):
            return queryset
        return queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        if _is_admin(self.request):
            serializer.save()
            return
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        if _is_admin(self.request):
            serializer.save()
            return
        if serializer.instance.user_id != self.request.user.id:
            raise PermissionDenied('No permission for this notification.')
        serializer.save()

    def perform_destroy(self, instance):
        if _is_admin(self.request) or instance.user_id == self.request.user.id:
            instance.delete()
            return
        raise PermissionDenied('No permission for this notification.')

class DashboardSummaryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        round_no = request.query_params.get('round', 0)
        if not year:
            return Response({'error': 'year parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        entries = BudgetEntry.objects.filter(year=year, supplemental_round=round_no).select_related('subject', 'organization')
        entries = _scope_queryset_by_org(entries, request, org_field='organization_id')
        if not entries.exists():
            return Response({
                'year': year, 'round': round_no, 'total_income': 0, 'total_expense': 0,
                'status_counts': {}, 'org_progress': []
            })
        
        # Detail aggregation
        summary_query = BudgetDetail.objects.filter(entry__year=year, entry__supplemental_round=round_no)
        summary_query = _scope_queryset_by_org(summary_query, request, org_field='entry__organization_id')
        
        income_normal = summary_query.filter(entry__subject__subject_type='income', is_rate=False).aggregate(
            total=Sum(F('price') * F('qty') * F('freq'))
        )['total'] or 0
        income_rate = sum(d.total_price for d in summary_query.filter(entry__subject__subject_type='income', is_rate=True))
        total_income = income_normal + income_rate
        
        expense_normal = summary_query.filter(entry__subject__subject_type='expense', is_rate=False).aggregate(
            total=Sum(F('price') * F('qty') * F('freq'))
        )['total'] or 0
        expense_rate = sum(d.total_price for d in summary_query.filter(entry__subject__subject_type='expense', is_rate=True))
        total_expense = expense_normal + expense_rate

        from collections import Counter
        status_counts = dict(Counter(entries.values_list('status', flat=True)))

        # Organization-wise progress (for departments)
        org_stats = []
        depts = Organization.objects.filter(org_type='dept')
        allowed_org_ids = _scope_org_ids_for_user(request)
        if allowed_org_ids is not None:
            allowed_dept_ids = set()
            for org in Organization.objects.filter(id__in=allowed_org_ids).only('id', 'parent_id', 'org_type'):
                if _is_team_organization(org):
                    if org.parent_id:
                        allowed_dept_ids.add(org.parent_id)
                else:
                    allowed_dept_ids.add(org.id)
            depts = depts.filter(id__in=allowed_dept_ids)
        for org in depts:
            # Include child teams in the count
            team_ids = list(Organization.objects.filter(parent=org).values_list('id', flat=True))
            scope_ids = [org.id] + team_ids
            
            org_entries = entries.filter(organization_id__in=scope_ids)
            if not org_entries.exists(): continue
            
            e_count = org_entries.count()
            f_count = org_entries.filter(status='FINALIZED').count()
            org_stats.append({
                'id': org.id,
                'name': org.name,
                'total': e_count,
                'finalized': f_count,
                'ratio': round((f_count / e_count) * 100, 1) if e_count > 0 else 0
            })

        return Response({
            'year': year,
            'round': round_no,
            'total_income': total_income,
            'total_expense': total_expense,
            'status_counts': status_counts,
            'org_progress': org_stats
        })

class BudgetBulkUpsertView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        """
        Expects: { "year": 2026, "round": 0, "entries": [ { "subject_code": "...", "org_code": "...", "details": [...] } ] }
        """
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='Bulk upsert is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        year = request.data.get('year')
        round_no = request.data.get('round', 0)
        entry_list = request.data.get('entries', [])
        
        if not year or not entry_list:
            return Response({'error': 'year and entries list required'}, status=status.HTTP_400_BAD_REQUEST)

        # Pre-cache orgs and subjects for speed
        org_map = {o.code: o for o in Organization.objects.all()}
        sub_map = {s.code: s for s in BudgetSubject.objects.filter(level=4)}
        allowed_org_ids = _scope_org_ids_for_user(request)

        created_count = 0
        updated_count = 0
        
        for item in entry_list:
            sc = item.get('subject_code')
            oc = item.get('org_code')
            if sc not in sub_map or oc not in org_map:
                continue
            if allowed_org_ids is not None and org_map[oc].id not in allowed_org_ids:
                continue

            entry, created = BudgetEntry.objects.get_or_create(
                year=year,
                supplemental_round=round_no,
                subject=sub_map[sc],
                organization=org_map[oc],
                defaults={'status': 'DRAFT'}
            )
            if created: created_count += 1
            else: updated_count += 1

            # If details provided, replace them
            details_data = item.get('details')
            if isinstance(details_data, list):
                entry.details.all().delete()
                for d in details_data:
                    # 유효성 검사: price/qty/freq는 숫자, name은 문자열
                    try:
                        price = int(d.get('price', 0))
                        qty = float(d.get('qty', 1))
                        freq = int(d.get('freq', 1))
                    except (TypeError, ValueError):
                        continue  # 잘못된 숫자 형식은 건너뜀
                    name = str(d.get('name', 'Bulk Item'))[:200]  # 최대 200자
                    if freq < 1:
                        freq = 1
                    BudgetDetail.objects.create(
                        entry=entry,
                        name=name,
                        price=price,
                        qty=qty,
                        freq=freq,
                        unit=str(d.get('unit', '식'))[:20],
                        source=str(d.get('source', 'SELF'))[:50],
                    )
        
        return Response({
            'status': 'ok',
            'created': created_count,
            'updated': updated_count
        })



class ERPNextViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def _require_erp_role(self, request):
        denied = _require_roles_response(
            request,
            {'MANAGER', 'ADMIN'},
            message='ERPNext integration is allowed only for MANAGER or ADMIN.',
        )
        if denied is not None:
            return denied
        return None

    def _parse_bool(self, value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.strip().lower() in ('1', 'true', 'yes', 'y', 'on')
        return default

    def _config_or_error(self):
        if not settings.ERPNEXT_BASE_URL:
            return Response({'error': 'ERPNEXT_BASE_URL is not configured'}, status=status.HTTP_400_BAD_REQUEST)
        if not settings.ERPNEXT_API_KEY or not settings.ERPNEXT_API_SECRET:
            return Response({'error': 'ERPNEXT_API_KEY/ERPNEXT_API_SECRET are not configured'}, status=status.HTTP_400_BAD_REQUEST)
        if not settings.ERPNEXT_COMPANY:
            return Response({'error': 'ERPNEXT_COMPANY is not configured'}, status=status.HTTP_400_BAD_REQUEST)
        return None

    @action(detail=False, methods=['get'])
    def me(self, request):
        denied = self._require_erp_role(request)
        if denied is not None:
            return denied
        config_error = self._config_or_error()
        if config_error:
            return config_error
        try:
            client = get_erpnext_client()
            return Response(client.get_logged_user())
        except ERPNextError as exc:
            return Response({'error': str(exc), 'details': exc.payload}, status=status.HTTP_502_BAD_GATEWAY)

    @action(detail=False, methods=['post'], url_path='budgets/sync')
    def sync_budgets(self, request):
        denied = self._require_erp_role(request)
        if denied is not None:
            return denied
        config_error = self._config_or_error()
        if config_error:
            return config_error

        year = request.data.get('year') or request.query_params.get('year')
        if not year:
            return Response({'error': 'year is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
        except ValueError:
            return Response({'error': 'year must be an integer'}, status=status.HTTP_400_BAD_REQUEST)

        company = request.data.get('company') or settings.ERPNEXT_COMPANY
        fiscal_year = request.data.get('fiscal_year') or settings.ERPNEXT_FISCAL_YEAR or str(year)
        budget_against = request.data.get('budget_against') or settings.ERPNEXT_BUDGET_AGAINST
        dry_run = self._parse_bool(request.data.get('dry_run', False), default=False)
        update_existing = self._parse_bool(request.data.get('update_existing', True), default=True)

        if budget_against != 'Cost Center':
            return Response({'error': 'Only Cost Center budgets are supported in this endpoint'}, status=status.HTTP_400_BAD_REQUEST)

        entries = (
            BudgetEntry.objects
            .filter(year=year)
            .select_related('organization', 'subject')
            .prefetch_related('details')
        )
        entries = _scope_queryset_by_org(entries, request, org_field='organization_id')

        skipped = {
            'organizations_missing_cost_center': [],
            'subjects_missing_account': [],
        }
        grouped = {}

        for entry in entries:
            org = entry.organization
            subject = entry.subject

            if not org.erpnext_cost_center:
                if org.code not in skipped['organizations_missing_cost_center']:
                    skipped['organizations_missing_cost_center'].append(org.code)
                continue

            if not subject.erpnext_account:
                if subject.code not in skipped['subjects_missing_account']:
                    skipped['subjects_missing_account'].append(subject.code)
                continue

            grouped.setdefault(org.erpnext_cost_center, {})
            grouped[org.erpnext_cost_center].setdefault(subject.erpnext_account, 0)
            grouped[org.erpnext_cost_center][subject.erpnext_account] += entry.total_amount

        payloads = []
        for cost_center, accounts in grouped.items():
            account_rows = [
                {'account': account, 'budget_amount': int(amount)}
                for account, amount in accounts.items()
            ]
            payloads.append({
                'company': company,
                'fiscal_year': fiscal_year,
                'budget_against': budget_against,
                'cost_center': cost_center,
                'accounts': account_rows,
            })

        if dry_run:
            return Response({'dry_run': True, 'payloads': payloads, 'skipped': skipped})

        client = get_erpnext_client()
        results = []

        for payload in payloads:
            try:
                existing_name = None
                if update_existing:
                    filters = [
                        ['Budget', 'company', '=', company],
                        ['Budget', 'fiscal_year', '=', fiscal_year],
                        ['Budget', 'budget_against', '=', budget_against],
                        ['Budget', 'cost_center', '=', payload['cost_center']],
                    ]
                    existing = client.list_resource('Budget', filters=filters, fields=['name'], limit=1)
                    data = existing.get('data') or []
                    if data:
                        existing_name = data[0].get('name')

                if existing_name:
                    response = client.update_resource('Budget', existing_name, payload)
                    results.append({'cost_center': payload['cost_center'], 'action': 'updated', 'name': existing_name})
                else:
                    response = client.create_resource('Budget', payload)
                    name = response.get('data', {}).get('name')
                    results.append({'cost_center': payload['cost_center'], 'action': 'created', 'name': name})
            except ERPNextError as exc:
                results.append({'cost_center': payload['cost_center'], 'action': 'error', 'error': str(exc)})

        return Response({'year': year, 'fiscal_year': fiscal_year, 'results': results, 'skipped': skipped})

    @action(detail=False, methods=['post'], url_path='closing-voucher')
    def create_closing_voucher(self, request):
        denied = self._require_erp_role(request)
        if denied is not None:
            return denied
        config_error = self._config_or_error()
        if config_error:
            return config_error

        data = request.data.copy()
        if 'company' not in data:
            data['company'] = settings.ERPNEXT_COMPANY
        if 'fiscal_year' not in data and settings.ERPNEXT_FISCAL_YEAR:
            data['fiscal_year'] = settings.ERPNEXT_FISCAL_YEAR

        try:
            client = get_erpnext_client()
            response = client.create_resource('Period Closing Voucher', data)
            return Response(response)
        except ERPNextError as exc:
            return Response({'error': str(exc), 'details': exc.payload}, status=status.HTTP_502_BAD_GATEWAY)


class SubmissionCommentViewSet(viewsets.ModelViewSet):
    """
    Notice board CRUD.
    - Create: requires logged-in user
    - Update/Delete: author only (soft delete)
    """
    serializer_class = SubmissionCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = SubmissionComment.objects.filter(is_deleted=False).select_related(
            'author', 'entry', 'org', 'subject', 'version', 'parent'
        ).prefetch_related('replies__author')
        allowed_org_ids = _scope_org_ids_for_user(self.request)
        if allowed_org_ids is not None:
            if not allowed_org_ids:
                return qs.none()
            qs = qs.filter(
                Q(org_id__in=allowed_org_ids)
                | (Q(org__isnull=True) & Q(entry__organization_id__in=allowed_org_ids))
            )

        version_id = self.request.query_params.get('version')
        if version_id:
            qs = qs.filter(version_id=version_id)

        entry_id = self.request.query_params.get('entry')
        if entry_id:
            qs = qs.filter(entry_id=entry_id)

        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        org_id = self.request.query_params.get('org')
        if org_id:
            qs = qs.filter(
                Q(org_id=org_id)
                | (Q(org__isnull=True) & Q(entry__organization_id=org_id))
            )

        project_id = self.request.query_params.get('entrusted_project')
        if project_id:
            qs = qs.filter(
                Q(entrusted_project_id=project_id)
                | (Q(entrusted_project__isnull=True) & Q(entry__entrusted_project_id=project_id))
            )
        elif not entry_id and subject_id:
            qs = qs.filter(entrusted_project__isnull=True)

        # Return top-level notices only (replies are nested in serializer)
        top_level = self.request.query_params.get('top_level')
        if top_level in ('1', 'true', 'True'):
            qs = qs.filter(parent__isnull=True)

        return qs

    def perform_create(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to create comments.')
        entry = serializer.validated_data.get('entry')
        org = serializer.validated_data.get('org')
        entrusted_project = serializer.validated_data.get('entrusted_project')

        # Entry-level comment should inherit scope when client omits org/project.
        if entry is not None:
            if org is None:
                org = entry.organization
            if entrusted_project is None:
                entrusted_project = entry.entrusted_project
        scope_org_id = getattr(org, 'id', None) if org is not None else None
        if scope_org_id is None and entry is not None:
            scope_org_id = getattr(entry, 'organization_id', None)
        if not _org_in_scope(self.request, scope_org_id):
            raise PermissionDenied('No permission for this organization.')

        serializer.save(
            author=self.request.user,
            org=org,
            entrusted_project=entrusted_project,
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.author_id != request.user.id:
            return Response({'detail': 'Only the author can edit this post.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.author_id != request.user.id:
            return Response({'detail': 'Only the author can delete this post.'}, status=status.HTTP_403_FORBIDDEN)
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class SupportingDocumentViewSet(viewsets.ModelViewSet):
    """
    Supporting document attachment CRUD.
    """
    serializer_class = SupportingDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = SupportingDocument.objects.select_related(
            'author', 'org', 'subject', 'version', 'entrusted_project'
        ).order_by('-created_at')
        qs = _scope_queryset_by_org(qs, self.request, org_field='org_id')
        
        version_id = self.request.query_params.get('version')
        if version_id:
            qs = qs.filter(version_id=version_id)

        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        org_id = self.request.query_params.get('org')
        if org_id:
            qs = qs.filter(org_id=org_id)

        return qs

    def perform_create(self, serializer):
        if not _can_write_budget_data(self.request):
            raise PermissionDenied('No permission to upload supporting documents.')
        org = serializer.validated_data.get('org')
        if not _org_in_scope(self.request, getattr(org, 'id', None)):
            raise PermissionDenied('No permission for this organization.')
        file_obj = self.request.data.get('file')
        file_size = file_obj.size if file_obj else 0
        filename = file_obj.name if file_obj else 'unknown'
        serializer.save(
            author=self.request.user,
            file_size=file_size,
            filename=filename
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        role = _user_role(request)
        if instance.author_id != request.user.id and role not in ('MANAGER', 'ADMIN'):
            return Response({'detail': 'No permission to edit.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        role = _user_role(request)
        if instance.author_id != request.user.id and role not in ('MANAGER', 'ADMIN'):
            return Response({'detail': 'No permission to edit.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        profile = getattr(request.user, 'profile', None)  # delete allowed only for author or MANAGER/ADMIN
        role = _normalize_role(getattr(profile, 'role', None)) if profile else 'STAFF'
        if instance.author_id != request.user.id and role not in ('MANAGER', 'ADMIN'):
            return Response({'detail': 'No permission to delete.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)
